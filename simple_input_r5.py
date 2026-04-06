import os
import functools
from typing import List, Tuple
import pandas as pd
import math


# master_menseki = DedicatedAreaTable

# コンクリート, 断熱材の熱伝導率λ
concrete_conductivity = 1.6
insulation_conductivity = 0.03

def to_json(input_xlsx_filepath, region):
    import openpyxl


    def convert_to_input_json(
        input_xlsx_filepath: str,
    ):
        book = openpyxl.load_workbook(input_xlsx_filepath)

        sheet_common = book['common']
        sheet_building = book['building']
        sheet_rooms = book['rooms']
        sheet_external_general_parts = book['external_general_parts']
        sheet_external_opaque_parts = book['external_opaque_parts']
        sheet_external_transparent_parts = book['external_transparent_parts']
        sheet_internals = book['internals']
        sheet_grounds = book['grounds']
        sheet_layers = book['layers']
        
        n_rooms = count_number_in_id_row(sheet=sheet_rooms)
        n_external_general_parts = count_number_in_id_row(sheet=sheet_external_general_parts)
        n_external_opaque_parts = count_number_in_id_row(sheet=sheet_external_opaque_parts)
        n_external_transparent_parts = count_number_in_id_row(sheet=sheet_external_transparent_parts)
        n_internals = count_number_in_id_row(sheet=sheet_internals)
        n_grounds = count_number_in_id_row(sheet=sheet_grounds)
        n_layers = count_number_in_id_row(sheet=sheet_layers)
        
        common = {
            'ac_method': sheet_common.cell(column=2, row=2).value,
            'weather': {
                'method': 'ees',
                'region': region,
            }
        }
        
        building = {
            "infiltration": {
                "method": "balance_residential",
                "c_value_estimate": "specify",
                "story": int(sheet_building.cell(column=2, row=2).value),
                "c_value": float(sheet_building.cell(column=3, row=2).value),
                "inside_pressure": sheet_building.cell(column=4, row=2).value
            }
        }
        
        rooms = [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "floor_area": float(row[4].value),
                "volume": float(row[5].value),
                "ventilation": {
                    "natural": float(row[6].value)
                },
                "furniture": {
                    "input_method": "default"
                },
                "schedule": {
                    "name": row[7].value
                }
            } for row in sheet_rooms.iter_rows(min_row=2, max_row=n_rooms+1)
        ]

        layers_master = [
            {
                "name": row[1].value,
                "layers": make_dictionary_of_layer(row)[0],
                "reversed_layers": make_dictionary_of_layer(row)[1]
            } for row in sheet_layers.iter_rows(min_row=2, max_row=n_layers+1)
        ]
        
        external_general_parts =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "external_general_part",
                "area": float(row[5].value),
                "h_c": get_h_c(direction=row[8].value),
                "is_solar_absorbed_inside": bool(row[6].value),
                "is_floor": bool(row[6].value),
                "layers": get_layers(layers_master, layer_name=row[7].value),
                "solar_shading_part": {"existence": False},
                "is_sun_striked_outside": row[9].value == 1.0,  # fix by kitamura
                "direction": row[8].value,
                "outside_emissivity": 0.9,
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[8].value, temp_dif_coef=float(row[9].value)),
                "outside_solar_absorption": 0.8,
                "temp_dif_coef": float(row[9].value)
            } for row in sheet_external_general_parts.iter_rows(min_row=2, max_row=n_external_general_parts+1)
            if float(row[5].value) > 0.0
        ]

        external_opaque_parts =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "external_opaque_part",
                "area": float(row[5].value),
                "h_c": get_h_c(direction=row[7].value),
                "is_solar_absorbed_inside": False,
                "is_floor": False,
                "solar_shading_part": {"existence": False},
                "is_sun_striked_outside": True,
                "direction": row[7].value,
                "outside_emissivity": 0.9,
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[7].value, temp_dif_coef=1.0),
                "u_value": float(row[6].value),
                "inside_heat_transfer_resistance": 0.11,
                "outside_solar_absorption": 0.8,
                "temp_dif_coef": 1.0
            } for row in sheet_external_opaque_parts.iter_rows(min_row=2, max_row=n_external_opaque_parts+1)
            if float(row[5].value) > 0.0
        ]

        external_transparent_parts =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "external_transparent_part",
                "area": float(row[5].value),
                "h_c": get_h_c(direction=row[10].value),
                "is_solar_absorbed_inside": False,
                "is_floor": False,
                "solar_shading_part": get_solar_shading(exist=bool(row[11].value), depth=row[12].value, d_h=row[13].value, d_e=row[14].value),
                "is_sun_striked_outside": True,
                "direction": row[10].value,
                "outside_emissivity": 0.9,
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[10].value, temp_dif_coef=1.0),
                "u_value": float(row[6].value),
                "inside_heat_transfer_resistance": 0.11,
                "eta_value": float(row[7].value),
                "incident_angle_characteristics": row[8].value,
                "glass_area_ratio": float(row[9].value),
                "temp_dif_coef": 1.0
            } for row in sheet_external_transparent_parts.iter_rows(min_row=2, max_row=n_external_transparent_parts+1)
            if float(row[5].value) > 0.0
        ]
        
        internals_2d =  [
            [
                {
                    "id": row[1].value,
                    "name": row[3].value,
                    "sub_name": row[5].value or '',
                    "connected_room_id": int(row[7].value),
                    "boundary_type": "internal",
                    "area": float(row[9].value),
                    "h_c": get_h_c(direction=row[11].value)[0],
                    "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[0],
                    "is_floor": get_is_floor(direction=row[11].value)[0],
                    "layers": get_layers(layers_master, layer_name=row[10].value, is_reverse=False),
                    "solar_shading_part": {"existence": False},
                    "rear_surface_boundary_id": row[2].value
                },
                {
                    "id": row[2].value,
                    "name": row[4].value,
                    "sub_name": row[6].value or '',
                    "connected_room_id": int(row[8].value),
                    "boundary_type": "internal",
                    "area": float(row[9].value),
                    "h_c": get_h_c(direction=row[11].value)[1],
                    "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[1],
                    "is_floor": get_is_floor(direction=row[11].value)[1],
                    "layers": get_layers(layers_master, layer_name=row[10].value, is_reverse=True),
                    "solar_shading_part": {"existence": False},
                    "rear_surface_boundary_id": row[1].value
                }
            ] for row in sheet_internals.iter_rows(min_row=2, max_row=n_internals+1)
            if float(row[9].value) > 0.0
        ]
        # flatten
        internals = sum(internals_2d, [])

        grounds =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "ground",
                "area": float(row[5].value),
                "is_solar_absorbed_inside": bool(row[7].value),
                "is_floor": True,
                "h_c": get_h_c(direction='bottom'),
                "layers": get_layers(layers_master, layer_name=row[6].value),
                "solar_shading_part": {"existence": False},
            } for row in sheet_grounds.iter_rows(min_row=2, max_row=n_grounds+1)
            if float(row[5].value) > 0.0
        ]

        # 各境界id及び裏面境界idを修正
        # ※面積0の境界がスキップされるため、修正前の時点では境界のidと実際の配列のインデックスにずれが発生してしまっている。
        boundaries = external_general_parts + external_opaque_parts + external_transparent_parts + internals + grounds
        boundaries_id_mapper = {_bd['id']: _new_id for _new_id, _bd in enumerate(boundaries)}
        for bd in boundaries:
            bd['id'] = boundaries_id_mapper[bd['id']]
            if bd['boundary_type'] == 'internal':
                bd['rear_surface_boundary_id'] = boundaries_id_mapper[bd['rear_surface_boundary_id']]

        ventilation_rate = 0.5
        V_MR, V_OR, V_NR = [_["volume"] for _ in rooms[:3]]
        v_vent_MR = ventilation_rate * (V_MR + V_NR * V_MR / (V_MR + V_OR))
        v_vent_OR = ventilation_rate * (V_OR + V_NR * V_OR / (V_MR + V_OR))
        mechanical_ventilations = [
            {
                "id": 0,
                "root_type": "type3",
                "volume": v_vent_MR,
                "root": [
                    0,
                    2
                ]
            },
            {
                "id": 1,
                "root_type": "type3",
                "volume": v_vent_OR,
                "root": [
                    1,
                    2
                ]
            }
        ]
        
        equipment_c_MR, equipment_h_MR = create_equipments(id=0, space_id=0, a_floor_is=rooms[0]['floor_area'])
        equipment_c_OR, equipment_h_OR = create_equipments(id=1, space_id=1, a_floor_is=rooms[1]['floor_area'])
        equipments = {
            "heating_equipments": [equipment_h_MR, equipment_h_OR],
            "cooling_equipments": [equipment_c_MR, equipment_c_OR]
        }

        return {
            "common":common,
            "building": building,
            "rooms": rooms,
            "boundaries": boundaries,
            "mechanical_ventilations": mechanical_ventilations,
            "equipments": equipments
        }


    def count_number_in_id_row(sheet):
        id_all = [row[1].value for row in sheet.rows][1:]
        return len(id_all) - (id_all).count(None)


    def make_dictionary_of_layer(row):
        n = int(row[2].value)
        # NOTE: 熱抵抗が 0 未満の layer は生成しない（heat_load_calc による計算時にエラーになるため）
        layer = [
            {
                "name": row[3+3*i].value,
                "thermal_resistance": float(row[4+3*i].value),
                "thermal_capacity": float(row[5+3*i].value)
            } for i in range(n) if float(row[4+3*i].value) > 0.0
        ]
        # Tuple(layer_list, reversed_layer_list)
        return layer, layer[::-1]


    def get_layers(layers_master, layer_name, is_reverse=False):
        # use variable 'layers_master' as global variable
        layers = list(filter(lambda d: d['name'] == layer_name , layers_master))
        if len(layers) > 1:
            raise Exception("Match over one layer.")
        if len(layers) == 0:
            raise Exception("Can't find the layer")
        if is_reverse:
            return layers[0]['reversed_layers']
        else:
            return layers[0]['layers']
        
        
    def get_h_c(direction):
        if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
            return 2.5
        elif direction == 'bottom':
            return 0.7
        elif direction == 'top':
            return 5.0
        elif direction == 'horizontal':
            return (2.5, 2.5)
        elif direction == 'upward':
            return (5.0, 0.7)
        elif direction == 'downward':
            return (0.7, 5.0)
        else:
            raise ValueError(direction)
        
        
    def get_outside_heat_transfer_resistance(direction, temp_dif_coef):
        is_parting = (temp_dif_coef != 1.0)

        if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
            return 0.04 if not is_parting else 0.11
        elif direction == 'bottom':
            return 0.15
        elif direction == 'top':
            return 0.04 if not is_parting else 0.09
        else:
            raise Exception()
        
        
    def get_solar_shading(exist: bool, depth=None, d_h=None, d_e=None):
        if exist:
            return {
                "existence": True,
                "input_method": "simple",
                "depth": float(depth),
                "d_h": float(d_h),
                "d_e": float(d_e)
            }
        else:
            return {
                "existence": False
            }


    def get_is_floor(direction):
        if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se', 'top']:
            return False
        elif direction == 'bottom':
            return True
        elif direction == 'horizontal':
            return (False, False)
        elif direction == 'upward':
            return (False, True)
        elif direction == 'downward':
            return (True, False)
        else:
            raise Exception()


    def create_equipments(id, space_id, a_floor_is):
        q_rtd_c = 190.5 * a_floor_is + 45.6
        q_rtd_h = 1.2090 * q_rtd_c - 85.1

        q_max_c = max(0.8462 * q_rtd_c + 1205.9, q_rtd_c)
        q_max_h = max(1.7597 * q_max_c - 413.7, q_rtd_h)
        
        q_min_c = 500
        q_min_h = 500

        v_max_c = 11.076 * (q_rtd_c / 1000.0) ** 0.3432
        v_max_h = 11.076 * (q_rtd_h / 1000.0) ** 0.3432

        v_min_c = v_max_c * 0.55
        v_min_h = v_max_h * 0.55
        
        bf_c = 0.2
        bf_h = 0.2

        cooling_equipment = {
            "id": id,
            "name": f"cooling_equipment no.{id}",
            "equipment_type": "rac",
            "property": {
                "space_id": space_id,
                "q_min": q_min_c,
                "q_max": q_max_c,
                "v_min": v_min_c,
                "v_max": v_max_c,
                "bf": bf_c
            }
        }
        
        heating_equipment = {
            "id": id,
            "name": f"heating_equipment no.{id}",
            "equipment_type": "rac",
            "property": {
                "space_id": space_id,
                "q_min": q_min_h,
                "q_max": q_max_h,
                "v_min": v_min_h,
                "v_max": v_max_h,
                "bf": bf_h
            }
        }
        
        return cooling_equipment, heating_equipment
    
    return convert_to_input_json(input_xlsx_filepath)





def get_wall_transfer_rate(R_fix, insulation_thickness, H, Rs):

    # 熱抵抗
    total_registance = sum([
        Rs,
        R_fix,
        1.0 / insulation_conductivity * insulation_thickness,
    ])

    return 1.0 / total_registance * H


def get_insulation_registance(wall_transfer_rate, R_fix, tempdiff_coeff, surface_registance):

    if wall_transfer_rate == 0.0 or tempdiff_coeff == 0.0:
        return 0.0

    # 壁体の熱抵抗
    total_registance = 1.0 / (wall_transfer_rate / tempdiff_coeff)

    # 断熱材の熱抵抗 = 壁体の熱抵抗からコンクリートと表面の熱抵抗を除いたもの
    insulation_registance = total_registance - surface_registance - R_fix

    return insulation_registance

# 3.3.10.6 推定住戸の開口部面積の合計
# 開口部面積の合計 = 外気に接する外皮の面積 * 開口部比率
def get_total_open_area(A_env_ex, r_env_op):
    return A_env_ex * r_env_op

# 3.3.10.7 推定住戸の外気に接する外皮に占める開口部面積の割合
def get_open_rate(eta_ac):
    return 1.0 / (1.0 + math.exp(-0.12852127 * (eta_ac - 13.35408535)))



def calc_eta_win(m_tran, A_env_win, neu_c, neu_h, DD_C, DD_H):

    # 3.3.9.1 推定住戸の窓の日射熱取得率(負荷計算への入力)
    # 窓に割り当てられた日射熱取得量を方位係数を考慮しつつ、面積で割って日射熱取得率にする


    def get_proportion_ratio(A_env_win_dir: float, neu_dir_c: float, neu_dir_h: float):

        # 3.3.9.9 取得日射熱補正係数
        solar_heat_acquisition_coeff_c = 0.93
        solar_heat_acquisition_coeff_h = 0.51

        return (A_env_win_dir * neu_dir_c * solar_heat_acquisition_coeff_c * DD_C \
                + A_env_win_dir * neu_dir_h * solar_heat_acquisition_coeff_h * DD_H) / (DD_C + DD_H)

    # 窓の日射熱取得率 [(W/㎡)/(W/㎡)]
    # NOTE: 以前 neu_c, neu_h の式の実装誤りがあったようだ。方位係数が正しく選ばれていなかったように思われる。
    A_env_win_south, A_env_win_east, A_env_win_north, A_env_win_west = A_env_win
    proportion_ratio_n = get_proportion_ratio(A_env_win_north, neu_c[1], neu_h[1])
    proportion_ratio_e = get_proportion_ratio(A_env_win_east, neu_c[3], neu_h[3])
    proportion_ratio_s = get_proportion_ratio(A_env_win_south, neu_c[5], neu_h[5])
    proportion_ratio_w = get_proportion_ratio(A_env_win_west, neu_c[7], neu_h[7])

    # NOTE: 窓のη値が0.0以下の場合heat_load_calcでは計算不可能のため、最小値を1e-8とする。
    eta_win = max(1e-8, m_tran / (proportion_ratio_n + proportion_ratio_e + proportion_ratio_s + proportion_ratio_w))

    return eta_win



def estimate(region, total_floor_area, main_floor_area, other_floor_area, A_env, ua, eta_ah, eta_ac, tatekata, structure, xlsx_path, has_vertical_internal="有"):
        
    ### ざっくり入力された面積から補正率を計算する ###

    ## 床面積の補正率

    # 入力された床面積の整理

    # 居室の面積
    A_MR = main_floor_area
    A_OR = other_floor_area
    A_NO = total_floor_area - main_floor_area - other_floor_area

    # 主たる居室の床面積の入力値との比率
    # その他の居室の床面積の入力値との比率
    # 非居室の床面積の入力値との比率
    # floor_area_rate = TriValue(
    #     A_MR / A_env_horz_MR,
    #     A_OR / A_env_horz_OR,
    #     A_NO / A_env_horz_NO,
    # )

    # -------------------------
    # ---- 3.3.10 面積計算 -----
    # -------------------------

    # 3.3.10.51	参照住戸の面積

    A_MR_ref, A_OR_ref, A_NO_ref = get_floor_area_ref(tatekata)

    # 参照住戸の面積
    area_table_ref = get_area_table_ref(tatekata)
    A_env_top_MR_ref, A_env_top_OR_ref, A_env_top_NO_ref, A_env_top_UF_ref = area_table_ref[0]
    A_env_north_MR_ref, A_env_north_OR_ref, A_env_north_NO_ref, A_env_north_UF_ref = area_table_ref[1]
    A_env_east_MR_ref, A_env_east_OR_ref, A_env_east_NO_ref, A_env_east_UF_ref = area_table_ref[2]
    A_env_south_MR_ref, A_env_south_OR_ref, A_env_south_NO_ref, A_env_south_UF_ref = area_table_ref[3]
    A_env_west_MR_ref, A_env_west_OR_ref, A_env_west_NO_ref, A_env_west_UF_ref = area_table_ref[4]
    A_env_bottom_MR_ref, A_env_bottom_OR_ref, A_env_bottom_NO_ref, A_env_bottom_UF_ref = area_table_ref[5]
    A_env_win_north_MR_ref, A_env_win_north_OR_ref, A_env_win_north_NO_ref = area_table_ref[6]
    A_env_win_east_MR_ref, A_env_win_east_OR_ref, A_env_win_east_NO_ref = area_table_ref[7]
    A_env_win_south_MR_ref, A_env_win_south_OR_ref, A_env_win_south_NO_ref = area_table_ref[8]
    A_env_win_west_MR_ref, A_env_win_west_OR_ref, A_env_win_west_NO_ref = area_table_ref[9]
    A_env_door_north_MR_ref, A_env_door_north_OR_ref, A_env_door_north_NO_ref = area_table_ref[10]
    A_env_door_west_MR_ref, A_env_door_west_OR_ref, A_env_door_west_NO_ref = area_table_ref[11]

    # 断熱方法によって値を読み変える
    # 床断熱の場合: 床下空間の北・東・南・西・下面の外皮面積を0とする
    # 基礎断熱の場合: 主たる居室・その他の居室・非居室の下面の外皮面積を0とする
    if tatekata == "戸建住宅":
        if structure == "基礎断熱":
            A_env_bottom_MR_ref = 0.0
            A_env_bottom_OR_ref = 0.0
            A_env_bottom_NO_ref = 0.0
        elif structure in ["床断熱", "床下断熱"]:
            A_env_bottom_UF_ref = 0.0
            A_env_south_UF_ref = 0.0
            A_env_east_UF_ref = 0.0
            A_env_north_UF_ref = 0.0
            A_env_west_UF_ref = 0.0
        else:
            raise ValueError(structure)

    # 参照住戸の間仕切りの面積
    partition_table_ref = get_partition_table_ref(tatekata)
    A_part_MR_OR_ref, A_part_MR_NO_ref, A_part_OR_NO_ref = partition_table_ref

    # 参照住戸の内壁床の面積
    partition_bottom_table_ref = get_partition_bottom_table_ref(tatekata)
    A_part_bottom_MR_MR_ref, A_part_bottom_MR_OR_ref, A_part_bottom_MR_NO_ref, A_part_bottom_MR_UF_ref, A_part_bottom_OR_MR_ref, A_part_bottom_OR_OR_ref, A_part_bottom_OR_NO_ref, A_part_bottom_OR_UF_ref, A_part_bottom_NO_MR_ref, A_part_bottom_NO_OR_ref, A_part_bottom_NO_NO_ref, A_part_bottom_NO_UF_ref = partition_bottom_table_ref

    # 戸建住宅(基礎断熱)以外の場合、床下空間に接する内壁床の面積を0とする。
    if not (tatekata == "戸建住宅" and structure == "基礎断熱"):
        A_part_bottom_MR_UF_ref = 0.0
        A_part_bottom_OR_UF_ref = 0.0
        A_part_bottom_NO_UF_ref = 0.0

    # 3.3.10.50	参照住戸の空間ごとの垂直の外皮の面積
    A_env_vert_MR_ref = A_env_south_MR_ref + A_env_east_MR_ref + A_env_north_MR_ref + A_env_west_MR_ref
    A_env_vert_OR_ref = A_env_south_OR_ref + A_env_east_OR_ref + A_env_north_OR_ref + A_env_west_OR_ref
    A_env_vert_NO_ref = A_env_south_NO_ref + A_env_east_NO_ref + A_env_north_NO_ref + A_env_west_NO_ref
    A_env_vert_UF_ref = A_env_south_UF_ref + A_env_east_UF_ref + A_env_north_UF_ref + A_env_west_UF_ref

    # 3.3.10.49	参照住戸の空間ごとの水平の外皮の面積
    A_env_horz_MR_ref = A_env_top_MR_ref + A_env_bottom_MR_ref
    A_env_horz_OR_ref = A_env_top_OR_ref + A_env_bottom_OR_ref
    A_env_horz_NO_ref = A_env_top_NO_ref + A_env_bottom_NO_ref
    A_env_horz_UF_ref = A_env_top_UF_ref + A_env_bottom_UF_ref

    # A_env_win_south_ref = A_env_win_south_MR_ref + A_env_win_south_NO_ref + A_env_win_south_NO_ref
    # A_env_win_east_ref = A_env_win_east_MR_ref + A_env_win_east_NO_ref + A_env_win_east_NO_ref
    # A_env_win_north_ref = A_env_win_north_MR_ref + A_env_win_north_NO_ref + A_env_win_north_NO_ref
    # A_env_win_west_ref = A_env_win_west_MR_ref + A_env_win_west_NO_ref + A_env_win_west_NO_ref

    # 3.3.10.48	参照住戸の空間ごとの不透明部位の面積の合計
    A_env_door_MR_ref = A_env_door_north_MR_ref + A_env_door_west_MR_ref
    A_env_door_OR_ref = A_env_door_north_OR_ref + A_env_door_west_OR_ref
    A_env_door_NO_ref = A_env_door_north_NO_ref + A_env_door_west_NO_ref

    # A_env_door_north_ref = A_env_door_north_MR_ref + A_env_door_north_OR_ref + A_env_door_north_NO_ref
    # A_env_door_west_ref = A_env_door_west_MR_ref + A_env_door_west_OR_ref + A_env_door_west_NO_ref

    # 3.3.10.47	参照住戸の空間ごとの透明部位の面積の合計
    A_env_win_MR_ref = A_env_win_south_MR_ref + A_env_win_east_MR_ref + A_env_win_north_MR_ref + A_env_win_west_MR_ref
    A_env_win_OR_ref = A_env_win_south_OR_ref + A_env_win_east_OR_ref + A_env_win_north_OR_ref + A_env_win_west_OR_ref
    A_env_win_NO_ref = A_env_win_south_NO_ref + A_env_win_east_NO_ref + A_env_win_north_NO_ref + A_env_win_west_NO_ref
    A_env_win_ref = A_env_win_MR_ref + A_env_win_OR_ref + A_env_win_NO_ref

    # 3.3.10.46	参照住戸の空間ごとの開口部の面積の合計
    A_env_op_MR_ref = A_env_win_MR_ref + A_env_door_MR_ref
    A_env_op_OR_ref = A_env_win_OR_ref + A_env_door_OR_ref
    A_env_op_NO_ref = A_env_win_NO_ref + A_env_door_NO_ref

    # 3.3.10.45	参照住戸の開口部の面積の合計
    A_env_op_ref = A_env_op_MR_ref + A_env_op_OR_ref + A_env_op_NO_ref

    # 3.3.10.44	参照住戸の空間ごとの外皮面積
    A_env_MR_ref = A_env_horz_MR_ref + A_env_vert_MR_ref
    A_env_OR_ref = A_env_horz_OR_ref + A_env_vert_OR_ref
    A_env_NO_ref = A_env_horz_NO_ref + A_env_vert_NO_ref
    A_env_UF_ref = A_env_horz_UF_ref + A_env_vert_UF_ref

    # 3.3.10.43	参照住戸の外皮面積の合計
    A_env_ref = A_env_MR_ref + A_env_OR_ref + A_env_NO_ref + A_env_UF_ref
                   
    # 3.3.10.42	推定住戸の空間ごとの下面の外皮面積
    A_env_bottom_MR = A_env_bottom_MR_ref * A_MR / A_MR_ref
    A_env_bottom_OR = A_env_bottom_OR_ref * A_OR / A_OR_ref
    A_env_bottom_NO = A_env_bottom_NO_ref * A_NO / A_NO_ref
    A_env_bottom_UF = A_env_bottom_UF_ref * (A_MR + A_OR + A_NO) / (A_MR_ref + A_OR_ref + A_NO_ref)

    # 3.3.10.41	推定住戸の下面の外皮面積の合計
    A_env_bottom = A_env_bottom_MR + A_env_bottom_OR + A_env_bottom_NO + A_env_bottom_UF

    # 3.3.10.40	推定住戸の空間ごとの上面の外皮面積
    A_env_top_MR = A_env_top_MR_ref * A_MR / A_MR_ref
    A_env_top_OR = A_env_top_OR_ref * A_OR / A_OR_ref
    A_env_top_NO = A_env_top_NO_ref * A_NO / A_NO_ref
    A_env_top_UF = 0
    A_env_top = A_env_top_MR + A_env_top_OR + A_env_top_NO + A_env_top_UF

    # 3.3.10.39	推定住戸における空間ごとの水平の外皮面積
    A_env_horz_MR = A_env_top_MR + A_env_bottom_MR
    A_env_horz_OR = A_env_top_OR + A_env_bottom_OR
    A_env_horz_NO = A_env_top_NO + A_env_bottom_NO
    A_env_horz_UF = A_env_top_UF + A_env_bottom_UF
    A_env_horz = A_env_horz_MR + A_env_horz_OR + A_env_horz_NO + A_env_horz_UF

    # 3.3.10.38	推定住戸における空間ごとの垂直の外皮の面積の合計
    A_env_vert_MR = max(A_env * A_env_MR_ref / A_env_ref - A_env_horz_MR, 0.00)
    A_env_vert_OR = max(A_env * A_env_OR_ref / A_env_ref - A_env_horz_OR, 0.00)
    A_env_vert_NO = max(A_env * A_env_NO_ref / A_env_ref - A_env_horz_NO, 0.00)
    A_env_vert_UF = max(A_env * A_env_UF_ref / A_env_ref - A_env_horz_UF, 0.00)
    A_env_vert =  A_env_vert_MR + A_env_vert_OR + A_env_vert_NO + A_env_vert_UF

    # 3.3.10.37	推定住戸における空間ごとの外皮面積の合計
    A_env_MR = A_env_horz_MR + A_env_vert_MR
    A_env_OR = A_env_horz_OR + A_env_vert_OR
    A_env_NO = A_env_horz_NO + A_env_vert_NO
    A_env_UF = A_env_horz_UF + A_env_vert_UF

    # 3.3.10.36	推定住戸の空間ごとの西向き外皮の面積の合計
    A_env_west_MR  = A_env_vert_MR * A_env_west_MR_ref  / A_env_vert_MR_ref
    A_env_west_OR  = A_env_vert_OR * A_env_west_OR_ref  / A_env_vert_OR_ref
    A_env_west_NO  = A_env_vert_NO * A_env_west_NO_ref  / A_env_vert_NO_ref
    A_env_west_UF  = A_env_vert_UF * A_env_west_UF_ref  / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_west = A_env_west_MR + A_env_west_OR + A_env_west_NO + A_env_west_UF

    # 3.3.10.35	推定住戸の空間ごとの南向きの外皮面積の合計
    A_env_south_MR = A_env_vert_MR * A_env_south_MR_ref / A_env_vert_MR_ref
    A_env_south_OR = A_env_vert_OR * A_env_south_OR_ref / A_env_vert_OR_ref
    A_env_south_NO = A_env_vert_NO * A_env_south_NO_ref / A_env_vert_NO_ref
    A_env_south_UF = A_env_vert_UF * A_env_south_UF_ref / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_south = A_env_south_MR + A_env_south_OR + A_env_south_NO + A_env_south_UF

    # 3.3.10.34	推定住戸の空間ごとの東向きの外皮面積の合計
    A_env_east_MR  = A_env_vert_MR * A_env_east_MR_ref  / A_env_vert_MR_ref
    A_env_east_OR  = A_env_vert_OR * A_env_east_OR_ref  / A_env_vert_OR_ref
    A_env_east_NO  = A_env_vert_NO * A_env_east_NO_ref  / A_env_vert_NO_ref
    A_env_east_UF  = A_env_vert_UF * A_env_east_UF_ref  / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_east = A_env_east_MR + A_env_east_OR + A_env_east_NO + A_env_east_UF

    # 3.3.10.33	推定住戸の空間ごとの北向きの外皮面積の合計
    A_env_north_MR = A_env_vert_MR * A_env_north_MR_ref / A_env_vert_MR_ref
    A_env_north_OR = A_env_vert_OR * A_env_north_OR_ref / A_env_vert_OR_ref
    A_env_north_NO = A_env_vert_NO * A_env_north_NO_ref / A_env_vert_NO_ref
    A_env_north_UF = A_env_vert_UF * A_env_north_UF_ref / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_north = A_env_north_MR  + A_env_north_OR + A_env_north_NO + A_env_north_UF


    # 計算誤差が発生するため、小数点以下10桁まで一致確認する。
    assert '{:.10f}'.format(A_env_MR + A_env_OR + A_env_NO + A_env_UF) == '{:.10f}'.format(A_env)

    # 3.3.10.32	推定住戸の床以外の外皮がすべて外気に接する場合の総外皮に占める外気に接する外皮面積の割合（共同住宅のみ計算）
    if tatekata == "共同住宅":
        # 下方向を除く5方位が全て外気に接する場合が、外気に接する外皮面積が最大になる。
        A_env_ex_max = A_env_top + A_env_vert
        r_env_ex_max = A_env_ex_max / A_env

    # 3.3.10.31	推定住戸の南向きまたは北向きの外皮のみが外気に接する場合の総外皮に占める外気に接する外皮の面積の割合(共同住宅のみ計算)
    if tatekata == "共同住宅":
        # 北面と南面のみ外気に接する場合が、外気に接する外皮面積が参照になる。
        A_env_ex_min = A_env_south + A_env_north
        r_env_ex_min = A_env_ex_min / A_env

    # 3.3.10.30	推定住戸の総外皮に占める外気に接する外皮面積の割合
    if tatekata == "戸建住宅":
        # 戸建てでは固定で考える
        r_env_ex = 1.0
    elif tatekata == "共同住宅":
        # UA値から外気に接する外皮面積の割合を推定
        r_dash_env_ex = 1.0 / (1.0 + math.exp(-9.10907512 * (ua - 1.05204145)))
        r_env_ex = min(max(r_env_ex_min, r_dash_env_ex), r_env_ex_max)
    else:
        raise ValueError(tatekata)
    
    # 3.3.10.29	推定住戸の外気に接する外皮の面積の合計
    A_env_ex = A_env * r_env_ex

    # 3.3.10.28 推定住戸の外気に接する外皮に占める開口部面積の割合
    r_env_op = get_open_rate(eta_ac)

    # 3.3.10.27 推定住戸の開口部面積の合計
    # 開口部面積の合計 = 外気に接する外皮の面積 * 開口部比率
    A_env_op = get_total_open_area(A_env_ex, r_env_op)

    # 3.3.10.26	推定住戸の間仕切りの面積
    # 間仕切り面積 = 参照住戸の間仕切り面積 * モデル住戸の屋根面積 / 参照住戸の屋根面積
    if has_vertical_internal == "有":
        A_part_MR_OR = A_part_MR_OR_ref * (A_env_vert_MR + A_env_vert_OR) / (A_env_vert_MR_ref + A_env_vert_OR_ref)
        A_part_MR_NO = A_part_MR_NO_ref * (A_env_vert_MR + A_env_vert_NO) / (A_env_vert_MR_ref + A_env_vert_NO_ref)
        A_part_OR_NO = A_part_OR_NO_ref * (A_env_vert_OR + A_env_vert_NO) / (A_env_vert_OR_ref + A_env_vert_NO_ref)
    elif has_vertical_internal == "無":
        A_part_MR_OR = 0.0
        A_part_MR_NO = 0.0
        A_part_OR_NO = 0.0
    else:
        raise ValueError(has_vertical_internal)

    # 推定住戸・参照住戸における、主たる居室・その他の居室・非居室の内壁床の面積の合計
    A_part_bottom_MR = max(A_MR - A_env_bottom_MR, 0.0)
    A_part_bottom_OR = max(A_OR - A_env_bottom_OR, 0.0)
    A_part_bottom_NO = max(A_NO - A_env_bottom_NO, 0.0)
    A_part_bottom_MR_ref = A_part_bottom_MR_MR_ref + A_part_bottom_MR_OR_ref + A_part_bottom_MR_NO_ref + A_part_bottom_MR_UF_ref
    A_part_bottom_OR_ref = A_part_bottom_OR_MR_ref + A_part_bottom_OR_OR_ref + A_part_bottom_OR_NO_ref + A_part_bottom_OR_UF_ref
    A_part_bottom_NO_ref = A_part_bottom_NO_MR_ref + A_part_bottom_NO_OR_ref + A_part_bottom_NO_NO_ref + A_part_bottom_NO_UF_ref

    # 推定住戸における各居室の内壁床の面積を、隣接する居室の違いに応じて割り振る際の比率は、参照住戸における比率と同じとする。
    A_part_bottom_MR_MR = A_part_bottom_MR * A_part_bottom_MR_MR_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_MR_OR = A_part_bottom_MR * A_part_bottom_MR_OR_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_MR_NO = A_part_bottom_MR * A_part_bottom_MR_NO_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_MR_UF = A_part_bottom_MR * A_part_bottom_MR_UF_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_OR_MR = A_part_bottom_OR * A_part_bottom_OR_MR_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_OR_OR = A_part_bottom_OR * A_part_bottom_OR_OR_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_OR_NO = A_part_bottom_OR * A_part_bottom_OR_NO_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_OR_UF = A_part_bottom_OR * A_part_bottom_OR_UF_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_NO_MR = A_part_bottom_NO * A_part_bottom_NO_MR_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0
    A_part_bottom_NO_OR = A_part_bottom_NO * A_part_bottom_NO_OR_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0
    A_part_bottom_NO_NO = A_part_bottom_NO * A_part_bottom_NO_NO_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0
    A_part_bottom_NO_UF = A_part_bottom_NO * A_part_bottom_NO_UF_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0

    # 3.3.10.25	推定住戸の外気に接する外皮面積の合計と外気面積の合計の比
    r_dashdash_env_ex = (A_env_ex - A_env_south - A_env_north) / (A_env_top + A_env_east + A_env_west)

    # 3.3.10.24	推定住戸の空間の下面の外皮面積の合計
    A_env_bottom_MR_ex = A_env_bottom_MR if tatekata == "戸建住宅" else 0.0
    A_env_bottom_OR_ex = A_env_bottom_OR if tatekata == "戸建住宅" else 0.0
    A_env_bottom_NO_ex = A_env_bottom_NO if tatekata == "戸建住宅" else 0.0
    A_env_bottom_UF_ex = A_env_bottom_UF if tatekata == "戸建住宅" else 0.0
    A_env_bottom_ex = A_env_bottom_MR_ex + A_env_bottom_OR_ex + A_env_bottom_NO_ex + A_env_bottom_UF_ex

    # 3.3.10.23	推定住戸の空間の北向きの外皮面積の合計
    A_env_north_MR_ex = A_env_north_MR
    A_env_north_OR_ex = A_env_north_OR
    A_env_north_NO_ex = A_env_north_NO
    A_env_north_UF_ex = A_env_north_UF

    # 3.3.10.22	推定住戸の空間の南向きの外皮面積の合計
    A_env_south_MR_ex = A_env_south_MR
    A_env_south_OR_ex = A_env_south_OR
    A_env_south_NO_ex = A_env_south_NO
    A_env_south_UF_ex = A_env_south_UF

    # 3.3.10.21	推定住戸の空間の西向きの外皮面積の合計
    A_env_west_MR_ex = A_env_west_MR if tatekata == "戸建住宅" else A_env_west_MR * r_dashdash_env_ex
    A_env_west_OR_ex = A_env_west_OR if tatekata == "戸建住宅" else A_env_west_OR * r_dashdash_env_ex
    A_env_west_NO_ex = A_env_west_NO if tatekata == "戸建住宅" else A_env_west_NO * r_dashdash_env_ex
    A_env_west_UF_ex = A_env_west_UF if tatekata == "戸建住宅" else A_env_west_UF * r_dashdash_env_ex

    # 3.3.10.20	推定住戸の空間の東向きの外皮面積の合計
    A_env_east_MR_ex = A_env_east_MR if tatekata == "戸建住宅" else A_env_east_MR * r_dashdash_env_ex
    A_env_east_OR_ex = A_env_east_OR if tatekata == "戸建住宅" else A_env_east_OR * r_dashdash_env_ex
    A_env_east_NO_ex = A_env_east_NO if tatekata == "戸建住宅" else A_env_east_NO * r_dashdash_env_ex
    A_env_east_UF_ex = A_env_east_UF if tatekata == "戸建住宅" else A_env_east_UF * r_dashdash_env_ex

    # 3.3.10.19	推定住戸の空間ごとの上面の外皮面積の合計
    A_env_top_MR_ex = A_env_top_MR if tatekata == "戸建住宅" else A_env_top_MR * r_dashdash_env_ex
    A_env_top_OR_ex = A_env_top_OR if tatekata == "戸建住宅" else A_env_top_OR * r_dashdash_env_ex
    A_env_top_NO_ex = A_env_top_NO if tatekata == "戸建住宅" else A_env_top_NO * r_dashdash_env_ex
    A_env_top_UF_ex = A_env_top_UF if tatekata == "戸建住宅" else A_env_top_UF * r_dashdash_env_ex
    A_env_top_ex = A_env_top_MR_ex + A_env_top_OR_ex + A_env_top_NO_ex + A_env_top_UF_ex

    # --- 参考出力用 ----
    A_env_horz_MR_ex = A_env_top_MR_ex + A_env_bottom_MR_ex
    A_env_horz_OR_ex = A_env_top_OR_ex + A_env_bottom_OR_ex
    A_env_horz_NO_ex = A_env_top_NO_ex + A_env_bottom_NO_ex
    A_env_horz_UF_ex = A_env_top_UF_ex + A_env_bottom_UF_ex
    A_env_horz_ex = A_env_horz_MR_ex + A_env_horz_OR_ex + A_env_horz_NO_ex + A_env_horz_UF_ex

    # --- 参考出力用 ----
    A_env_vert_MR_ex = A_env_south_MR_ex + A_env_east_MR_ex + A_env_north_MR_ex + A_env_west_MR_ex
    A_env_vert_OR_ex = A_env_south_OR_ex + A_env_east_OR_ex + A_env_north_OR_ex + A_env_west_OR_ex
    A_env_vert_NO_ex = A_env_south_NO_ex + A_env_east_NO_ex + A_env_north_NO_ex + A_env_west_NO_ex
    A_env_vert_UF_ex = A_env_south_UF_ex + A_env_east_UF_ex + A_env_north_UF_ex + A_env_west_UF_ex

    # --- 参考出力用 ----
    A_env_MR_ex = A_env_horz_MR_ex + A_env_vert_MR_ex
    A_env_OR_ex = A_env_horz_OR_ex + A_env_vert_OR_ex
    A_env_NO_ex = A_env_horz_NO_ex + A_env_vert_NO_ex
    A_env_UF_ex = A_env_horz_UF_ex + A_env_vert_UF_ex

    # 計算誤差が発生するため、小数点以下10桁まで一致確認する。
    assert '{:.10f}'.format(A_env_MR_ex + A_env_OR_ex + A_env_NO_ex + A_env_UF_ex) == '{:.10f}'.format(A_env_ex)


    # 3.3.10.18	推定住戸の空間ごとの西向きの不透明開口部の面積の合計
    A_env_door_west_MR = min(A_env_op * A_env_door_west_MR_ref / A_env_op_ref, A_env_west_MR_ex)
    A_env_door_west_OR = min(A_env_op * A_env_door_west_OR_ref / A_env_op_ref, A_env_west_OR_ex)
    A_env_door_west_NO = min(A_env_op * A_env_door_west_NO_ref / A_env_op_ref, A_env_west_NO_ex)
    A_env_door_west = A_env_door_west_MR + A_env_door_west_OR + A_env_door_west_NO

    # 3.3.10.17	推定住戸の空間ごとの北向きの不透明開口部の面積の合計
    A_env_door_north_MR = min(A_env_op * A_env_door_north_MR_ref / A_env_op_ref, A_env_north_MR_ex)
    A_env_door_north_OR = min(A_env_op * A_env_door_north_OR_ref / A_env_op_ref, A_env_north_OR_ex)
    A_env_door_north_NO = min(A_env_op * A_env_door_north_NO_ref / A_env_op_ref, A_env_north_NO_ex)
    A_env_door_north = A_env_door_north_MR + A_env_door_north_OR + A_env_door_north_NO

    # 3.3.10.16	推定住戸の不透明部位の面積の合計
    A_env_door = A_env_door_north + A_env_door_west

    # 3.3.10.15	推定住戸の空間ごとの西向きの透明部位の面積の合計
    A_env_win_west_MR = min(A_env_op * A_env_win_west_MR_ref / A_env_op_ref, A_env_west_MR_ex)
    A_env_win_west_OR = min(A_env_op * A_env_win_west_OR_ref / A_env_op_ref, A_env_west_OR_ex)
    A_env_win_west_NO = min(A_env_op * A_env_win_west_NO_ref / A_env_op_ref, A_env_west_NO_ex)
    A_env_win_west = A_env_win_west_MR + A_env_win_west_OR + A_env_win_west_NO

    # 3.3.10.14	推定住戸の空間ごとの南向きの透明部位の面積の合計
    A_env_win_south_MR = min(A_env_op * A_env_win_south_MR_ref / A_env_op_ref, A_env_south_MR_ex)
    A_env_win_south_OR = min(A_env_op * A_env_win_south_OR_ref / A_env_op_ref, A_env_south_OR_ex)
    A_env_win_south_NO = min(A_env_op * A_env_win_south_NO_ref / A_env_op_ref, A_env_south_NO_ex)
    A_env_win_south = A_env_win_south_MR + A_env_win_south_OR + A_env_win_south_NO

    # 3.3.10.13	推定住戸の空間ごとの東向きの透明部位の面積の合計
    A_env_win_east_MR = min(A_env_op * A_env_win_east_MR_ref / A_env_op_ref, A_env_east_MR_ex)
    A_env_win_east_OR = min(A_env_op * A_env_win_east_OR_ref / A_env_op_ref, A_env_east_OR_ex)
    A_env_win_east_NO = min(A_env_op * A_env_win_east_NO_ref / A_env_op_ref, A_env_east_NO_ex)
    A_env_win_east = A_env_win_east_MR + A_env_win_east_OR + A_env_win_east_NO

    # 3.3.10.12	推定住戸の空間ごとの北向きの透明部位の面積の合計
    A_env_win_north_MR = min(A_env_op * A_env_win_north_MR_ref / A_env_op_ref, A_env_north_MR_ex)
    A_env_win_north_OR = min(A_env_op * A_env_win_north_OR_ref / A_env_op_ref, A_env_north_OR_ex)
    A_env_win_north_NO = min(A_env_op * A_env_win_north_NO_ref / A_env_op_ref, A_env_north_NO_ex)
    A_env_win_north = A_env_win_north_MR + A_env_win_north_OR + A_env_win_north_NO

    # 3.3.10.11	推定住戸の透明部位の面積の合計
    A_env_win = A_env_win_south + A_env_win_east + A_env_win_north + A_env_win_west

    # ---- 外気に接する外壁等の計算 ----

    # 3.3.10.10	推定住戸の空間ごとの西向きの外気に接しない外皮面積の合計
    A_env_wall_west_ex_MR = A_env_west_MR_ex - A_env_win_west_MR - A_env_door_west_MR
    A_env_wall_west_ex_OR = A_env_west_OR_ex - A_env_win_west_OR - A_env_door_west_OR
    A_env_wall_west_ex_NO = A_env_west_NO_ex - A_env_win_west_NO - A_env_door_west_NO
    A_env_wall_west_ex = A_env_wall_west_ex_MR + A_env_wall_west_ex_OR + A_env_wall_west_ex_NO

    # 3.3.10.9	推定住戸の空間ごとの南向きの外気に接しない外皮面積の合計
    A_env_wall_south_ex_MR = A_env_south_MR_ex - A_env_win_south_MR
    A_env_wall_south_ex_OR = A_env_south_OR_ex - A_env_win_south_OR
    A_env_wall_south_ex_NO = A_env_south_NO_ex - A_env_win_south_NO
    A_env_wall_south_ex = A_env_wall_south_ex_MR + A_env_wall_south_ex_OR + A_env_wall_south_ex_NO

    # 3.3.10.8	推定住戸の空間ごとの東向きの外気に接しない外皮面積の合計
    A_env_wall_east_ex_MR = A_env_east_MR_ex - A_env_win_east_MR
    A_env_wall_east_ex_OR = A_env_east_OR_ex - A_env_win_east_OR
    A_env_wall_east_ex_NO = A_env_east_NO_ex - A_env_win_east_NO
    A_env_wall_east_ex = A_env_wall_east_ex_MR + A_env_wall_east_ex_OR + A_env_wall_east_ex_NO

    # 3.3.10.7	推定住戸の空間ごとの北向きの外気に接しない外皮面積の合計
    A_env_wall_north_ex_MR = A_env_north_MR_ex - A_env_win_north_MR - A_env_door_north_MR
    A_env_wall_north_ex_OR = A_env_north_OR_ex - A_env_win_north_OR - A_env_door_north_OR
    A_env_wall_north_ex_NO = A_env_north_NO_ex - A_env_win_north_NO - A_env_door_north_NO
    A_env_wall_north_ex = A_env_wall_north_ex_MR + A_env_wall_north_ex_OR + A_env_wall_north_ex_NO

    # !!! 未定義 !!!
    A_env_wall_vert_ex_MR = A_env_wall_south_ex_MR + A_env_wall_east_ex_MR + A_env_wall_north_ex_MR + A_env_wall_west_ex_MR
    A_env_wall_vert_ex_OR = A_env_wall_south_ex_OR + A_env_wall_east_ex_OR + A_env_wall_north_ex_OR + A_env_wall_west_ex_OR
    A_env_wall_vert_ex_NO = A_env_wall_south_ex_NO + A_env_wall_east_ex_NO + A_env_wall_north_ex_NO + A_env_wall_west_ex_NO
    A_env_wall_vert_ex = A_env_wall_vert_ex_MR + A_env_wall_vert_ex_OR + A_env_wall_vert_ex_NO


    # 3.3.10.6	推定住戸の空間ごとの下面の外気に接しない外皮面積の合計
    A_env_bottom_MR_in = A_env_bottom_MR - A_env_bottom_MR_ex
    A_env_bottom_OR_in = A_env_bottom_OR - A_env_bottom_OR_ex
    A_env_bottom_NO_in = A_env_bottom_NO - A_env_bottom_NO_ex
    A_env_bottom_UF_in = A_env_bottom_UF - A_env_bottom_UF_ex
    A_env_bottom_in = A_env_bottom_MR_in + A_env_bottom_OR_in + A_env_bottom_NO_in + A_env_bottom_UF_in

    # # 3.3.10.4	推定住戸の外気に接する外皮の面積
    # external_envelope_area_wons = A_env_ex - A_env_ex_min
    # if external_envelope_rate == r_env_ex_min:
    #     external_envelope_area_wons = 0.0

    # 3.3.10.5	推定住戸の空間ごとの西向きの外気に接しない外皮面積の合計
    A_env_west_MR_in = A_env_west_MR - A_env_west_MR_ex
    A_env_west_OR_in = A_env_west_OR - A_env_west_OR_ex
    A_env_west_NO_in = A_env_west_NO - A_env_west_NO_ex
    A_env_west_UF_in = A_env_west_UF - A_env_west_UF_ex
    A_env_west_in = A_env_west_MR_in + A_env_west_OR_in + A_env_west_NO_in + A_env_west_UF_in

    # 3.3.10.4	推定住戸の空間ごとの南向きの外気に接しない外皮面積の合計
    A_env_south_MR_in = A_env_south_MR - A_env_south_MR_ex
    A_env_south_OR_in = A_env_south_OR - A_env_south_OR_ex
    A_env_south_NO_in = A_env_south_NO - A_env_south_NO_ex
    A_env_south_UF_in = A_env_south_UF - A_env_south_UF_ex
    A_env_south_in = A_env_south_MR_in + A_env_south_OR_in + A_env_south_NO_in + A_env_south_UF_in

    # 3.3.10.3	推定住戸の空間ごとの東向きの外気に接しない外皮面積の合計
    A_env_east_MR_in = A_env_east_MR - A_env_east_MR_ex
    A_env_east_OR_in = A_env_east_OR - A_env_east_OR_ex
    A_env_east_NO_in = A_env_east_NO - A_env_east_NO_ex
    A_env_east_UF_in = A_env_east_UF - A_env_east_UF_ex
    A_env_east_in = A_env_east_MR_in + A_env_east_OR_in + A_env_east_NO_in + A_env_east_UF_in

    # 3.3.10.2	推定住戸の空間ごとの北向きの外気に接しない外皮面積の合計
    A_env_north_MR_in = A_env_north_MR - A_env_north_MR_ex
    A_env_north_OR_in = A_env_north_OR - A_env_north_OR_ex
    A_env_north_NO_in = A_env_north_NO - A_env_north_NO_ex
    A_env_north_UF_in = A_env_north_UF - A_env_north_UF_ex
    A_env_north_in = A_env_north_MR_in + A_env_north_OR_in + A_env_north_NO_in + A_env_north_UF_in

    # 3.3.10.1	推定住戸の空間ごとの上面の外気に接しない外皮面積の合計
    A_env_top_MR_in = A_env_top_MR - A_env_top_MR_ex
    A_env_top_OR_in = A_env_top_OR - A_env_top_OR_ex
    A_env_top_NO_in = A_env_top_NO - A_env_top_NO_ex
    A_env_top_UF_in = A_env_top_UF - A_env_top_UF_ex
    A_env_top_in = A_env_top_MR_in + A_env_top_OR_in + A_env_top_NO_in + A_env_top_UF_in

    # --- 参考出力用 ---
    A_env_horz_MR_in = A_env_horz_MR - A_env_horz_MR_ex
    A_env_horz_OR_in = A_env_horz_OR - A_env_horz_OR_ex
    A_env_horz_NO_in = A_env_horz_NO - A_env_horz_NO_ex
    A_env_horz_UF_in = A_env_horz_UF - A_env_horz_UF_ex
    A_env_horz_in = A_env_horz_MR_in + A_env_horz_OR_in + A_env_horz_NO_in + A_env_horz_UF_in

    # --- 参考出力用 ---
    A_env_vert_MR_in = A_env_vert_MR - A_env_vert_MR_ex
    A_env_vert_OR_in = A_env_vert_OR - A_env_vert_OR_ex
    A_env_vert_NO_in = A_env_vert_NO - A_env_vert_NO_ex
    A_env_vert_UF_in = A_env_vert_UF - A_env_vert_UF_ex
    A_env_vert_in = A_env_vert_MR_in + A_env_vert_OR_in + A_env_vert_NO_in + A_env_vert_UF_in

    # --- 参考出力用 ---
    A_env_MR_in = A_env_MR - A_env_MR_ex
    A_env_OR_in = A_env_OR - A_env_OR_ex
    A_env_NO_in = A_env_NO - A_env_NO_ex
    A_env_UF_in = A_env_UF - A_env_UF_ex
    A_env_in = A_env_MR_in + A_env_OR_in + A_env_NO_in + A_env_UF_in


    # 3.3.11 居室の容積 (室内高=2.4m想定)
    V_room_MR = 2.4 * A_MR
    V_room_OR = 2.4 * A_OR
    V_room_NO = 2.4 * A_NO
    V_room_UF = 0.4 * A_env_bottom_UF

    # 3.3.12 居室の換気量 (1時間当たり5回想定)
    V_vent_MR = 5.0 * V_room_MR
    V_vent_OR = 5.0 * V_room_OR
    V_vent_NO = 5.0 * V_room_NO
    V_vent_UF = 5.0 * V_room_UF


    # 表11 参照住戸の壁体構成等
    
    # 参照住戸の外気に接する屋根・天井のコンクリートの厚み
    t_roof_ex_concre_ref = 0.150
    t_wall_ex_concre_ref = 0.135
    t_floor_ex_concre_ref = 0.150
    t_base_ex_concre_ref = 0.12
    R_fix_roof = t_roof_ex_concre_ref / concrete_conductivity if tatekata == "共同住宅" else 0.043
    R_fix_wall = t_wall_ex_concre_ref / concrete_conductivity if tatekata == "共同住宅" else 0.296
    R_fix_floor = t_floor_ex_concre_ref / concrete_conductivity if tatekata == "共同住宅" else 0.12 / 1.6
    R_fix_base = t_base_ex_concre_ref / concrete_conductivity

    ##concrete_thickness = TriValue(0.150, 0.135, 0.150)

    # 表12 参照住戸の開口部の熱貫流率
    U_win_ref = 2.0    #窓の基準U
    U_door_ref = 6.0   #ドアの基準U
    if region == 8:
        # 8地域の場合は基準をmax相当にする
        U_win_ref = 10.0
        U_door_ref = 10.0
    U_floor_ex_ref = 0.19
    # U_base_ref = 0.48
    Psi_base_ref = 0.99

    # 表面熱伝達抵抗
    Rs_roof_ex_ref = 0.13
    Rs_wall_ex_ref = 0.15
    Rs_floor_ex_ref = 0.30
    Rs_base_ref = 0.15

    # 基準断熱厚 (断熱材の基準厚は0.05mとする, 外気に接する床は0.2)
    t_roof_ex_ins_ref = 0.20
    t_wall_ex_ins_ref = 0.15
    t_floor_ex_ins_ref = 0.10
    t_base_ins_ref = 0.10

    # ---------------------------------------------
    # ----- 3.3.8 熱貫流率および断熱材厚みの推定 -----
    # ---------------------------------------------

    # 3.3.8.10 温度差係数
    H_os = 1.0      # 外気に接する部位の温度差係数
    H_is = 0.0      # 外気に接しない部位および隣接住戸と接する内壁の温度差係数
    H_floor = 0.7   # 外気に通じる床裏の温度差係数
    
    # 3.3.8.9	推定住戸の部位の種類ごとの面積


    # 3.3.8.8	推定住戸が無断熱であると仮定した場合の熱貫流率
    # 最悪U [W/m2K]
    U_roof_ex_max = get_wall_transfer_rate(R_fix=R_fix_roof, insulation_thickness=0.0, H=H_os, Rs=Rs_roof_ex_ref)
    U_wall_ex_max = get_wall_transfer_rate(R_fix=R_fix_wall, insulation_thickness=0.0, H=H_os, Rs=Rs_wall_ex_ref)
    U_floor_ex_max = get_wall_transfer_rate(R_fix=R_fix_floor, insulation_thickness=0.0, H=H_floor, Rs=Rs_floor_ex_ref)
    U_base_max = get_wall_transfer_rate(R_fix=R_fix_base, insulation_thickness=0.0, H=H_os, Rs=Rs_base_ref)  # 無断熱相当の基礎壁の熱貫流率
    Psi_base_max = 0.99    # 無断熱相当の基礎の線熱貫流率
    U_win_max = 10      #窓の最悪U追加
    U_door_max = 10     #ドアの最悪U追加

    # 3.3.8.7 推定住戸が無断熱であると仮定した場合の熱損失量
    # 最悪q [W/K] = 最悪U * 面積
    q_roof_max  = U_roof_ex_max  * A_env_top_ex
    q_wall_max  = U_wall_ex_max  * A_env_wall_vert_ex
    q_floor_max = U_floor_ex_max * (A_env_bottom_ex - A_env_bottom_UF_ex)
    q_base_max = U_base_max * A_env_vert_UF
    q_win_max   = U_win_max  * A_env_win
    q_door_max  = U_door_max * A_env_door
    q_max = q_roof_max + q_wall_max + q_floor_max + q_base_max + q_win_max + q_door_max

    # 3.3.8.6 参照住戸の部位の熱貫流率
    # 基準U [W/m2K]
    U_roof_ex_ref = get_wall_transfer_rate(R_fix=R_fix_roof, insulation_thickness=t_roof_ex_ins_ref, H=H_os, Rs=Rs_roof_ex_ref)
    U_wall_ex_ref = get_wall_transfer_rate(R_fix=R_fix_wall, insulation_thickness=t_wall_ex_ins_ref, H=H_os, Rs=Rs_wall_ex_ref)
    U_floor_ex_ref = get_wall_transfer_rate(R_fix=R_fix_floor, insulation_thickness=t_floor_ex_ins_ref, H=H_floor, Rs=Rs_floor_ex_ref)
    U_base_ref = get_wall_transfer_rate(R_fix=R_fix_base, insulation_thickness=t_base_ins_ref, H=H_os, Rs=Rs_base_ref)
    # Psi_base_ref = 0.99    # 基礎の線熱貫流率

    # 3.3.8.5 推定住戸の熱貫流率が参照住戸と同等であると仮定した場合の熱損失量
    # 基準q [W/K] = 基準U * 面積
    q_roof_basis = U_roof_ex_ref * A_env_top_ex
    q_wall_basis = U_wall_ex_ref * A_env_wall_vert_ex
    q_floor_basis = U_floor_ex_ref *  (A_env_bottom_ex - A_env_bottom_UF_ex)
    q_base_basis = U_base_ref * A_env_vert_UF
    q_win_basis = U_win_ref * A_env_win
    q_door_basis = U_door_ref * A_env_door
    q_basis = q_roof_basis + q_wall_basis + q_floor_basis + q_base_basis + q_win_basis + q_door_basis

    q_roof_margin = q_roof_max - q_roof_basis
    q_wall_margin = q_wall_max - q_wall_basis
    q_floor_margin = q_floor_max - q_floor_basis
    q_base_margin = q_base_max - q_base_basis
    q_win_margin = q_win_max - q_win_basis
    q_door_margin = q_door_max - q_door_basis
    q_margin = q_roof_margin + q_wall_margin + q_floor_margin + q_base_margin + q_win_margin + q_door_margin

    # 差分q = 入力された条件におけるq値 - 基準qの合計
    q_diff = A_env * ua - q_basis

    r_q_margin_plus = q_diff / q_margin
    r_q_margin_minus =  - q_diff / q_basis 

    # 3.3.8.4 設計住戸の熱損失量に合わせるための調整量
    def f_offset(x, y):

        # q値の割り当て [W/K] = 基準q値 or (最悪q - 基準q値)
        if q_diff <= 0:
            return -y * r_q_margin_minus
        else:
            # x-y: q余裕
            return (x - y) * r_q_margin_plus

    # 3.3.8.3 推定住戸の部位の熱損失量 [W/K]
    q_roof_ex  = q_roof_basis  + f_offset(q_roof_max,  q_roof_basis)
    q_wall_ex  = q_wall_basis  + f_offset(q_wall_max,  q_wall_basis)
    q_floor_ex = q_floor_basis + f_offset(q_floor_max, q_floor_basis)
    q_base     = q_base_basis  + f_offset(q_base_max,  q_base_basis)
    q_win      = q_win_basis   + f_offset(q_win_max,   q_win_basis)
    q_door     = q_door_basis  + f_offset(q_door_max,  q_door_basis)

    # 3.3.8.2 推定住戸の部位の熱貫流率(負荷計算への入力) [W/㎡K]
    U_roof_ex  = q_roof_ex  / A_env_top_ex if A_env_top_ex > 0.0 else 0.0
    U_wall_ex  = q_wall_ex  / A_env_wall_vert_ex if A_env_wall_vert_ex > 0.0 else 0.0
    U_floor_ex = q_floor_ex / A_env_bottom_ex if A_env_bottom_ex > 0.0 else 0.0
    U_base     = q_base     / A_env_vert_UF if A_env_vert_UF > 0.0 else 0.0
    U_win      = q_win      / A_env_win if A_env_win > 0.0 else 0.0
    U_door     = q_door     / A_env_door if A_env_door > 0.0 else 0.0
    # NOTE: 基礎のあたりの処理が怪しい

    # 3.3.8.1 推定住戸の断熱材の熱抵抗(負荷計算への入力) [m]
    R_roof_ex_ins  = get_insulation_registance(U_roof_ex, R_fix_roof, H_os, Rs_roof_ex_ref)
    R_wall_ex_ins  = get_insulation_registance(U_wall_ex, R_fix_wall, H_os, Rs_wall_ex_ref)
    R_floor_ex_ins = get_insulation_registance(U_floor_ex, R_fix_floor, H_floor, Rs_floor_ex_ref)
    R_base_ins = get_insulation_registance(U_base, R_fix_base, H_os, Rs_base_ref)

    # --------------------------------
    # ----- 3.3.9 窓の日射熱取得率 -----
    # --------------------------------

    # 3.3.9.8 暖冷房期間
    DD_H, DD_C = get_master_days(region)

    # 3.3.9.7 方位係数
    neu_c, neu_h = get_neu_avg(region)

    # 3.3.9.6 設計住戸の年間平均日射熱取得率
    # --------------------------------------------------------
    # 年間平均日射熱取得率は冷房期平均日射熱取得率および暖房期平均日射熱取得率を冷房期間および
    # 暖房期間の日数で案分して求める。
    eta_avg = (eta_ac * DD_C + eta_ah * DD_H) / (DD_H + DD_C)

    # 3.3.9.5 設計住戸の外皮全体の日射熱取得 [W/(W/㎡)]
    # --------------------------------------------------------
    # 外皮全体の日射熱取得量は年間平均日射熱取得率と外皮の面積の合計から求められる。
    m = eta_avg / 100 * A_env

    # 3.3.9.4 推定住戸の冷房期の不透明部位の日射熱取得量と暖房期の不透明部位の日射熱取得量
    # 窓以外の日射熱取得量は熱損失率から計算される  [W/(W・㎡)]
    # NOTE: 戸建住宅に対応するため、基礎壁からの日射熱取得の計算式が追加されている。
    m_opaq_C = \
        A_env_top_ex * U_roof_ex * neu_c[0] * 0.034 \
        + A_env_wall_south_ex * U_wall_ex * neu_c[5] * 0.034 \
        + A_env_wall_east_ex * U_wall_ex * neu_c[3] * 0.034 \
        + A_env_wall_north_ex * U_wall_ex * neu_c[1] * 0.034 \
        + A_env_wall_west_ex * U_wall_ex * neu_c[7] * 0.034 \
        + A_env_south_UF_ex * U_base * neu_c[5] * 0.034 \
        + A_env_east_UF_ex * U_base * neu_c[3] * 0.034 \
        + A_env_north_UF_ex * U_base * neu_c[1] * 0.034 \
        + A_env_west_UF_ex * U_base * neu_c[7] * 0.034 \
        + A_env_door_north  * U_door * neu_c[1] * 0.034 \
        + A_env_door_west  * U_door * neu_c[7] * 0.034
    m_opaq_H = \
        A_env_top_ex * U_roof_ex * neu_h[0] * 0.034 \
        + A_env_wall_south_ex * U_wall_ex * neu_h[5] * 0.034 \
        + A_env_wall_east_ex * U_wall_ex * neu_h[3] * 0.034 \
        + A_env_wall_north_ex * U_wall_ex * neu_h[1] * 0.034 \
        + A_env_wall_west_ex * U_wall_ex * neu_h[7] * 0.034 \
        + A_env_south_UF_ex * U_base * neu_h[5] * 0.034 \
        + A_env_east_UF_ex * U_base * neu_h[3] * 0.034 \
        + A_env_north_UF_ex * U_base * neu_h[1] * 0.034 \
        + A_env_west_UF_ex * U_base * neu_h[7] * 0.034 \
        + A_env_door_north * U_door * neu_h[1] * 0.034 \
        + A_env_door_west * U_door * neu_h[7] * 0.034
    
    # 3.3.9.3 推定住戸の不透明部位の日射熱取得量
    # ---------------------------------------------------------------
    # 不透明部位の日射熱取得量は冷房期の不透明部位の日射熱取得量と暖房期の不透明部位の日射熱取得量を
    # 冷房期間および暖房期間の日数で案分して求める。
    m_opaq = (m_opaq_C * DD_C + m_opaq_H * DD_H) / (DD_C + DD_H)

    # 3.3.9.2 推定住戸の透明部位(窓)の日射熱取得量 [W/(W/㎡)]
    # ---------------------------------------------------------------
    # 透明部位（窓）の日射熱取得量は、外皮全体の日射熱取得量から不透明部位の日射熱取得量を減じること
    # で求める。
    m_tran = m - m_opaq

    # 3.3.9.1 推定住戸の窓の日射熱取得率(負荷計算への入力) [(W/㎡)/(W/㎡)]
    # NOTE: 日射熱取得率の按分方法がExcelと異なる
    eta_win = calc_eta_win(m_tran, (A_env_win_south, A_env_win_east, A_env_win_north, A_env_win_west), neu_c, neu_h, DD_C, DD_H)

    # ------------------------------------------------------------------------

    print("計算条件")
    print("-------------------------------------------------")
    print("地域区分: {}".format(region))
    print("延床面積: {} [㎡]".format(total_floor_area))
    print(" 主たる居室: {} [㎡]".format(main_floor_area))
    print(" その他居室: {} [㎡]".format(other_floor_area))
    print("外皮総面積: {} [㎡]".format(A_env))
    print("外皮平均熱貫流率: {} [W/K]".format(ua))
    print("暖房期平均日射熱取得率: {}".format(eta_ah))
    print("冷房期平均日射熱取得率: {}".format(eta_ac))
    print("")
    print("計算結果")
    print("-------------------------------------------------")
    print("外皮面積")
    print("  主たる居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_MR, A_env_east_MR, A_env_north_MR, A_env_west_MR))
    print("  その他居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_OR, A_env_east_OR, A_env_north_OR, A_env_west_OR))
    print("  非居室:     {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_NO, A_env_east_NO, A_env_north_NO, A_env_west_NO))
    print("外気に接する屋根の面積")
    print("  主たる居室: {:3.2f} [㎡]".format(A_env_top_MR_ex))
    print("  その他居室: {:3.2f} [㎡]".format(A_env_top_OR_ex))
    print("  非居室:     {:3.2f} [㎡]".format(A_env_top_NO_ex))
    print("外気に接する外壁・基礎壁の面積")
    print("  主たる居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_wall_south_ex_MR, A_env_wall_east_ex_MR, A_env_wall_north_ex_MR, A_env_wall_west_ex_MR))
    print("  その他居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_wall_south_ex_OR, A_env_wall_east_ex_OR, A_env_wall_north_ex_OR, A_env_wall_west_ex_OR))
    print("  非居室:     {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_wall_south_ex_NO, A_env_wall_east_ex_NO, A_env_wall_north_ex_NO, A_env_wall_west_ex_NO))
    print("  床下空間:   {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_UF_ex, A_env_east_UF_ex, A_env_north_UF_ex, A_env_west_UF_ex))
    print("窓面積")
    print("  主たる居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_win_south_MR, A_env_win_east_MR, A_env_win_north_MR, A_env_win_west_MR))
    print("  その他居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_win_south_OR, A_env_win_east_OR, A_env_win_north_OR, A_env_win_west_OR))
    print("  非居室:     {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_win_south_NO, A_env_win_east_NO, A_env_win_north_NO, A_env_win_west_NO))
    print("ドア面積")
    print("  主たる居室: --- --- {:3.2f} {:3.2f} [㎡]".format(A_env_door_north_MR, A_env_door_west_MR))
    print("  その他居室: --- --- {:3.2f} {:3.2f} [㎡]".format(A_env_door_north_OR, A_env_door_west_OR))
    print("  非居室:     --- --- {:3.2f} {:3.2f} [㎡]".format(A_env_door_north_NO, A_env_door_west_NO))
    print("間仕切り")
    print("  主居室 - その他居室: {:3.2f} [㎡]".format(A_part_MR_OR))
    print("  主居室 - 非居室: {:3.2f} [㎡]".format(A_part_MR_NO))
    print("  その他居室 - 非居室: {:3.2f} [㎡]".format(A_part_OR_NO))
    print("内壁床")
    print("  主居室 - 主居室: {:3.2f} [㎡]".format(A_part_bottom_MR_MR))
    print("  主居室 - その他居室: {:3.2f} [㎡]".format(A_part_bottom_MR_OR))
    print("  主居室 - 非居室: {:3.2f} [㎡]".format(A_part_bottom_MR_NO))
    print("  主居室 - 床下空間: {:3.2f} [㎡]".format(A_part_bottom_MR_UF))
    print("  その他居室 - 主居室: {:3.2f} [㎡]".format(A_part_bottom_OR_MR))
    print("  その他居室 - その他居室: {:3.2f} [㎡]".format(A_part_bottom_OR_OR))
    print("  その他居室 - 非居室: {:3.2f} [㎡]".format(A_part_bottom_OR_NO))
    print("  その他居室 - 床下空間: {:3.2f} [㎡]".format(A_part_bottom_OR_UF))
    print("  非居室 - 主居室: {:3.2f} [㎡]".format(A_part_bottom_NO_MR))
    print("  非居室 - その他居室: {:3.2f} [㎡]".format(A_part_bottom_NO_OR))
    print("  非居室 - 非居室: {:3.2f} [㎡]".format(A_part_bottom_NO_NO))
    print("  非居室 - 床下空間: {:3.2f} [㎡]".format(A_part_bottom_NO_UF))
    print("熱貫流率")
    print("  外気に接する屋根: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_roof_ex, R_roof_ex_ins))
    print("  外気に接する外壁等: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_wall_ex, R_wall_ex_ins))
    print("  外気に接する床下: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_floor_ex, R_floor_ex_ins))
    print("  外気に接する基礎壁: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_base, R_base_ins))
    print("  窓: {:3.2f} [W/㎡K]".format(U_win))
    print("  ドア: {:3.2f} [W/㎡K]".format(U_door))
    print("窓の日射熱取得率: {:3.2f} [(W/㎡)/(W/㎡)]".format(eta_win))

    # 入力Excel のテンプレートファイルを読み込み
    template_xlsx = _get_template_xlsx(tatekata, structure)

    # 入力シート内の各種変数名を変換する辞書を取得
    varname_mapper = {
        # rooms 
        # 居室の面積
        '#MR_A': A_MR,
        '#OR_A': A_OR,
        '#NO_A': A_NO,
        '#UF_A': A_env_bottom_UF,
        '#MR_VOL': V_room_MR,
        '#OR_VOL': V_room_OR,
        '#NO_VOL': V_room_NO,
        '#UF_VOL': V_room_UF,
        '#MR_VENT': V_vent_MR,
        '#OR_VENT': V_vent_OR,
        '#NO_VENT': V_vent_NO,
        '#UF_VENT': V_vent_UF,

        # external_general_parts
        # 垂直外壁と屋根・天井の面積
        # 外気に接する外壁等
        '#MR_A_C_EW': A_env_top_MR_ex,   #2F天井相当(屋根)
        '#OR_A_C_EW': A_env_top_OR_ex,   #2F天井相当(屋根)
        '#NO_A_C_EW': A_env_top_NO_ex,   #2F天井相当(屋根)
        '#MR_A_S_EW': A_env_wall_south_ex_MR,
        '#OR_A_S_EW': A_env_wall_south_ex_OR,
        '#NO_A_S_EW': A_env_wall_south_ex_NO,
        '#MR_A_E_EW': A_env_wall_east_ex_MR,
        '#OR_A_E_EW': A_env_wall_east_ex_OR,
        '#NO_A_E_EW': A_env_wall_east_ex_NO,
        '#MR_A_N_EW': A_env_wall_north_ex_MR,
        '#OR_A_N_EW': A_env_wall_north_ex_OR,
        '#NO_A_N_EW': A_env_wall_north_ex_NO,
        '#MR_A_W_EW': A_env_wall_west_ex_MR,
        '#OR_A_W_EW': A_env_wall_west_ex_OR,
        '#NO_A_W_EW': A_env_wall_west_ex_NO,
        '#MR_A_F_EW': A_env_bottom_MR_ex, #1F床相当(床断熱の場合)
        '#OR_A_F_EW': A_env_bottom_OR_ex, #1F床相当(床断熱の場合)
        '#NO_A_F_EW': A_env_bottom_NO_ex, #1F床相当(床断熱の場合)
        '#UF_A_S_EW': A_env_south_UF_ex, #床下の基礎(基礎断熱の場合)
        '#UF_A_E_EW': A_env_east_UF_ex,  #床下の基礎(基礎断熱の場合)
        '#UF_A_N_EW': A_env_north_UF_ex, #床下の基礎(基礎断熱の場合)
        '#UF_A_W_EW': A_env_west_UF_ex,  #床下の基礎(基礎断熱の場合)
        # 外気に接しない外壁等
        '#MR_A_C_PW': A_env_top_MR_in,   #2F天井相当(戸境壁)
        '#OR_A_C_PW': A_env_top_OR_in,   #2F天井相当(戸境壁)
        '#NO_A_C_PW': A_env_top_NO_in,   #2F天井相当(戸境壁)
        '#MR_A_E_PW': A_env_east_MR_in,  #戸境壁
        '#OR_A_E_PW': A_env_east_OR_in,  #戸境壁
        '#NO_A_E_PW': A_env_east_NO_in,  #戸境壁
        '#MR_A_W_PW': A_env_west_MR_in,  #戸境壁
        '#OR_A_W_PW': A_env_west_OR_in,  #戸境壁
        '#NO_A_W_PW': A_env_west_NO_in,  #戸境壁
        '#MR_A_F_PW': A_env_bottom_MR_in,    #1F床相当
        '#OR_A_F_PW': A_env_bottom_OR_in,     #1F床相当
        '#NO_A_F_PW': A_env_bottom_NO_in,     #1F床相当

        # external_opaque_parts
        # ドアの熱貫流率と面積
        '#DOOR_U': U_door,
        '#MR_DOOR_A_N': A_env_door_north_MR,
        '#OR_DOOR_A_N': A_env_door_north_OR,
        '#NO_DOOR_A_N': A_env_door_north_NO,
        '#MR_DOOR_A_W': A_env_door_west_MR,
        '#OR_DOOR_A_W': A_env_door_west_OR,
        '#NO_DOOR_A_W': A_env_door_west_NO,

        # external_transparent_parts
        # 窓の熱貫流率、日射熱取得率と面積
        '#WINDOW_U': U_win,
        '#WINDOW_ETA': eta_win,
        '#MR_A_WIN_S': A_env_win_south_MR,
        '#MR_A_WIN_E': A_env_win_east_MR,
        '#MR_A_WIN_N': A_env_win_north_MR,
        '#MR_A_WIN_W': A_env_win_west_MR,
        '#OR_A_WIN_S': A_env_win_south_OR,
        '#OR_A_WIN_E': A_env_win_east_OR,
        '#OR_A_WIN_N': A_env_win_north_OR,
        '#OR_A_WIN_W': A_env_win_west_OR,
        '#NO_A_WIN_S': A_env_win_south_NO,
        '#NO_A_WIN_E': A_env_win_east_NO,
        '#NO_A_WIN_N': A_env_win_north_NO,
        '#NO_A_WIN_W': A_env_win_west_NO,

        # partitions
        # 間仕切りの面積
        '#IN_MR_OR': A_part_MR_OR,
        '#IN_MR_NO': A_part_MR_NO,
        '#IN_OR_NO': A_part_OR_NO,

        # 内壁床の面積
        # NOTE: 同じ室用途どうしで接する内壁床の面積は、「温度差係数 0 の外気に接する床」として割り振る点に注意
        # ref. 2025年3月14日 暖冷房負荷評価枠組検討TG コアMTG ( https://youworks.atlassian.net/l/cp/AvAiqBoB )
        '#IN_MR_MR_F': A_part_bottom_MR_MR,
        '#IN_MR_OR_F': A_part_bottom_MR_OR,
        '#IN_MR_NO_F': A_part_bottom_MR_OR,
        '#IN_MR_UF_F': A_part_bottom_MR_UF,
        '#IN_OR_MR_F': A_part_bottom_OR_MR,
        '#IN_OR_OR_F': A_part_bottom_OR_OR,
        '#IN_OR_NO_F': A_part_bottom_OR_NO,
        '#IN_OR_UF_F': A_part_bottom_OR_UF,
        '#IN_NO_MR_F': A_part_bottom_NO_MR,
        '#IN_NO_OR_F': A_part_bottom_NO_OR,
        '#IN_NO_NO_F': A_part_bottom_NO_NO,
        '#IN_NO_UF_F': A_part_bottom_NO_UF,

        # layers
        # 屋根と垂直外壁の断熱材熱抵抗
        '#CEILING_R_INSULATION': R_roof_ex_ins,
        '#WALL_R_INSULATION': R_wall_ex_ins,
        '#FLOOR_R_INSULATION': R_floor_ex_ins,
        '#BASE_R_INSULATION': R_base_ins,
    }

    # 入力シート内の各種変数名を変換
    with pd.ExcelWriter(xlsx_path) as writer:
        for (sheet_name, df) in template_xlsx.items():
            df.replace(varname_mapper).to_excel(writer, sheet_name=sheet_name, index=False)



# --------------------------------------------------------------------------------
# 3.3.9.7 方位係数
# --------------------------------------------------------------------------------

def get_neu_avg(region: int) -> Tuple[
        Tuple[float,float,float,float,float,float,float,float,float,float],
        Tuple[float,float,float,float,float,float,float,float,float,float]
    ]:
    """指定された地域の区分の冷房期と暖房期の方位係数を返します。

    Args:
        region (int): 地域の区分1-8

    Returns:
        冷房期の方位係数と暖房期の方位係数をタプルで返します。
        それぞの方位係数は方位ごとに細分化されており、
        上面, 北, 北東, 東, 南東, 南, 南西, 西, 北西, 下面の順に格納されます。
    
    Notes:
        平成28年省エネルギー基準に準拠したエネルギー消費性能の評価に関する技術情報
        （住宅エネルギー消費性能の算定方法 第三章 暖冷房負荷と外皮性能 
        第二節 外皮性能 付録C 方位係数に定義される暖房期の方位係数と冷房期の方位係数) より
    """

    ## 方位係数
    ## index: 屋根・上面, 北, 北東, 東, 南東, 南, 南西, 西, 北西, 下面

    # 冷房期の方位係数
    neu_c = [
        [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
        [0.329, 0.341, 0.335, 0.322, 0.373, 0.341, 0.307, 0.325],
        [0.430, 0.412, 0.390, 0.426, 0.437, 0.431, 0.415, 0.414],
        [0.545, 0.503, 0.468, 0.518, 0.500, 0.512, 0.509, 0.515],
        [0.560, 0.527, 0.487, 0.508, 0.500, 0.498, 0.490, 0.528],
        [0.502, 0.507, 0.476, 0.437, 0.472, 0.434, 0.412, 0.480],
        [0.526, 0.548, 0.550, 0.481, 0.520, 0.491, 0.479, 0.517],
        [0.508, 0.529, 0.553, 0.481, 0.518, 0.504, 0.495, 0.505],
        [0.411, 0.428, 0.447, 0.401, 0.442, 0.427, 0.406, 0.411],
        [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    ]

    # 暖房期の方位係数
    neu_h = [
        [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 0.0],
        [0.260, 0.263, 0.284, 0.256, 0.238, 0.261, 0.227, 0.000],
        [0.333, 0.341, 0.348, 0.330, 0.310, 0.325, 0.281, 0.000],
        [0.564, 0.554, 0.540, 0.531, 0.568, 0.579, 0.543, 0.000],
        [0.823, 0.766, 0.751, 0.724, 0.846, 0.833, 0.843, 0.000],
        [0.935, 0.856, 0.851, 0.815, 0.983, 0.936, 1.023, 0.000],
        [0.790, 0.753, 0.750, 0.723, 0.815, 0.763, 0.848, 0.000],
        [0.535, 0.544, 0.542, 0.527, 0.538, 0.523, 0.548, 0.000],
        [0.325, 0.341, 0.351, 0.326, 0.297, 0.317, 0.284, 0.000],
        [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    ]

    return [neu_c[_][region - 1] for _ in range(10)], [neu_h[_][region - 1] for _ in range(10)]


# --------------------------------------------------------------------------------
# 3.3.9.8 暖冷房期間
# --------------------------------------------------------------------------------

def get_master_days(region: int) -> Tuple[int, int]:
    """指定された地域の区分の暖房期間と冷房期間の日数を返します。

    Args:
        region (int): 地域の区分1-8

    Returns:
        暖房期間の日数と冷房期間の日数をタプルで返します。
    
    Notes:
        平成28年省エネルギー基準に準拠したエネルギー消費性能の評価に関する技術情報
        （住宅エネルギー消費性能の算定方法 第十一章 その他)より
    """
    return (
        (257, 53),
        (252, 48),
        (244, 53),
        (242, 53),
        (218, 57),
        (169, 117),
        (122, 152),
        (0, 265),
    )[region - 1]


# --------------------------------------------------------------------------------
# 3.3.10.51	参照住戸の面積
# --------------------------------------------------------------------------------

def get_floor_area_ref(tatekata: str) -> Tuple[float,float,float]:
    """参照住戸の床面積

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        主たる居室、その他の居室、非居室の面積
    """
    if tatekata == "共同住宅":
        # 表5 参照住戸の床面積（共同住宅の場合）
        A_MR_ref, A_OR_ref, A_NO_ref = 24.23, 29.75, 16.02
    elif tatekata == "戸建住宅":
        # 表6 参照住戸の床面積（戸建住宅の場合）
        A_MR_ref, A_OR_ref, A_NO_ref = 29.81, 51.35, 38.93
    else:
        raise ValueError(tatekata)
    return A_MR_ref, A_OR_ref, A_NO_ref


def get_area_table_ref(tatekata: str) -> Tuple[Tuple[float]] :
    """参照住戸の面積を一覧表

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        2次元配列で面積を返します。
        1次元目: 主たる居室 / その他の居室 / 非居室 / 床下空間
        2次元目: 外皮(上面/北/東/南/西/下面) / 窓(北/東/南/西) / ドア(北/西)
        それぞれ、対応する記号は次の通りです。
        主たる居室,その他の居室,非居室,床下空間 = MR,OR,NO,UF
        外皮,窓,ドア = "env", "env,win", "env,door"
        上面,北,東,南,西,下面 = top,north,east,south,west,bottom
        例) 参照住戸の主たる居室の外皮の上面の面積の合計の記号は A_env_top_MR_ref
    """
    if tatekata == "共同住宅":
        # 表7 参照住戸の面積(共同住宅の場合)
        return (
            # 主たる居室 / その他の居室 / 非居室 / 床下空間
            (24.23, 29.75, 16.02, 0.00),    # 外皮-上面
            (0.00, 11.80, 4.16, 0.00),      # 外皮-北
            (0.00, 21.59, 8.05, 0.00),      # 外皮-東
            (9.52, 6.45, 0.00, 0.00),       # 外皮-南 
            (17.21, 10.06, 2.37, 0.00),     # 外皮-西
            (24.23, 29.75, 16.02, 0.00),    # 外皮-下面
            (0.00, 2.53, 0.00),             # 窓-北
            (0.00, 0.00, 0.00),             # 窓-東
            (4.52, 3.24, 0.00),             # 窓-南
            (0.00, 0.00, 0.00),             # 窓-西
            (0.00, 0.00, 1.76),             # ドア-北
            (0.00, 0.00, 0.00),             # ドア-西
        )
    
    elif tatekata == "戸建住宅":
        # 表8 参照住戸の面積(戸建住宅の場合)
        return (
            (0.00, 34.79, 17.40, 0.00),     # 外皮-上面
            (5.12, 6.77, 39.08, 2.91),      # 外皮-北
            (17.20, 8.74, 4.36, 3.28),      # 外皮-東
            (14.21, 29.26, 0.00, 2.91),     # 外皮-南 
            (0.00, 17.48, 13.20, 3.28),     # 外皮-西
            (29.81, 16.56, 21.53, 55.48),   # 外皮-下面
            (0.00, 4.59, 3.15),             # 窓-北
            (3.13, 0.66, 0.00),             # 窓-東
            (6.94, 8.17, 0.00),             # 窓-南
            (0.00, 0.99, 1.08),             # 窓-西
            (1.62, 0.00, 0.00),             # ドア-北
            (0.00, 0.00, 1.89),             # ドア-西
        )
    
    else:
        raise ValueError(tatekata)


def get_partition_table_ref(tatekata: str) -> Tuple[float]:
    """参照住戸の間仕切りの面積

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        以下の順番で参照住戸の間仕切り面積をリストで返します。
        1.主たる居室とその他の居室 A_part,MR->OR,ref
        2.主たる居室と非居室 A_part,MR->NO,ref
        3.その他の居室と非居室 A_part,OR->NO,ref
    """
    if tatekata == "共同住宅":
        # 表9 参照住戸の間仕切りの面積(共同住宅)
        return (12.53, 16.19, 40.51)
    elif tatekata == "戸建住宅":
        # 表10 参照住戸の間仕切りの面積(戸建住宅)
        return (8.64, 17.20, 29.51)
    else:
        raise ValueError(tatekata)


def get_partition_bottom_table_ref(tatekata: str) -> Tuple[float]:
    """参照住戸の内壁床の面積

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        以下の順番で参照住戸の内壁床の面積をリストで返します。
        1.主たる居室の床面積のうち、別の主たる居室に接する部分      A_part,bottom,MR->MR,ref
        2.主たる居室の床面積のうち、その他の居室に接する部分        A_part,bottom,MR->OR,ref
        3.主たる居室の床面積のうち、非居室に接する部分              A_part,bottom,MR->NO,ref
        4.主たる居室の床面積のうち、床下空間に接する部分            A_part,bottom,MR->UF,ref
        5.その他の居室の床面積のうち、主たる居室に接する部分        A_part,bottom,OR->MR,ref
        6.その他の居室の床面積のうち、別のその他の居室に接する部分  A_part,bottom,OR->OR,ref
        7.その他の居室の床面積のうち、非居室に接する部分            A_part,bottom,OR->NO,ref
        8.その他の居室の床面積のうち、床下空間に接する部分          A_part,bottom,OR->UF,ref
        9.非居室の床面積のうち、主たる居室に接する部分              A_part,bottom,NO->MR,ref
        10.非居室の床面積のうち、その他の居室に接する部分           A_part,bottom,NO->OR,ref
        11.非居室の床面積のうち、別の非居室に接する部分             A_part,bottom,NO->NO,ref
        12.非居室の床面積のうち、床下空間に接する部分               A_part,bottom,NO->UF,ref

    """
    if tatekata == "共同住宅":
        return (00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00)
    elif tatekata == "戸建住宅":
        return (00.00, 00.00, 00.00, 29.81, 21.53, 13.25, 00.00, 16.56, 04.14, 00.00, 12.42, 21.53)
    else:
        raise ValueError(tatekata)


@functools.lru_cache
def _get_template_xlsx(tatekata, structure) -> dict[str, pd.DataFrame]:

    if tatekata == "戸建住宅":
        if structure == "床断熱" or structure == '床下断熱':
            template_xlsx_path = 'simple_input_excel_template_kodate_yukadan.xlsx'
        elif structure == "基礎断熱":
            template_xlsx_path = 'simple_input_excel_template_kodate_kisodan.xlsx'
        else:
            raise ValueError(structure)
    elif tatekata == "共同住宅":
            template_xlsx_path = 'simple_input_excel_template_kyodo.xlsx'
    else:
        raise ValueError(tatekata)
    
    if not os.path.isabs(template_xlsx_path):
        template_xlsx_path = os.path.join(os.path.dirname(__file__), 'templates', template_xlsx_path)

    return pd.read_excel(template_xlsx_path, sheet_name=None)


if __name__ == "__main__":
    # 各種入力
    region = 3
    estimate(
        region=region,
        total_floor_area=83.38,
        main_floor_area=29.225,
        other_floor_area=34.47,
        A_env=264.12,
        ua=1.991226,
        eta_ah=2,
        eta_ac=1,
        tatekata="戸建住宅",
        structure="基礎断熱",
        xlsx_path='test.xlsx'
    )

    import json
    input_json = to_json('test.xlsx', region)
    print(input_json)
    with open('test.json', mode='w') as input_json_file:
        json.dump(input_json, input_json_file, indent=4)    
