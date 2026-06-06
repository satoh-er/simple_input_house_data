import os
import functools
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
import pandas as pd
import math

SIMPLE_INPUT_R5_MODIFIED_VERSION = "2026-06-06-verification-1"


# master_menseki = DedicatedAreaTable

# コンクリート, 断熱材の熱伝導率λ
concrete_conductivity = 1.6
insulation_conductivity = 0.03

def to_json(input_xlsx_filepath, region):
    import openpyxl


    def convert_to_input_json(
        input_xlsx_filepath: str,
    ):
        book = openpyxl.load_workbook(input_xlsx_filepath)

        sheet_common = book['common']
        sheet_building = book['building']
        sheet_rooms = book['rooms']
        sheet_external_general_parts = book['external_general_parts']
        sheet_external_opaque_parts = book['external_opaque_parts']
        sheet_external_transparent_parts = book['external_transparent_parts']
        sheet_internals = book['internals']
        sheet_grounds = book['grounds']
        sheet_layers = book['layers']
        
        n_rooms = count_number_in_id_row(sheet=sheet_rooms)
        n_external_general_parts = count_number_in_id_row(sheet=sheet_external_general_parts)
        n_external_opaque_parts = count_number_in_id_row(sheet=sheet_external_opaque_parts)
        n_external_transparent_parts = count_number_in_id_row(sheet=sheet_external_transparent_parts)
        n_internals = count_number_in_id_row(sheet=sheet_internals)
        n_grounds = count_number_in_id_row(sheet=sheet_grounds)
        n_layers = count_number_in_id_row(sheet=sheet_layers)
        
        common = {
            'ac_method': sheet_common.cell(column=2, row=2).value,
            'weather': {
                'method': 'ees',
                'region': region,
            }
        }
        
        building = {
            "infiltration": {
                "method": "balance_residential",
                "c_value_estimate": "specify",
                "story": int(sheet_building.cell(column=2, row=2).value),
                "c_value": float(sheet_building.cell(column=3, row=2).value),
                "inside_pressure": sheet_building.cell(column=4, row=2).value
            }
        }
        
        rooms = [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "floor_area": float(row[4].value),
                "volume": float(row[5].value),
                "ventilation": {
                    "natural": float(row[6].value)
                },
                "furniture": {
                    "input_method": "default"
                },
                "schedule": {
                    "name": row[7].value
                }
            } for row in sheet_rooms.iter_rows(min_row=2, max_row=n_rooms+1)
        ]

        layers_master = [
            {
                "name": row[1].value,
                "layers": make_dictionary_of_layer(row)[0],
                "reversed_layers": make_dictionary_of_layer(row)[1]
            } for row in sheet_layers.iter_rows(min_row=2, max_row=n_layers+1)
        ]
        
        external_general_parts =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "external_general_part",
                "area": float(row[5].value),
                "h_c": get_h_c(direction=row[8].value),
                "is_solar_absorbed_inside": bool(row[6].value),
                "is_floor": bool(row[6].value),
                "layers": get_layers(layers_master, layer_name=row[7].value),
                "solar_shading_part": {"existence": False},
                "is_sun_striked_outside": row[9].value == 1.0,  # fix by kitamura
                "direction": row[8].value,
                "outside_emissivity": 0.9,
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[8].value, temp_dif_coef=float(row[9].value)),
                "outside_solar_absorption": 0.8,
                "temp_dif_coef": float(row[9].value)
            } for row in sheet_external_general_parts.iter_rows(min_row=2, max_row=n_external_general_parts+1)
            if float(row[5].value) > 0.0
        ]

        external_opaque_parts =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "external_opaque_part",
                "area": float(row[5].value),
                "h_c": get_h_c(direction=row[7].value),
                "is_solar_absorbed_inside": False,
                "is_floor": False,
                "solar_shading_part": {"existence": False},
                "is_sun_striked_outside": True,
                "direction": row[7].value,
                "outside_emissivity": 0.9,
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[7].value, temp_dif_coef=1.0),
                "u_value": float(row[6].value),
                "inside_heat_transfer_resistance": 0.11,
                "outside_solar_absorption": 0.8,
                "temp_dif_coef": 1.0
            } for row in sheet_external_opaque_parts.iter_rows(min_row=2, max_row=n_external_opaque_parts+1)
            if float(row[5].value) > 0.0
        ]

        external_transparent_parts =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "external_transparent_part",
                "area": float(row[5].value),
                "h_c": get_h_c(direction=row[10].value),
                "is_solar_absorbed_inside": False,
                "is_floor": False,
                "solar_shading_part": get_solar_shading(exist=bool(row[11].value), depth=row[12].value, d_h=row[13].value, d_e=row[14].value),
                "is_sun_striked_outside": True,
                "direction": row[10].value,
                "outside_emissivity": 0.9,
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[10].value, temp_dif_coef=1.0),
                "u_value": float(row[6].value),
                "inside_heat_transfer_resistance": 0.11,
                "eta_value": float(row[7].value),
                "incident_angle_characteristics": row[8].value,
                "glass_area_ratio": float(row[9].value),
                "temp_dif_coef": 1.0
            } for row in sheet_external_transparent_parts.iter_rows(min_row=2, max_row=n_external_transparent_parts+1)
            if float(row[5].value) > 0.0
        ]
        
        internals_2d =  [
            [
                {
                    "id": row[1].value,
                    "name": row[3].value,
                    "sub_name": row[5].value or '',
                    "connected_room_id": int(row[7].value),
                    "boundary_type": "internal",
                    "area": float(row[9].value),
                    "h_c": get_h_c(direction=row[11].value)[0],
                    "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[0],
                    "is_floor": get_is_floor(direction=row[11].value)[0],
                    "layers": get_layers(layers_master, layer_name=row[10].value, is_reverse=False),
                    "solar_shading_part": {"existence": False},
                    "rear_surface_boundary_id": row[2].value
                },
                {
                    "id": row[2].value,
                    "name": row[4].value,
                    "sub_name": row[6].value or '',
                    "connected_room_id": int(row[8].value),
                    "boundary_type": "internal",
                    "area": float(row[9].value),
                    "h_c": get_h_c(direction=row[11].value)[1],
                    "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[1],
                    "is_floor": get_is_floor(direction=row[11].value)[1],
                    "layers": get_layers(layers_master, layer_name=row[10].value, is_reverse=True),
                    "solar_shading_part": {"existence": False},
                    "rear_surface_boundary_id": row[1].value
                }
            ] for row in sheet_internals.iter_rows(min_row=2, max_row=n_internals+1)
            if float(row[9].value) > 0.0
        ]
        # flatten
        internals = sum(internals_2d, [])

        grounds =  [
            {
                "id": row[1].value,
                "name": row[2].value,
                "sub_name": row[3].value or '',
                "connected_room_id": int(row[4].value),
                "boundary_type": "ground",
                "area": float(row[5].value),
                "is_solar_absorbed_inside": bool(row[7].value),
                "is_floor": True,
                "h_c": get_h_c(direction='bottom'),
                "layers": get_layers(layers_master, layer_name=row[6].value),
                "solar_shading_part": {"existence": False},
            } for row in sheet_grounds.iter_rows(min_row=2, max_row=n_grounds+1)
            if float(row[5].value) > 0.0
        ]

        # 各境界id及び裏面境界idを修正
        # ※面積0の境界がスキップされるため、修正前の時点では境界のidと実際の配列のインデックスにずれが発生してしまっている。
        boundaries = external_general_parts + external_opaque_parts + external_transparent_parts + internals + grounds
        boundaries_id_mapper = {_bd['id']: _new_id for _new_id, _bd in enumerate(boundaries)}
        for bd in boundaries:
            bd['id'] = boundaries_id_mapper[bd['id']]
            if bd['boundary_type'] == 'internal':
                bd['rear_surface_boundary_id'] = boundaries_id_mapper[bd['rear_surface_boundary_id']]

        ventilation_rate = 0.5
        V_MR, V_OR, V_NR = [_["volume"] for _ in rooms[:3]]
        v_vent_MR = ventilation_rate * (V_MR + V_NR * V_MR / (V_MR + V_OR))
        v_vent_OR = ventilation_rate * (V_OR + V_NR * V_OR / (V_MR + V_OR))
        mechanical_ventilations = [
            {
                "id": 0,
                "root_type": "type3",
                "volume": v_vent_MR,
                "root": [
                    0,
                    2
                ]
            },
            {
                "id": 1,
                "root_type": "type3",
                "volume": v_vent_OR,
                "root": [
                    1,
                    2
                ]
            }
        ]
        
        equipment_c_MR, equipment_h_MR = create_equipments(id=0, space_id=0, a_floor_is=rooms[0]['floor_area'])
        equipment_c_OR, equipment_h_OR = create_equipments(id=1, space_id=1, a_floor_is=rooms[1]['floor_area'])
        equipments = {
            "heating_equipments": [equipment_h_MR, equipment_h_OR],
            "cooling_equipments": [equipment_c_MR, equipment_c_OR]
        }

        return {
            "common":common,
            "building": building,
            "rooms": rooms,
            "boundaries": boundaries,
            "mechanical_ventilations": mechanical_ventilations,
            "equipments": equipments
        }


    def count_number_in_id_row(sheet):
        id_all = [row[1].value for row in sheet.rows][1:]
        return len(id_all) - (id_all).count(None)


    def make_dictionary_of_layer(row):
        n = int(row[2].value)
        # NOTE: 熱抵抗が 0 未満の layer は生成しない（heat_load_calc による計算時にエラーになるため）
        layer = [
            {
                "name": row[3+3*i].value,
                "thermal_resistance": float(row[4+3*i].value),
                "thermal_capacity": float(row[5+3*i].value)
            } for i in range(n) if float(row[4+3*i].value) > 0.0
        ]
        # Tuple(layer_list, reversed_layer_list)
        return layer, layer[::-1]


    def get_layers(layers_master, layer_name, is_reverse=False):
        # use variable 'layers_master' as global variable
        layers = list(filter(lambda d: d['name'] == layer_name , layers_master))
        if len(layers) > 1:
            raise Exception("Match over one layer.")
        if len(layers) == 0:
            raise Exception("Can't find the layer")
        if is_reverse:
            return layers[0]['reversed_layers']
        else:
            return layers[0]['layers']
        
        
    def get_h_c(direction):
        if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
            return 2.5
        elif direction == 'bottom':
            return 0.7
        elif direction == 'top':
            return 5.0
        elif direction == 'horizontal':
            return (2.5, 2.5)
        elif direction == 'upward':
            return (5.0, 0.7)
        elif direction == 'downward':
            return (0.7, 5.0)
        else:
            raise ValueError(direction)
        
        
    def get_outside_heat_transfer_resistance(direction, temp_dif_coef):
        is_parting = (temp_dif_coef != 1.0)

        if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
            return 0.04 if not is_parting else 0.11
        elif direction == 'bottom':
            return 0.15
        elif direction == 'top':
            return 0.04 if not is_parting else 0.09
        else:
            raise Exception()
        
        
    def get_solar_shading(exist: bool, depth=None, d_h=None, d_e=None):
        if exist:
            return {
                "existence": True,
                "input_method": "simple",
                "depth": float(depth),
                "d_h": float(d_h),
                "d_e": float(d_e)
            }
        else:
            return {
                "existence": False
            }


    def get_is_floor(direction):
        if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se', 'top']:
            return False
        elif direction == 'bottom':
            return True
        elif direction == 'horizontal':
            return (False, False)
        elif direction == 'upward':
            return (False, True)
        elif direction == 'downward':
            return (True, False)
        else:
            raise Exception()


    def create_equipments(id, space_id, a_floor_is):
        q_rtd_c = 190.5 * a_floor_is + 45.6
        q_rtd_h = 1.2090 * q_rtd_c - 85.1

        q_max_c = max(0.8462 * q_rtd_c + 1205.9, q_rtd_c)
        q_max_h = max(1.7597 * q_max_c - 413.7, q_rtd_h)
        
        q_min_c = 500
        q_min_h = 500

        v_max_c = 11.076 * (q_rtd_c / 1000.0) ** 0.3432
        v_max_h = 11.076 * (q_rtd_h / 1000.0) ** 0.3432

        v_min_c = v_max_c * 0.55
        v_min_h = v_max_h * 0.55
        
        bf_c = 0.2
        bf_h = 0.2

        cooling_equipment = {
            "id": id,
            "name": f"cooling_equipment no.{id}",
            "equipment_type": "rac",
            "property": {
                "space_id": space_id,
                "q_min": q_min_c,
                "q_max": q_max_c,
                "v_min": v_min_c,
                "v_max": v_max_c,
                "bf": bf_c
            }
        }
        
        heating_equipment = {
            "id": id,
            "name": f"heating_equipment no.{id}",
            "equipment_type": "rac",
            "property": {
                "space_id": space_id,
                "q_min": q_min_h,
                "q_max": q_max_h,
                "v_min": v_min_h,
                "v_max": v_max_h,
                "bf": bf_h
            }
        }
        
        return cooling_equipment, heating_equipment
    
    return convert_to_input_json(input_xlsx_filepath)





def get_wall_transfer_rate(R_fix, insulation_thickness, H, Rs):

    # 熱抵抗
    total_registance = sum([
        Rs,
        R_fix,
        1.0 / insulation_conductivity * insulation_thickness,
    ])

    return 1.0 / total_registance * H


def get_insulation_registance(wall_transfer_rate, R_fix, tempdiff_coeff, surface_registance):

    if wall_transfer_rate == 0.0 or tempdiff_coeff == 0.0:
        return 0.0

    # 壁体の熱抵抗
    total_registance = 1.0 / (wall_transfer_rate / tempdiff_coeff)

    # 断熱材の熱抵抗 = 壁体の熱抵抗からコンクリートと表面の熱抵抗を除いたもの
    insulation_registance = total_registance - surface_registance - R_fix

    return insulation_registance

# 3.3.10.6 暖冷房負荷モデルの開口部面積の合計
# 開口部面積の合計 = 外気に接する外皮の面積 * 開口部比率
def get_total_open_area(A_env_ex, r_env_op):
    return A_env_ex * r_env_op

# 3.4.2.4.1 暖冷房負荷モデルの外気に接する外皮に占める開口部面積の割合
def get_open_rate(eta_ac):
    return 1.0 / (1.0 + math.exp(-0.129 * (eta_ac - 13.35)))



def calc_eta_win(m_tran, A_env_win, neu_c, neu_h, DD_C, DD_H):

    # 3.3.9.1 暖冷房負荷モデルの窓の日射熱取得率(負荷計算への入力)
    # 窓に割り当てられた日射熱取得量を方位係数を考慮しつつ、面積で割って日射熱取得率にする


    def get_proportion_ratio(A_env_win_dir: float, neu_dir_c: float, neu_dir_h: float):

        # 3.3.9.9 取得日射熱補正係数
        solar_heat_acquisition_coeff_c = 0.93
        solar_heat_acquisition_coeff_h = 0.51

        return (A_env_win_dir * neu_dir_c * solar_heat_acquisition_coeff_c * DD_C \
                + A_env_win_dir * neu_dir_h * solar_heat_acquisition_coeff_h * DD_H) / (DD_C + DD_H)

    # 窓の日射熱取得率 [(W/㎡)/(W/㎡)]
    # NOTE: 以前 neu_c, neu_h の式の実装誤りがあったようだ。方位係数が正しく選ばれていなかったように思われる。
    A_env_win_south, A_env_win_east, A_env_win_north, A_env_win_west = A_env_win
    proportion_ratio_n = get_proportion_ratio(A_env_win_north, neu_c[1], neu_h[1])
    proportion_ratio_e = get_proportion_ratio(A_env_win_east, neu_c[3], neu_h[3])
    proportion_ratio_s = get_proportion_ratio(A_env_win_south, neu_c[5], neu_h[5])
    proportion_ratio_w = get_proportion_ratio(A_env_win_west, neu_c[7], neu_h[7])

    # NOTE: 窓のη値が0.0以下の場合heat_load_calcでは計算不可能のため、最小値を1e-8とする。
    eta_win = max(1e-8, m_tran / (proportion_ratio_n + proportion_ratio_e + proportion_ratio_s + proportion_ratio_w))

    return eta_win



def estimate(region, total_floor_area, main_floor_area, other_floor_area, A_env, ua, eta_ah, eta_ac, tatekata, structure, xlsx_path, has_vertical_internal="有"):
        
    ### ざっくり入力された面積から補正率を計算する ###

    ## 床面積の補正率

    # 入力された床面積の整理

    # 居室の面積
    A_MR = main_floor_area
    A_OR = other_floor_area
    A_NO = total_floor_area - main_floor_area - other_floor_area

    # 主たる居室の床面積の入力値との比率
    # その他の居室の床面積の入力値との比率
    # 非居室の床面積の入力値との比率
    # floor_area_rate = TriValue(
    #     A_MR / A_env_horz_MR,
    #     A_OR / A_env_horz_OR,
    #     A_NO / A_env_horz_NO,
    # )

    # -------------------------
    # ---- 3.3.10 面積計算 -----
    # -------------------------

    # 3.3.10.51	参照住戸の面積

    A_MR_ref, A_OR_ref, A_NO_ref = get_floor_area_ref(tatekata)

    # 参照住戸の面積
    area_table_ref = get_area_table_ref(tatekata)
    A_env_top_MR_ref, A_env_top_OR_ref, A_env_top_NO_ref, A_env_top_UF_ref = area_table_ref[0]
    A_env_north_MR_ref, A_env_north_OR_ref, A_env_north_NO_ref, A_env_north_UF_ref = area_table_ref[1]
    A_env_east_MR_ref, A_env_east_OR_ref, A_env_east_NO_ref, A_env_east_UF_ref = area_table_ref[2]
    A_env_south_MR_ref, A_env_south_OR_ref, A_env_south_NO_ref, A_env_south_UF_ref = area_table_ref[3]
    A_env_west_MR_ref, A_env_west_OR_ref, A_env_west_NO_ref, A_env_west_UF_ref = area_table_ref[4]
    A_env_bottom_MR_ref, A_env_bottom_OR_ref, A_env_bottom_NO_ref, A_env_bottom_UF_ref = area_table_ref[5]
    A_env_win_north_MR_ref, A_env_win_north_OR_ref, A_env_win_north_NO_ref = area_table_ref[6]
    A_env_win_east_MR_ref, A_env_win_east_OR_ref, A_env_win_east_NO_ref = area_table_ref[7]
    A_env_win_south_MR_ref, A_env_win_south_OR_ref, A_env_win_south_NO_ref = area_table_ref[8]
    A_env_win_west_MR_ref, A_env_win_west_OR_ref, A_env_win_west_NO_ref = area_table_ref[9]
    A_env_door_north_MR_ref, A_env_door_north_OR_ref, A_env_door_north_NO_ref = area_table_ref[10]
    A_env_door_west_MR_ref, A_env_door_west_OR_ref, A_env_door_west_NO_ref = area_table_ref[11]

    # 断熱方法によって値を読み変える
    # 床断熱の場合: 床下空間の北・東・南・西・下面の外皮面積を0とする
    # 基礎断熱の場合: 主たる居室・その他の居室・非居室の下面の外皮面積を0とする
    if tatekata == "戸建住宅":
        if structure == "基礎断熱":
            A_env_bottom_MR_ref = 0.0
            A_env_bottom_OR_ref = 0.0
            A_env_bottom_NO_ref = 0.0
        elif structure in ["床断熱", "床下断熱"]:
            A_env_bottom_UF_ref = 0.0
            A_env_south_UF_ref = 0.0
            A_env_east_UF_ref = 0.0
            A_env_north_UF_ref = 0.0
            A_env_west_UF_ref = 0.0
        else:
            raise ValueError(structure)

    # 参照住戸の間仕切りの面積
    partition_table_ref = get_partition_table_ref(tatekata)
    A_part_MR_OR_ref, A_part_MR_NO_ref, A_part_OR_NO_ref = partition_table_ref

    # 参照住戸の内壁床の面積
    partition_bottom_table_ref = get_partition_bottom_table_ref(tatekata)
    A_part_bottom_MR_MR_ref, A_part_bottom_MR_OR_ref, A_part_bottom_MR_NO_ref, A_part_bottom_MR_UF_ref, A_part_bottom_OR_MR_ref, A_part_bottom_OR_OR_ref, A_part_bottom_OR_NO_ref, A_part_bottom_OR_UF_ref, A_part_bottom_NO_MR_ref, A_part_bottom_NO_OR_ref, A_part_bottom_NO_NO_ref, A_part_bottom_NO_UF_ref = partition_bottom_table_ref

    # 戸建住宅(基礎断熱)以外の場合、床下空間に接する内壁床の面積を0とする。
    if not (tatekata == "戸建住宅" and structure == "基礎断熱"):
        A_part_bottom_MR_UF_ref = 0.0
        A_part_bottom_OR_UF_ref = 0.0
        A_part_bottom_NO_UF_ref = 0.0

    # 3.3.10.50	参照住戸の空間ごとの垂直の外皮の面積
    A_env_vert_MR_ref = A_env_south_MR_ref + A_env_east_MR_ref + A_env_north_MR_ref + A_env_west_MR_ref
    A_env_vert_OR_ref = A_env_south_OR_ref + A_env_east_OR_ref + A_env_north_OR_ref + A_env_west_OR_ref
    A_env_vert_NO_ref = A_env_south_NO_ref + A_env_east_NO_ref + A_env_north_NO_ref + A_env_west_NO_ref
    A_env_vert_UF_ref = A_env_south_UF_ref + A_env_east_UF_ref + A_env_north_UF_ref + A_env_west_UF_ref

    # 3.3.10.49	参照住戸の空間ごとの水平の外皮の面積
    A_env_horz_MR_ref = A_env_top_MR_ref + A_env_bottom_MR_ref
    A_env_horz_OR_ref = A_env_top_OR_ref + A_env_bottom_OR_ref
    A_env_horz_NO_ref = A_env_top_NO_ref + A_env_bottom_NO_ref
    A_env_horz_UF_ref = A_env_top_UF_ref + A_env_bottom_UF_ref

    # A_env_win_south_ref = A_env_win_south_MR_ref + A_env_win_south_NO_ref + A_env_win_south_NO_ref
    # A_env_win_east_ref = A_env_win_east_MR_ref + A_env_win_east_NO_ref + A_env_win_east_NO_ref
    # A_env_win_north_ref = A_env_win_north_MR_ref + A_env_win_north_NO_ref + A_env_win_north_NO_ref
    # A_env_win_west_ref = A_env_win_west_MR_ref + A_env_win_west_NO_ref + A_env_win_west_NO_ref
    
    # 3.3.10.48	参照住戸の空間ごとの不透明部位の面積の合計
    A_env_door_MR_ref = A_env_door_north_MR_ref + A_env_door_west_MR_ref
    A_env_door_OR_ref = A_env_door_north_OR_ref + A_env_door_west_OR_ref
    A_env_door_NO_ref = A_env_door_north_NO_ref + A_env_door_west_NO_ref

    # A_env_door_north_ref = A_env_door_north_MR_ref + A_env_door_north_OR_ref + A_env_door_north_NO_ref
    # A_env_door_west_ref = A_env_door_west_MR_ref + A_env_door_west_OR_ref + A_env_door_west_NO_ref

    # 3.3.10.47	参照住戸の空間ごとの透明部位の面積の合計
    A_env_win_MR_ref = A_env_win_south_MR_ref + A_env_win_east_MR_ref + A_env_win_north_MR_ref + A_env_win_west_MR_ref
    A_env_win_OR_ref = A_env_win_south_OR_ref + A_env_win_east_OR_ref + A_env_win_north_OR_ref + A_env_win_west_OR_ref
    A_env_win_NO_ref = A_env_win_south_NO_ref + A_env_win_east_NO_ref + A_env_win_north_NO_ref + A_env_win_west_NO_ref
    A_env_win_ref = A_env_win_MR_ref + A_env_win_OR_ref + A_env_win_NO_ref

    # 3.3.10.46	参照住戸の空間ごとの開口部の面積の合計
    A_env_op_MR_ref = A_env_win_MR_ref + A_env_door_MR_ref
    A_env_op_OR_ref = A_env_win_OR_ref + A_env_door_OR_ref
    A_env_op_NO_ref = A_env_win_NO_ref + A_env_door_NO_ref

    # 3.3.10.45	参照住戸の開口部の面積の合計
    A_env_op_ref = A_env_op_MR_ref + A_env_op_OR_ref + A_env_op_NO_ref

    # 3.3.10.44	参照住戸の空間ごとの外皮面積
    A_env_MR_ref = A_env_horz_MR_ref + A_env_vert_MR_ref
    A_env_OR_ref = A_env_horz_OR_ref + A_env_vert_OR_ref
    A_env_NO_ref = A_env_horz_NO_ref + A_env_vert_NO_ref
    A_env_UF_ref = A_env_horz_UF_ref + A_env_vert_UF_ref

    # 3.3.10.43	参照住戸の外皮面積の合計
    A_env_ref = A_env_MR_ref + A_env_OR_ref + A_env_NO_ref + A_env_UF_ref

    # 3.4.2.1.1	暖冷房負荷モデルの空間ごとの上面の外皮面積
    # TODO: 変数名の記載順序を空間名_方位名にしたい
    A_env_top_MR = A_env_top_MR_ref * A_MR / A_MR_ref
    A_env_top_OR = A_env_top_OR_ref * A_OR / A_OR_ref
    A_env_top_NO = A_env_top_NO_ref * A_NO / A_NO_ref
    A_env_top_UF = 0
    A_env_top = A_env_top_MR + A_env_top_OR + A_env_top_NO + A_env_top_UF
                   
    # 3.4.2.1.2	暖冷房負荷モデルの空間ごとの下面の外皮面積
    A_env_bottom_MR = A_env_bottom_MR_ref * A_MR / A_MR_ref
    A_env_bottom_OR = A_env_bottom_OR_ref * A_OR / A_OR_ref
    A_env_bottom_NO = A_env_bottom_NO_ref * A_NO / A_NO_ref
    A_env_bottom_UF = A_env_bottom_UF_ref * (A_MR + A_OR + A_NO) / (A_MR_ref + A_OR_ref + A_NO_ref)

    # 3.4.2.2.1	暖冷房負荷モデルにおける床下空間の総垂直外皮面積
    A_env_vert_UF = A_env_vert_UF_ref * A_env_bottom_UF / A_env_bottom_UF_ref if A_env_bottom_UF_ref > 0.0 else 0.0

    # 3.4.2.2.2	暖冷房負荷モデルにおける空間ごとの総垂直外皮面積
    A_env_vert = max(A_env - A_env_top_MR - A_env_bottom_MR - A_env_vert_UF, 0.00)
    A_env_vert_MR = max(A_env_vert * A_MR / total_floor_area, 0.00)
    A_env_vert_OR = max(A_env_vert * A_OR / total_floor_area, 0.00)
    A_env_vert_NO = max(A_env_vert * A_NO / total_floor_area, 0.00)

    # 3.4.2.2.3	暖冷房負荷モデルの空間ごとの方位ごと垂直外皮面積（西）
    A_env_west_MR  = A_env_vert_MR * A_env_west_MR_ref  / A_env_vert_MR_ref
    A_env_west_OR  = A_env_vert_OR * A_env_west_OR_ref  / A_env_vert_OR_ref
    A_env_west_NO  = A_env_vert_NO * A_env_west_NO_ref  / A_env_vert_NO_ref
    A_env_west_UF  = A_env_vert_UF * A_env_west_UF_ref  / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_west = A_env_west_MR + A_env_west_OR + A_env_west_NO + A_env_west_UF

    # 3.4.2.2.3	暖冷房負荷モデルの空間ごとの方位ごと垂直外皮面積（南）
    A_env_south_MR = A_env_vert_MR * A_env_south_MR_ref / A_env_vert_MR_ref
    A_env_south_OR = A_env_vert_OR * A_env_south_OR_ref / A_env_vert_OR_ref
    A_env_south_NO = A_env_vert_NO * A_env_south_NO_ref / A_env_vert_NO_ref
    A_env_south_UF = A_env_vert_UF * A_env_south_UF_ref / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_south = A_env_south_MR + A_env_south_OR + A_env_south_NO + A_env_south_UF

    # 3.4.2.2.3	暖冷房負荷モデルの空間ごとの方位ごと垂直外皮面積（東）
    A_env_east_MR  = A_env_vert_MR * A_env_east_MR_ref  / A_env_vert_MR_ref
    A_env_east_OR  = A_env_vert_OR * A_env_east_OR_ref  / A_env_vert_OR_ref
    A_env_east_NO  = A_env_vert_NO * A_env_east_NO_ref  / A_env_vert_NO_ref
    A_env_east_UF  = A_env_vert_UF * A_env_east_UF_ref  / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_east = A_env_east_MR + A_env_east_OR + A_env_east_NO + A_env_east_UF

    # 3.4.2.2.3	暖冷房負荷モデルの空間ごとの方位ごと垂直外皮面積（北）
    A_env_north_MR = A_env_vert_MR * A_env_north_MR_ref / A_env_vert_MR_ref
    A_env_north_OR = A_env_vert_OR * A_env_north_OR_ref / A_env_vert_OR_ref
    A_env_north_NO = A_env_vert_NO * A_env_north_NO_ref / A_env_vert_NO_ref
    A_env_north_UF = A_env_vert_UF * A_env_north_UF_ref / A_env_vert_UF_ref if A_env_vert_UF_ref > 0.0 else 0.0
    A_env_north = A_env_north_MR  + A_env_north_OR + A_env_north_NO + A_env_north_UF

    # 3.4.2.3.1	暖冷房負荷モデルの外気に接する総外皮面積の推定
    A_env_ex_min = A_env_south_MR + A_env_south_OR + A_env_south_NO \
                   + A_env_north_MR + A_env_north_OR + A_env_north_NO \
                   + A_env_bottom_MR + A_env_bottom_OR + A_env_bottom_NO
    r_env_ex_min = A_env_ex_min / A_env

    r_env_ex_max = 1.0

    if tatekata == "戸建住宅":
        # 戸建てでは、外気に接する外皮面積は総外皮面積と同じと考える。
        r_env_ex = 1.0
    elif tatekata == "共同住宅":
        # 共同住宅では、外気に接する外皮面積は総外皮面積のうち、南面と北面の外皮面積と下面の外皮面積が外気に接すると考える。
        r_env_ex = min(max(r_env_ex_min, 1.0 / (1.0 + math.exp(-9.10907512 * (ua - 1.05204145)))), r_env_ex_max)
    else:
        raise ValueError(tatekata)
    
    A_env_ex = A_env * r_env_ex

    # 3.4.2.3.2 暖冷房負荷モデルの空間ごと、方位ごとの外気に接する外皮面積の推定
    if tatekata == "戸建住宅":
        r_env_ex_MR_top = 1.0
        r_env_ex_OR_top = 1.0
        r_env_ex_NO_top = 1.0
        r_env_ex_MR_south = 1.0
        r_env_ex_OR_south = 1.0
        r_env_ex_NO_south = 1.0
        r_env_ex_UF_south = 1.0
        r_env_ex_MR_east = 1.0
        r_env_ex_OR_east = 1.0
        r_env_ex_NO_east = 1.0
        r_env_ex_UF_east = 1.0
        r_env_ex_MR_north = 1.0
        r_env_ex_OR_north = 1.0
        r_env_ex_NO_north = 1.0
        r_env_ex_UF_north = 1.0
        r_env_ex_MR_west = 1.0
        r_env_ex_OR_west = 1.0
        r_env_ex_NO_west = 1.0
        r_env_ex_UF_west = 1.0
        r_env_ex_MR_bottom = 1.0
        r_env_ex_OR_bottom = 1.0
        r_env_ex_NO_bottom = 1.0
        r_env_ex_UF_bottom = 1.0
    elif tatekata == "共同住宅":
        denom = A_env_top_MR + A_env_top_OR + A_env_top_NO \
                + A_env_east_MR + A_env_east_OR + A_env_east_NO \
                + A_env_west_MR + A_env_west_OR + A_env_west_NO
        r_env_ex_top = \
            max((A_env_ex - A_env_south_MR - A_env_south_OR - A_env_south_NO - A_env_north_MR - A_env_north_OR - A_env_north_NO - A_env_bottom_MR - A_env_bottom_OR - A_env_bottom_NO) \
                / (denom if denom > 0.0 else 1.0 ), 0.0)
        r_env_ex_bottom = 1.0
        r_env_ex_south = 1.0
        r_env_ex_north = 1.0
        r_env_ex_east = \
            max((A_env_ex - A_env_south_MR - A_env_south_OR - A_env_south_NO - A_env_north_MR - A_env_north_OR - A_env_north_NO) \
                / (denom if denom > 0.0 else 1.0), 0.0)
        r_env_ex_west = r_env_ex_east
        r_env_ex_MR_top = r_env_ex_top
        r_env_ex_OR_top = r_env_ex_top
        r_env_ex_NO_top = r_env_ex_top
        r_env_ex_UF_top = 0.0
        r_env_ex_MR_south = r_env_ex_south
        r_env_ex_OR_south = r_env_ex_south
        r_env_ex_NO_south = r_env_ex_south
        r_env_ex_UF_south = 0.0
        r_env_ex_MR_east = r_env_ex_east
        r_env_ex_OR_east = r_env_ex_east
        r_env_ex_NO_east = r_env_ex_east
        r_env_ex_UF_east = 0.0
        r_env_ex_MR_north = r_env_ex_north
        r_env_ex_OR_north = r_env_ex_north
        r_env_ex_NO_north = r_env_ex_north
        r_env_ex_UF_north = 0.0
        r_env_ex_MR_west = r_env_ex_west
        r_env_ex_OR_west = r_env_ex_west
        r_env_ex_NO_west = r_env_ex_west
        r_env_ex_UF_west = 0.0
        r_env_ex_MR_bottom = r_env_ex_bottom
        r_env_ex_OR_bottom = r_env_ex_bottom
        r_env_ex_NO_bottom = r_env_ex_bottom
        r_env_ex_UF_bottom = 0.0
    else:
        raise ValueError(tatekata)

    A_env_top_MR_ex = A_env_top_MR * r_env_ex_MR_top
    A_env_top_OR_ex = A_env_top_OR * r_env_ex_OR_top
    A_env_top_NO_ex = A_env_top_NO * r_env_ex_NO_top
    A_env_top_UF_ex = A_env_top_UF * r_env_ex_UF_top
    A_env_south_MR_ex = A_env_south_MR * r_env_ex_MR_south
    A_env_south_OR_ex = A_env_south_OR * r_env_ex_OR_south
    A_env_south_NO_ex = A_env_south_NO * r_env_ex_NO_south
    A_env_south_UF_ex = A_env_south_UF * r_env_ex_UF_south
    A_env_east_MR_ex = A_env_east_MR * r_env_ex_MR_east
    A_env_east_OR_ex = A_env_east_OR * r_env_ex_OR_east
    A_env_east_NO_ex = A_env_east_NO * r_env_ex_NO_east
    A_env_east_UF_ex = A_env_east_UF * r_env_ex_UF_east
    A_env_north_MR_ex = A_env_north_MR * r_env_ex_MR_north
    A_env_north_OR_ex = A_env_north_OR * r_env_ex_OR_north
    A_env_north_NO_ex = A_env_north_NO * r_env_ex_NO_north
    A_env_north_UF_ex = A_env_north_UF * r_env_ex_UF_north
    A_env_west_MR_ex = A_env_west_MR * r_env_ex_MR_west
    A_env_west_OR_ex = A_env_west_OR * r_env_ex_OR_west
    A_env_west_NO_ex = A_env_west_NO * r_env_ex_NO_west
    A_env_west_UF_ex = A_env_west_UF * r_env_ex_UF_west
    A_env_bottom_MR_ex = A_env_bottom_MR * r_env_ex_MR_bottom
    A_env_bottom_OR_ex = A_env_bottom_OR * r_env_ex_OR_bottom
    A_env_bottom_NO_ex = A_env_bottom_NO * r_env_ex_NO_bottom
    A_env_bottom_UF_ex = A_env_bottom_UF * r_env_ex_UF_bottom

    A_env_top_ex = A_env_top_MR_ex + A_env_top_OR_ex + A_env_top_NO_ex + A_env_top_UF_ex
    A_env_south_ex = A_env_south_MR_ex + A_env_south_OR_ex + A_env_south_NO_ex + A_env_south_UF_ex
    A

    # 3.4.2.4.1 暖冷房負荷モデル総開口面積の推定
    r_env_op = get_open_rate(eta_ac)
    A_env_op = A_env_ex * r_env_op

    # 3.4.2.4.2 暖冷房負荷モデルの空間ごとの開口部面積の推定
    sum_A_r_A_env_op_ref_A_r_ref = A_MR * A_env_op_MR_ref / A_MR_ref + A_OR * A_env_op_OR_ref / A_OR_ref + A_NO * A_env_op_NO_ref / A_NO_ref
    A_env_op_MR = A_env_op * A_MR * (A_env_op_MR_ref / A_MR_ref) / sum_A_r_A_env_op_ref_A_r_ref if sum_A_r_A_env_op_ref_A_r_ref > 0.0 else 0.0
    A_env_op_OR = A_env_op * A_OR * (A_env_op_OR_ref / A_OR_ref) / sum_A_r_A_env_op_ref_A_r_ref if sum_A_r_A_env_op_ref_A_r_ref > 0.0 else 0.0
    A_env_op_NO = A_env_op * A_NO * (A_env_op_NO_ref / A_NO_ref) / sum_A_r_A_env_op_ref_A_r_ref if sum_A_r_A_env_op_ref_A_r_ref > 0.0 else 0.0

    # 3.4.2.4.3 暖冷房負荷モデルの空間ごとの方位ごとの窓面積の推定
    A_env_win_south_MR = min(A_env_op_MR * A_env_win_south_MR_ref / A_env_op_MR_ref, A_env_south_MR_ex)
    A_env_win_north_MR = min(A_env_op_MR * A_env_win_north_MR_ref / A_env_op_MR_ref, A_env_north_MR_ex)
    A_env_win_east_MR = min(A_env_op_MR * A_env_win_east_MR_ref / A_env_op_MR_ref, A_env_east_MR_ex)
    A_env_win_west_MR = min(A_env_op_MR * A_env_win_west_MR_ref / A_env_op_MR_ref, A_env_west_MR_ex)
    A_env_win_south_OR = min(A_env_op_OR * A_env_win_south_OR_ref / A_env_op_OR_ref, A_env_south_OR_ex)
    A_env_win_north_OR = min(A_env_op_OR * A_env_win_north_OR_ref / A_env_op_OR_ref, A_env_north_OR_ex)
    A_env_win_east_OR = min(A_env_op_OR * A_env_win_east_OR_ref / A_env_op_OR_ref, A_env_east_OR_ex)
    A_env_win_west_OR = min(A_env_op_OR * A_env_win_west_OR_ref / A_env_op_OR_ref, A_env_west_OR_ex)
    A_env_win_south_NO = min(A_env_op_NO * A_env_win_south_NO_ref / A_env_op_NO_ref, A_env_south_NO_ex)
    A_env_win_north_NO = min(A_env_op_NO * A_env_win_north_NO_ref / A_env_op_NO_ref, A_env_north_NO_ex)
    A_env_win_east_NO = min(A_env_op_NO * A_env_win_east_NO_ref / A_env_op_NO_ref, A_env_east_NO_ex)
    A_env_win_west_NO = min(A_env_op_NO * A_env_win_west_NO_ref / A_env_op_NO_ref, A_env_west_NO_ex)

    # 3.4.2.4.4 暖冷房負荷モデルの空間ごとの方位ごとのドア面積の推定
    A_env_door_north_MR = min(A_env_op_MR * A_env_door_north_MR_ref / A_env_op_MR_ref, A_env_north_MR_ex - A_env_win_north_MR)
    A_env_door_west_MR = min(A_env_op_MR * A_env_door_west_MR_ref / A_env_op_MR_ref, A_env_west_MR_ex - A_env_win_west_MR)
    A_env_door_south_MR = 0.0
    A_env_door_east_MR = 0.0
    A_env_door_north_OR = min(A_env_op_OR * A_env_door_north_OR_ref / A_env_op_OR_ref, A_env_north_OR_ex - A_env_win_north_OR)
    A_env_door_west_OR = min(A_env_op_OR * A_env_door_west_OR_ref / A_env_op_OR_ref, A_env_west_OR_ex - A_env_win_west_OR)
    A_env_door_south_OR = 0.0
    A_env_door_east_OR = 0.0
    A_env_door_north_NO = min(A_env_op_NO * A_env_door_north_NO_ref / A_env_op_NO_ref, A_env_north_NO_ex - A_env_win_north_NO)
    A_env_door_west_NO = min(A_env_op_NO * A_env_door_west_NO_ref / A_env_op_NO_ref, A_env_west_NO_ex - A_env_win_west_NO)
    A_env_door_south_NO = 0.0
    A_env_door_east_NO = 0.0

    # 3.4.2.5 暖冷房負荷モデルの空間ごと、方位ごとの外気に接する外壁等の面積の推定
    A_env_wall_south_MR_ex = max(A_env_south_MR_ex - A_env_win_south_MR - A_env_door_south_MR, 0.0)
    A_env_wall_east_MR_ex = max(A_env_east_MR_ex - A_env_win_east_MR - A_env_door_east_MR, 0.0)
    A_env_wall_north_MR_ex = max(A_env_north_MR_ex - A_env_win_north_MR - A_env_door_north_MR, 0.0)
    A_env_wall_west_MR_ex = max(A_env_west_MR_ex - A_env_win_west_MR - A_env_door_west_MR, 0.0)
    A_env_wall_south_OR_ex = max(A_env_south_OR_ex - A_env_win_south_OR - A_env_door_south_OR, 0.0)
    A_env_wall_east_OR_ex = max(A_env_east_OR_ex - A_env_win_east_OR - A_env_door_east_OR, 0.0)
    A_env_wall_north_OR_ex = max(A_env_north_OR_ex - A_env_win_north_OR - A_env_door_north_OR, 0.0)
    A_env_wall_west_OR_ex = max(A_env_west_OR_ex - A_env_win_west_OR - A_env_door_west_OR, 0.0)
    A_env_wall_south_NO_ex = max(A_env_south_NO_ex - A_env_win_south_NO - A_env_door_south_NO, 0.0)
    A_env_wall_east_NO_ex = max(A_env_east_NO_ex - A_env_win_east_NO - A_env_door_east_NO, 0.0)
    A_env_wall_north_NO_ex = max(A_env_north_NO_ex - A_env_win_north_NO - A_env_door_north_NO, 0.0)
    A_env_wall_west_NO_ex = max(A_env_west_NO_ex - A_env_win_west_NO - A_env_door_west_NO, 0.0)
    A_env_wall_bottom_MR_ex = A_env_bottom_MR
    A_env_wall_bottom_OR_ex = A_env_bottom_OR
    A_env_wall_bottom_NO_ex = A_env_bottom_NO
    A_env_wall_bottom_UF_ex = A_env_bottom_UF
    A_env_wall_top_MR_ex = A_env_top_MR
    A_env_wall_top_OR_ex = A_env_top_OR
    A_env_wall_top_NO_ex = A_env_top_NO
    A_env_wall_top_UF_ex = A_env_top_UF

    # 3.4.2.6 暖冷房負荷モデルの空間ごと、方位ごとの外気に接しない外壁等の面積の割合の推定
    A_env_south_MR_in = max(A_env_south_MR - A_env_south_MR_ex, 0.0)
    A_env_east_MR_in = max(A_env_east_MR - A_env_east_MR_ex, 0.0)
    A_env_north_MR_in = max(A_env_north_MR - A_env_north_MR_ex, 0.0)
    A_env_west_MR_in = max(A_env_west_MR - A_env_west_MR_ex, 0.0)
    A_env_south_OR_in = max(A_env_south_OR - A_env_south_OR_ex, 0.0)
    A_env_east_OR_in = max(A_env_east_OR - A_env_east_OR_ex, 0.0)
    A_env_north_OR_in = max(A_env_north_OR - A_env_north_OR_ex, 0.0)
    A_env_west_OR_in = max(A_env_west_OR - A_env_west_OR_ex, 0.0)
    A_env_south_NO_in = max(A_env_south_NO - A_env_south_NO_ex, 0.0)
    A_env_east_NO_in = max(A_env_east_NO - A_env_east_NO_ex, 0.0)
    A_env_north_NO_in = max(A_env_north_NO - A_env_north_NO_ex, 0.0)
    A_env_west_NO_in = max(A_env_west_NO - A_env_west_NO_ex, 0.0)
    A_env_top_MR_in = max(A_env_top_MR - A_env_top_MR_ex, 0.0)
    A_env_top_OR_in = max(A_env_top_OR - A_env_top_OR_ex, 0.0)
    A_env_top_NO_in = max(A_env_top_NO - A_env_top_NO_ex, 0.0)
    A_env_top_UF_in = max(A_env_top_UF - A_env_top_UF_ex, 0.0)
    A_env_bottom_MR_in = max(A_env_bottom_MR - A_env_bottom_MR_ex, 0.0)
    A_env_bottom_OR_in = max(A_env_bottom_OR - A_env_bottom_OR_ex, 0.0)
    A_env_bottom_NO_in = max(A_env_bottom_NO - A_env_bottom_NO_ex, 0.0)
    A_env_bottom_UF_in = max(A_env_bottom_UF - A_env_bottom_UF_ex, 0.0)

    # 3.4.7.1 暖冷房負荷モデルの間仕切り面積の推定
    A_part_vert_MR_OR = A_part_MR_OR_ref * (A_env_vert_MR + A_env_vert_OR) / (A_env_vert_MR_ref + A_env_vert_OR_ref)
    A_part_vert_MR_NO = A_part_MR_NO_ref * (A_env_vert_MR + A_env_vert_NO) / (A_env_vert_MR_ref + A_env_vert_NO_ref)
    A_part_vert_OR_NO = A_part_OR_NO_ref * (A_env_vert_OR + A_env_vert_NO) / (A_env_vert_OR_ref + A_env_vert_NO_ref)

    # 3.4.7.2 暖冷房負荷モデルの空間ごとの内壁床面積の推定
    A_part_bottom_MR = max(A_MR - A_env_bottom_MR, 0.0)
    A_part_bottom_OR = max(A_OR - A_env_bottom_OR, 0.0)
    A_part_bottom_NO = max(A_NO - A_env_bottom_NO, 0.0)
    A_part_bottom_UF = max(A_UF - A_env_bottom_UF, 0.0)
    # 3.3.10.31	暖冷房負荷モデルの南向きまたは北向きの外皮のみが外気に接する場合の総外皮に占める外気に接する外皮の面積の割合(共同住宅のみ計算)
    if tatekata == "共同住宅":
        # 北面と南面のみ外気に接する場合が、外気に接する外皮面積が参照になる。
        A_env_ex_min = A_env_south + A_env_north
        r_env_ex_min = A_env_ex_min / A_env

    # 3.3.10.30	暖冷房負荷モデルの総外皮に占める外気に接する外皮面積の割合
    if tatekata == "戸建住宅":
        # 戸建てでは固定で考える
        r_env_ex = 1.0
    elif tatekata == "共同住宅":
        # UA値から外気に接する外皮面積の割合を推定
        r_dash_env_ex = 1.0 / (1.0 + math.exp(-9.10907512 * (ua - 1.05204145)))
        r_env_ex = min(max(r_env_ex_min, r_dash_env_ex), r_env_ex_max)
    else:
        raise ValueError(tatekata)

    # 3.3.10.26	暖冷房負荷モデルの間仕切りの面積
    # 間仕切り面積 = 参照住戸の間仕切り面積 * モデル住戸の屋根面積 / 参照住戸の屋根面積
    if has_vertical_internal == "有":
        A_part_MR_OR = A_part_MR_OR_ref * (A_env_vert_MR + A_env_vert_OR) / (A_env_vert_MR_ref + A_env_vert_OR_ref)
        A_part_MR_NO = A_part_MR_NO_ref * (A_env_vert_MR + A_env_vert_NO) / (A_env_vert_MR_ref + A_env_vert_NO_ref)
        A_part_OR_NO = A_part_OR_NO_ref * (A_env_vert_OR + A_env_vert_NO) / (A_env_vert_OR_ref + A_env_vert_NO_ref)
    elif has_vertical_internal == "無":
        A_part_MR_OR = 0.0
        A_part_MR_NO = 0.0
        A_part_OR_NO = 0.0
    else:
        raise ValueError(has_vertical_internal)

    # 暖冷房負荷モデル・参照住戸における、主たる居室・その他の居室・非居室の内壁床の面積の合計
    A_part_bottom_MR = max(A_MR - A_env_bottom_MR, 0.0)
    A_part_bottom_OR = max(A_OR - A_env_bottom_OR, 0.0)
    A_part_bottom_NO = max(A_NO - A_env_bottom_NO, 0.0)
    A_part_bottom_MR_ref = A_part_bottom_MR_MR_ref + A_part_bottom_MR_OR_ref + A_part_bottom_MR_NO_ref + A_part_bottom_MR_UF_ref
    A_part_bottom_OR_ref = A_part_bottom_OR_MR_ref + A_part_bottom_OR_OR_ref + A_part_bottom_OR_NO_ref + A_part_bottom_OR_UF_ref
    A_part_bottom_NO_ref = A_part_bottom_NO_MR_ref + A_part_bottom_NO_OR_ref + A_part_bottom_NO_NO_ref + A_part_bottom_NO_UF_ref

    # 暖冷房負荷モデルにおける各居室の内壁床の面積を、隣接する居室の違いに応じて割り振る際の比率は、参照住戸における比率と同じとする。
    A_part_bottom_MR_MR = A_part_bottom_MR * A_part_bottom_MR_MR_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_MR_OR = A_part_bottom_MR * A_part_bottom_MR_OR_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_MR_NO = A_part_bottom_MR * A_part_bottom_MR_NO_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_MR_UF = A_part_bottom_MR * A_part_bottom_MR_UF_ref / A_part_bottom_MR_ref if A_part_bottom_MR_ref > 0.0 else 0.0
    A_part_bottom_OR_MR = A_part_bottom_OR * A_part_bottom_OR_MR_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_OR_OR = A_part_bottom_OR * A_part_bottom_OR_OR_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_OR_NO = A_part_bottom_OR * A_part_bottom_OR_NO_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_OR_UF = A_part_bottom_OR * A_part_bottom_OR_UF_ref / A_part_bottom_OR_ref if A_part_bottom_OR_ref > 0.0 else 0.0
    A_part_bottom_NO_MR = A_part_bottom_NO * A_part_bottom_NO_MR_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0
    A_part_bottom_NO_OR = A_part_bottom_NO * A_part_bottom_NO_OR_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0
    A_part_bottom_NO_NO = A_part_bottom_NO * A_part_bottom_NO_NO_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0
    A_part_bottom_NO_UF = A_part_bottom_NO * A_part_bottom_NO_UF_ref / A_part_bottom_NO_ref if A_part_bottom_NO_ref > 0.0 else 0.0

    # 3.4.3 居室の容積 (室内高=2.4m想定)
    V_room_MR = 2.4 * A_MR
    V_room_OR = 2.4 * A_OR
    V_room_NO = 2.4 * A_NO
    V_room_UF = 0.4 * A_env_bottom_UF

    # 表13、14　壁体構成
    if tatekata == "共同住宅":
        u_roof_noins_ex = 3.653
        r_roof_noins_ex = 0.274
        u_wall_noins_ex = 4.267
        r_wall_noins_ex = 0.234
        u_floor_in = 2.540
        r_floor_in = 0.394
        u_wall_UF_noins_ex = 0.0
        r_wall_UF_noins_ex = 0.0
        u_floor_noins_ex = 0.0
        u_win_noins_ex = 6.51
        u_door_noins_ex = 6.51
    elif tatekata == "戸建住宅":
        u_roof_noins_ex = 4.481
        r_roof_noins_ex = 0.227
        u_wall_noins_ex = 2.240
        r_wall_noins_ex = 0.431
        u_floor_noins_ex = 2.667
        u_floor_in = 0.0
        r_floor_noins_ex = 0.375
        u_wall_UF_noins_ex = 2.667
        r_wall_UF_noins_ex = 0.225
        u_win_noins_ex = 6.51
        u_door_noins_ex = 6.51
    else:
        raise ValueError(tatekata)
    
    # 3.4.4	暖冷房負荷モデルの熱貫流率の推定
    q_model_noins = u_roof_noins_ex * A_env_top_ex \
                    + u_wall_noins_ex * A_env_wall_vert_ex \
                    + u_floor_noins_ex * (A_env_bottom_ex - A_env_bottom_UF_ex) \
                    + u_win_noins_ex * A_env_win \
                    + u_door_noins_ex * A_env_door

    # 3.3.8.10 温度差係数
    H_os = 1.0      # 外気に接する部位の温度差係数
    H_is = 0.0      # 外気に接しない部位および隣接住戸と接する内壁の温度差係数
    H_floor = 0.7   # 外気に通じる床裏の温度差係数
    
    # 3.3.8.9	暖冷房負荷モデルの部位の種類ごとの面積


    # 3.3.8.8	暖冷房負荷モデルが無断熱であると仮定した場合の熱貫流率
    # 最悪U [W/m2K]
    U_roof_ex_max = get_wall_transfer_rate(R_fix=R_fix_roof, insulation_thickness=0.0, H=H_os, Rs=Rs_roof_ex_ref)
    U_wall_ex_max = get_wall_transfer_rate(R_fix=R_fix_wall, insulation_thickness=0.0, H=H_os, Rs=Rs_wall_ex_ref)
    U_floor_ex_max = get_wall_transfer_rate(R_fix=R_fix_floor, insulation_thickness=0.0, H=H_floor, Rs=Rs_floor_ex_ref)
    U_base_max = get_wall_transfer_rate(R_fix=R_fix_base, insulation_thickness=0.0, H=H_os, Rs=Rs_base_ref)  # 無断熱相当の基礎壁の熱貫流率
    Psi_base_max = 0.99    # 無断熱相当の基礎の線熱貫流率
    U_win_max = 10      #窓の最悪U追加
    U_door_max = 10     #ドアの最悪U追加

    # 3.3.8.7 暖冷房負荷モデルが無断熱であると仮定した場合の熱損失量
    # 最悪q [W/K] = 最悪U * 面積
    q_roof_max  = U_roof_ex_max  * A_env_top_ex
    q_wall_max  = U_wall_ex_max  * A_env_wall_vert_ex
    q_floor_max = U_floor_ex_max * (A_env_bottom_ex - A_env_bottom_UF_ex)
    q_base_max = U_base_max * A_env_vert_UF
    q_win_max   = U_win_max  * A_env_win
    q_door_max  = U_door_max * A_env_door
    q_max = q_roof_max + q_wall_max + q_floor_max + q_base_max + q_win_max + q_door_max

    # 3.3.8.6 参照住戸の部位の熱貫流率
    # 基準U [W/m2K]
    U_roof_ex_ref = get_wall_transfer_rate(R_fix=R_fix_roof, insulation_thickness=t_roof_ex_ins_ref, H=H_os, Rs=Rs_roof_ex_ref)
    U_wall_ex_ref = get_wall_transfer_rate(R_fix=R_fix_wall, insulation_thickness=t_wall_ex_ins_ref, H=H_os, Rs=Rs_wall_ex_ref)
    U_floor_ex_ref = get_wall_transfer_rate(R_fix=R_fix_floor, insulation_thickness=t_floor_ex_ins_ref, H=H_floor, Rs=Rs_floor_ex_ref)
    U_base_ref = get_wall_transfer_rate(R_fix=R_fix_base, insulation_thickness=t_base_ins_ref, H=H_os, Rs=Rs_base_ref)
    # Psi_base_ref = 0.99    # 基礎の線熱貫流率

    # 3.3.8.5 暖冷房負荷モデルの熱貫流率が参照住戸と同等であると仮定した場合の熱損失量
    # 基準q [W/K] = 基準U * 面積
    q_roof_basis = U_roof_ex_ref * A_env_top_ex
    q_wall_basis = U_wall_ex_ref * A_env_wall_vert_ex
    q_floor_basis = U_floor_ex_ref *  (A_env_bottom_ex - A_env_bottom_UF_ex)
    q_base_basis = U_base_ref * A_env_vert_UF
    q_win_basis = U_win_ref * A_env_win
    q_door_basis = U_door_ref * A_env_door
    q_basis = q_roof_basis + q_wall_basis + q_floor_basis + q_base_basis + q_win_basis + q_door_basis

    q_roof_margin = q_roof_max - q_roof_basis
    q_wall_margin = q_wall_max - q_wall_basis
    q_floor_margin = q_floor_max - q_floor_basis
    q_base_margin = q_base_max - q_base_basis
    q_win_margin = q_win_max - q_win_basis
    q_door_margin = q_door_max - q_door_basis
    q_margin = q_roof_margin + q_wall_margin + q_floor_margin + q_base_margin + q_win_margin + q_door_margin

    # 差分q = 入力された条件におけるq値 - 基準qの合計
    q_diff = A_env * ua - q_basis

    r_q_margin_plus = q_diff / q_margin
    r_q_margin_minus =  - q_diff / q_basis 

    # 3.3.8.4 設計住戸の熱損失量に合わせるための調整量
    def f_offset(x, y):

        # q値の割り当て [W/K] = 基準q値 or (最悪q - 基準q値)
        if q_diff <= 0:
            return -y * r_q_margin_minus
        else:
            # x-y: q余裕
            return (x - y) * r_q_margin_plus

    # 3.3.8.3 暖冷房負荷モデルの部位の熱損失量 [W/K]
    q_roof_ex  = q_roof_basis  + f_offset(q_roof_max,  q_roof_basis)
    q_wall_ex  = q_wall_basis  + f_offset(q_wall_max,  q_wall_basis)
    q_floor_ex = q_floor_basis + f_offset(q_floor_max, q_floor_basis)
    q_base     = q_base_basis  + f_offset(q_base_max,  q_base_basis)
    q_win      = q_win_basis   + f_offset(q_win_max,   q_win_basis)
    q_door     = q_door_basis  + f_offset(q_door_max,  q_door_basis)

    # 3.3.8.2 暖冷房負荷モデルの部位の熱貫流率(負荷計算への入力) [W/㎡K]
    U_roof_ex  = q_roof_ex  / A_env_top_ex if A_env_top_ex > 0.0 else 0.0
    U_wall_ex  = q_wall_ex  / A_env_wall_vert_ex if A_env_wall_vert_ex > 0.0 else 0.0
    U_floor_ex = q_floor_ex / A_env_bottom_ex if A_env_bottom_ex > 0.0 else 0.0
    U_base     = q_base     / A_env_vert_UF if A_env_vert_UF > 0.0 else 0.0
    U_win      = q_win      / A_env_win if A_env_win > 0.0 else 0.0
    U_door     = q_door     / A_env_door if A_env_door > 0.0 else 0.0
    # NOTE: 基礎のあたりの処理が怪しい

    # 3.3.8.1 暖冷房負荷モデルの断熱材の熱抵抗(負荷計算への入力) [m]
    R_roof_ex_ins  = get_insulation_registance(U_roof_ex, R_fix_roof, H_os, Rs_roof_ex_ref)
    R_wall_ex_ins  = get_insulation_registance(U_wall_ex, R_fix_wall, H_os, Rs_wall_ex_ref)
    R_floor_ex_ins = get_insulation_registance(U_floor_ex, R_fix_floor, H_floor, Rs_floor_ex_ref)
    R_base_ins = get_insulation_registance(U_base, R_fix_base, H_os, Rs_base_ref)

    # --------------------------------
    # ----- 3.3.9 窓の日射熱取得率 -----
    # --------------------------------

    # 3.3.9.8 暖冷房期間
    DD_H, DD_C = get_master_days(region)

    # 3.3.9.7 方位係数
    neu_c, neu_h = get_neu_avg(region)

    # 3.3.9.6 設計住戸の年間平均日射熱取得率
    # --------------------------------------------------------
    # 年間平均日射熱取得率は冷房期平均日射熱取得率および暖房期平均日射熱取得率を冷房期間および
    # 暖房期間の日数で案分して求める。
    eta_avg = (eta_ac * DD_C + eta_ah * DD_H) / (DD_H + DD_C)

    # 3.3.9.5 設計住戸の外皮全体の日射熱取得 [W/(W/㎡)]
    # --------------------------------------------------------
    # 外皮全体の日射熱取得量は年間平均日射熱取得率と外皮の面積の合計から求められる。
    m = eta_avg / 100 * A_env

    # 3.3.9.4 暖冷房負荷モデルの冷房期の不透明部位の日射熱取得量と暖房期の不透明部位の日射熱取得量
    # 窓以外の日射熱取得量は熱損失率から計算される  [W/(W・㎡)]
    # NOTE: 戸建住宅に対応するため、基礎壁からの日射熱取得の計算式が追加されている。
    m_opaq_C = \
        A_env_top_ex * U_roof_ex * neu_c[0] * 0.034 \
        + A_env_wall_south_ex * U_wall_ex * neu_c[5] * 0.034 \
        + A_env_wall_east_ex * U_wall_ex * neu_c[3] * 0.034 \
        + A_env_wall_north_ex * U_wall_ex * neu_c[1] * 0.034 \
        + A_env_wall_west_ex * U_wall_ex * neu_c[7] * 0.034 \
        + A_env_south_UF_ex * U_base * neu_c[5] * 0.034 \
        + A_env_east_UF_ex * U_base * neu_c[3] * 0.034 \
        + A_env_north_UF_ex * U_base * neu_c[1] * 0.034 \
        + A_env_west_UF_ex * U_base * neu_c[7] * 0.034 \
        + A_env_door_north  * U_door * neu_c[1] * 0.034 \
        + A_env_door_west  * U_door * neu_c[7] * 0.034
    m_opaq_H = \
        A_env_top_ex * U_roof_ex * neu_h[0] * 0.034 \
        + A_env_wall_south_ex * U_wall_ex * neu_h[5] * 0.034 \
        + A_env_wall_east_ex * U_wall_ex * neu_h[3] * 0.034 \
        + A_env_wall_north_ex * U_wall_ex * neu_h[1] * 0.034 \
        + A_env_wall_west_ex * U_wall_ex * neu_h[7] * 0.034 \
        + A_env_south_UF_ex * U_base * neu_h[5] * 0.034 \
        + A_env_east_UF_ex * U_base * neu_h[3] * 0.034 \
        + A_env_north_UF_ex * U_base * neu_h[1] * 0.034 \
        + A_env_west_UF_ex * U_base * neu_h[7] * 0.034 \
        + A_env_door_north * U_door * neu_h[1] * 0.034 \
        + A_env_door_west * U_door * neu_h[7] * 0.034
    
    # 3.3.9.3 暖冷房負荷モデルの不透明部位の日射熱取得量
    # ---------------------------------------------------------------
    # 不透明部位の日射熱取得量は冷房期の不透明部位の日射熱取得量と暖房期の不透明部位の日射熱取得量を
    # 冷房期間および暖房期間の日数で案分して求める。
    m_opaq = (m_opaq_C * DD_C + m_opaq_H * DD_H) / (DD_C + DD_H)

    # 3.3.9.2 暖冷房負荷モデルの透明部位(窓)の日射熱取得量 [W/(W/㎡)]
    # ---------------------------------------------------------------
    # 透明部位（窓）の日射熱取得量は、外皮全体の日射熱取得量から不透明部位の日射熱取得量を減じること
    # で求める。
    m_tran = m - m_opaq

    # 3.3.9.1 暖冷房負荷モデルの窓の日射熱取得率(負荷計算への入力) [(W/㎡)/(W/㎡)]
    # NOTE: 日射熱取得率の按分方法がExcelと異なる
    eta_win = calc_eta_win(m_tran, (A_env_win_south, A_env_win_east, A_env_win_north, A_env_win_west), neu_c, neu_h, DD_C, DD_H)

    # ------------------------------------------------------------------------

    print("計算条件")
    print("-------------------------------------------------")
    print("地域区分: {}".format(region))
    print("延床面積: {} [㎡]".format(total_floor_area))
    print(" 主たる居室: {} [㎡]".format(main_floor_area))
    print(" その他居室: {} [㎡]".format(other_floor_area))
    print("外皮総面積: {} [㎡]".format(A_env))
    print("外皮平均熱貫流率: {} [W/K]".format(ua))
    print("暖房期平均日射熱取得率: {}".format(eta_ah))
    print("冷房期平均日射熱取得率: {}".format(eta_ac))
    print("")
    print("計算結果")
    print("-------------------------------------------------")
    print("外皮面積")
    print("  主たる居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_MR, A_env_east_MR, A_env_north_MR, A_env_west_MR))
    print("  その他居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_OR, A_env_east_OR, A_env_north_OR, A_env_west_OR))
    print("  非居室:     {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_NO, A_env_east_NO, A_env_north_NO, A_env_west_NO))
    print("外気に接する屋根の面積")
    print("  主たる居室: {:3.2f} [㎡]".format(A_env_top_MR_ex))
    print("  その他居室: {:3.2f} [㎡]".format(A_env_top_OR_ex))
    print("  非居室:     {:3.2f} [㎡]".format(A_env_top_NO_ex))
    print("外気に接する外壁・基礎壁の面積")
    print("  主たる居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_wall_south_ex_MR, A_env_wall_east_ex_MR, A_env_wall_north_ex_MR, A_env_wall_west_ex_MR))
    print("  その他居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_wall_south_ex_OR, A_env_wall_east_ex_OR, A_env_wall_north_ex_OR, A_env_wall_west_ex_OR))
    print("  非居室:     {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_wall_south_ex_NO, A_env_wall_east_ex_NO, A_env_wall_north_ex_NO, A_env_wall_west_ex_NO))
    print("  床下空間:   {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_south_UF_ex, A_env_east_UF_ex, A_env_north_UF_ex, A_env_west_UF_ex))
    print("窓面積")
    print("  主たる居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_win_south_MR, A_env_win_east_MR, A_env_win_north_MR, A_env_win_west_MR))
    print("  その他居室: {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_win_south_OR, A_env_win_east_OR, A_env_win_north_OR, A_env_win_west_OR))
    print("  非居室:     {:3.2f} {:3.2f} {:3.2f} {:3.2f} [㎡]".format(A_env_win_south_NO, A_env_win_east_NO, A_env_win_north_NO, A_env_win_west_NO))
    print("ドア面積")
    print("  主たる居室: --- --- {:3.2f} {:3.2f} [㎡]".format(A_env_door_north_MR, A_env_door_west_MR))
    print("  その他居室: --- --- {:3.2f} {:3.2f} [㎡]".format(A_env_door_north_OR, A_env_door_west_OR))
    print("  非居室:     --- --- {:3.2f} {:3.2f} [㎡]".format(A_env_door_north_NO, A_env_door_west_NO))
    print("間仕切り")
    print("  主居室 - その他居室: {:3.2f} [㎡]".format(A_part_MR_OR))
    print("  主居室 - 非居室: {:3.2f} [㎡]".format(A_part_MR_NO))
    print("  その他居室 - 非居室: {:3.2f} [㎡]".format(A_part_OR_NO))
    print("内壁床")
    print("  主居室 - 主居室: {:3.2f} [㎡]".format(A_part_bottom_MR_MR))
    print("  主居室 - その他居室: {:3.2f} [㎡]".format(A_part_bottom_MR_OR))
    print("  主居室 - 非居室: {:3.2f} [㎡]".format(A_part_bottom_MR_NO))
    print("  主居室 - 床下空間: {:3.2f} [㎡]".format(A_part_bottom_MR_UF))
    print("  その他居室 - 主居室: {:3.2f} [㎡]".format(A_part_bottom_OR_MR))
    print("  その他居室 - その他居室: {:3.2f} [㎡]".format(A_part_bottom_OR_OR))
    print("  その他居室 - 非居室: {:3.2f} [㎡]".format(A_part_bottom_OR_NO))
    print("  その他居室 - 床下空間: {:3.2f} [㎡]".format(A_part_bottom_OR_UF))
    print("  非居室 - 主居室: {:3.2f} [㎡]".format(A_part_bottom_NO_MR))
    print("  非居室 - その他居室: {:3.2f} [㎡]".format(A_part_bottom_NO_OR))
    print("  非居室 - 非居室: {:3.2f} [㎡]".format(A_part_bottom_NO_NO))
    print("  非居室 - 床下空間: {:3.2f} [㎡]".format(A_part_bottom_NO_UF))
    print("熱貫流率")
    print("  外気に接する屋根: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_roof_ex, R_roof_ex_ins))
    print("  外気に接する外壁等: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_wall_ex, R_wall_ex_ins))
    print("  外気に接する床下: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_floor_ex, R_floor_ex_ins))
    print("  外気に接する基礎壁: {:3.2f} [W/㎡K] (断熱材熱抵抗= {:3.2f} [㎡K/W])".format(U_base, R_base_ins))
    print("  窓: {:3.2f} [W/㎡K]".format(U_win))
    print("  ドア: {:3.2f} [W/㎡K]".format(U_door))
    print("窓の日射熱取得率: {:3.2f} [(W/㎡)/(W/㎡)]".format(eta_win))

    # 入力Excel のテンプレートファイルを読み込み
    template_xlsx = _get_template_xlsx(tatekata, structure)

    # 入力シート内の各種変数名を変換する辞書を取得
    varname_mapper = {
        # rooms 
        # 居室の面積
        '#MR_A': A_MR,
        '#OR_A': A_OR,
        '#NO_A': A_NO,
        '#UF_A': A_env_bottom_UF,
        '#MR_VOL': V_room_MR,
        '#OR_VOL': V_room_OR,
        '#NO_VOL': V_room_NO,
        '#UF_VOL': V_room_UF,
        '#MR_VENT': V_vent_MR,
        '#OR_VENT': V_vent_OR,
        '#NO_VENT': V_vent_NO,
        '#UF_VENT': V_vent_UF,

        # external_general_parts
        # 垂直外壁と屋根・天井の面積
        # 外気に接する外壁等
        '#MR_A_C_EW': A_env_top_MR_ex,   #2F天井相当(屋根)
        '#OR_A_C_EW': A_env_top_OR_ex,   #2F天井相当(屋根)
        '#NO_A_C_EW': A_env_top_NO_ex,   #2F天井相当(屋根)
        '#MR_A_S_EW': A_env_wall_south_ex_MR,
        '#OR_A_S_EW': A_env_wall_south_ex_OR,
        '#NO_A_S_EW': A_env_wall_south_ex_NO,
        '#MR_A_E_EW': A_env_wall_east_ex_MR,
        '#OR_A_E_EW': A_env_wall_east_ex_OR,
        '#NO_A_E_EW': A_env_wall_east_ex_NO,
        '#MR_A_N_EW': A_env_wall_north_ex_MR,
        '#OR_A_N_EW': A_env_wall_north_ex_OR,
        '#NO_A_N_EW': A_env_wall_north_ex_NO,
        '#MR_A_W_EW': A_env_wall_west_ex_MR,
        '#OR_A_W_EW': A_env_wall_west_ex_OR,
        '#NO_A_W_EW': A_env_wall_west_ex_NO,
        '#MR_A_F_EW': A_env_bottom_MR_ex, #1F床相当(床断熱の場合)
        '#OR_A_F_EW': A_env_bottom_OR_ex, #1F床相当(床断熱の場合)
        '#NO_A_F_EW': A_env_bottom_NO_ex, #1F床相当(床断熱の場合)
        '#UF_A_S_EW': A_env_south_UF_ex, #床下の基礎(基礎断熱の場合)
        '#UF_A_E_EW': A_env_east_UF_ex,  #床下の基礎(基礎断熱の場合)
        '#UF_A_N_EW': A_env_north_UF_ex, #床下の基礎(基礎断熱の場合)
        '#UF_A_W_EW': A_env_west_UF_ex,  #床下の基礎(基礎断熱の場合)
        # 外気に接しない外壁等
        '#MR_A_C_PW': A_env_top_MR_in,   #2F天井相当(戸境壁)
        '#OR_A_C_PW': A_env_top_OR_in,   #2F天井相当(戸境壁)
        '#NO_A_C_PW': A_env_top_NO_in,   #2F天井相当(戸境壁)
        '#MR_A_E_PW': A_env_east_MR_in,  #戸境壁
        '#OR_A_E_PW': A_env_east_OR_in,  #戸境壁
        '#NO_A_E_PW': A_env_east_NO_in,  #戸境壁
        '#MR_A_W_PW': A_env_west_MR_in,  #戸境壁
        '#OR_A_W_PW': A_env_west_OR_in,  #戸境壁
        '#NO_A_W_PW': A_env_west_NO_in,  #戸境壁
        '#MR_A_F_PW': A_env_bottom_MR_in,    #1F床相当
        '#OR_A_F_PW': A_env_bottom_OR_in,     #1F床相当
        '#NO_A_F_PW': A_env_bottom_NO_in,     #1F床相当

        # external_opaque_parts
        # ドアの熱貫流率と面積
        '#DOOR_U': U_door,
        '#MR_DOOR_A_N': A_env_door_north_MR,
        '#OR_DOOR_A_N': A_env_door_north_OR,
        '#NO_DOOR_A_N': A_env_door_north_NO,
        '#MR_DOOR_A_W': A_env_door_west_MR,
        '#OR_DOOR_A_W': A_env_door_west_OR,
        '#NO_DOOR_A_W': A_env_door_west_NO,

        # external_transparent_parts
        # 窓の熱貫流率、日射熱取得率と面積
        '#WINDOW_U': U_win,
        '#WINDOW_ETA': eta_win,
        '#MR_A_WIN_S': A_env_win_south_MR,
        '#MR_A_WIN_E': A_env_win_east_MR,
        '#MR_A_WIN_N': A_env_win_north_MR,
        '#MR_A_WIN_W': A_env_win_west_MR,
        '#OR_A_WIN_S': A_env_win_south_OR,
        '#OR_A_WIN_E': A_env_win_east_OR,
        '#OR_A_WIN_N': A_env_win_north_OR,
        '#OR_A_WIN_W': A_env_win_west_OR,
        '#NO_A_WIN_S': A_env_win_south_NO,
        '#NO_A_WIN_E': A_env_win_east_NO,
        '#NO_A_WIN_N': A_env_win_north_NO,
        '#NO_A_WIN_W': A_env_win_west_NO,

        # partitions
        # 間仕切りの面積
        '#IN_MR_OR': A_part_MR_OR,
        '#IN_MR_NO': A_part_MR_NO,
        '#IN_OR_NO': A_part_OR_NO,

        # 内壁床の面積
        # NOTE: 同じ室用途どうしで接する内壁床の面積は、「温度差係数 0 の外気に接する床」として割り振る点に注意
        # ref. 2025年3月14日 暖冷房負荷評価枠組検討TG コアMTG ( https://youworks.atlassian.net/l/cp/AvAiqBoB )
        '#IN_MR_MR_F': A_part_bottom_MR_MR,
        '#IN_MR_OR_F': A_part_bottom_MR_OR,
        '#IN_MR_NO_F': A_part_bottom_MR_OR,
        '#IN_MR_UF_F': A_part_bottom_MR_UF,
        '#IN_OR_MR_F': A_part_bottom_OR_MR,
        '#IN_OR_OR_F': A_part_bottom_OR_OR,
        '#IN_OR_NO_F': A_part_bottom_OR_NO,
        '#IN_OR_UF_F': A_part_bottom_OR_UF,
        '#IN_NO_MR_F': A_part_bottom_NO_MR,
        '#IN_NO_OR_F': A_part_bottom_NO_OR,
        '#IN_NO_NO_F': A_part_bottom_NO_NO,
        '#IN_NO_UF_F': A_part_bottom_NO_UF,

        # layers
        # 屋根と垂直外壁の断熱材熱抵抗
        '#CEILING_R_INSULATION': R_roof_ex_ins,
        '#WALL_R_INSULATION': R_wall_ex_ins,
        '#FLOOR_R_INSULATION': R_floor_ex_ins,
        '#BASE_R_INSULATION': R_base_ins,
    }

    # 入力シート内の各種変数名を変換
    with pd.ExcelWriter(xlsx_path) as writer:
        for (sheet_name, df) in template_xlsx.items():
            df.replace(varname_mapper).to_excel(writer, sheet_name=sheet_name, index=False)



# --------------------------------------------------------------------------------
# 3.3.9.7 方位係数
# --------------------------------------------------------------------------------

def get_neu_avg(region: int) -> Tuple[
        Tuple[float,float,float,float,float,float,float,float,float,float],
        Tuple[float,float,float,float,float,float,float,float,float,float]
    ]:
    """指定された地域の区分の冷房期と暖房期の方位係数を返します。

    Args:
        region (int): 地域の区分1-8

    Returns:
        冷房期の方位係数と暖房期の方位係数をタプルで返します。
        それぞの方位係数は方位ごとに細分化されており、
        上面, 北, 北東, 東, 南東, 南, 南西, 西, 北西, 下面の順に格納されます。
    
    Notes:
        平成28年省エネルギー基準に準拠したエネルギー消費性能の評価に関する技術情報
        （住宅エネルギー消費性能の算定方法 第三章 暖冷房負荷と外皮性能 
        第二節 外皮性能 付録C 方位係数に定義される暖房期の方位係数と冷房期の方位係数) より
    """

    ## 方位係数
    ## index: 屋根・上面, 北, 北東, 東, 南東, 南, 南西, 西, 北西, 下面

    # 冷房期の方位係数
    neu_c = [
        [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
        [0.329, 0.341, 0.335, 0.322, 0.373, 0.341, 0.307, 0.325],
        [0.430, 0.412, 0.390, 0.426, 0.437, 0.431, 0.415, 0.414],
        [0.545, 0.503, 0.468, 0.518, 0.500, 0.512, 0.509, 0.515],
        [0.560, 0.527, 0.487, 0.508, 0.500, 0.498, 0.490, 0.528],
        [0.502, 0.507, 0.476, 0.437, 0.472, 0.434, 0.412, 0.480],
        [0.526, 0.548, 0.550, 0.481, 0.520, 0.491, 0.479, 0.517],
        [0.508, 0.529, 0.553, 0.481, 0.518, 0.504, 0.495, 0.505],
        [0.411, 0.428, 0.447, 0.401, 0.442, 0.427, 0.406, 0.411],
        [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    ]

    # 暖房期の方位係数
    neu_h = [
        [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 0.0],
        [0.260, 0.263, 0.284, 0.256, 0.238, 0.261, 0.227, 0.000],
        [0.333, 0.341, 0.348, 0.330, 0.310, 0.325, 0.281, 0.000],
        [0.564, 0.554, 0.540, 0.531, 0.568, 0.579, 0.543, 0.000],
        [0.823, 0.766, 0.751, 0.724, 0.846, 0.833, 0.843, 0.000],
        [0.935, 0.856, 0.851, 0.815, 0.983, 0.936, 1.023, 0.000],
        [0.790, 0.753, 0.750, 0.723, 0.815, 0.763, 0.848, 0.000],
        [0.535, 0.544, 0.542, 0.527, 0.538, 0.523, 0.548, 0.000],
        [0.325, 0.341, 0.351, 0.326, 0.297, 0.317, 0.284, 0.000],
        [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    ]

    return [neu_c[_][region - 1] for _ in range(10)], [neu_h[_][region - 1] for _ in range(10)]


# --------------------------------------------------------------------------------
# 3.3.9.8 暖冷房期間
# --------------------------------------------------------------------------------

def get_master_days(region: int) -> Tuple[int, int]:
    """指定された地域の区分の暖房期間と冷房期間の日数を返します。

    Args:
        region (int): 地域の区分1-8

    Returns:
        暖房期間の日数と冷房期間の日数をタプルで返します。
    
    Notes:
        平成28年省エネルギー基準に準拠したエネルギー消費性能の評価に関する技術情報
        （住宅エネルギー消費性能の算定方法 第十一章 その他)より
    """
    return (
        (257, 53),
        (252, 48),
        (244, 53),
        (242, 53),
        (218, 57),
        (169, 117),
        (122, 152),
        (0, 265),
    )[region - 1]


# --------------------------------------------------------------------------------
# 3.3.10.51	参照住戸の面積
# --------------------------------------------------------------------------------

def get_floor_area_ref(tatekata: str) -> Tuple[float,float,float]:
    """参照住戸の床面積

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        主たる居室、その他の居室、非居室の面積
    """
    if tatekata == "共同住宅":
        # 表5 参照住戸の床面積（共同住宅の場合）
        A_MR_ref, A_OR_ref, A_NO_ref = 24.23, 29.75, 16.02
    elif tatekata == "戸建住宅":
        # 表6 参照住戸の床面積（戸建住宅の場合）
        A_MR_ref, A_OR_ref, A_NO_ref = 29.81, 51.35, 38.93
    else:
        raise ValueError(tatekata)
    return A_MR_ref, A_OR_ref, A_NO_ref


def get_area_table_ref(tatekata: str) -> Tuple[Tuple[float]] :
    """参照住戸の面積を一覧表

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        2次元配列で面積を返します。
        1次元目: 主たる居室 / その他の居室 / 非居室 / 床下空間
        2次元目: 外皮(上面/北/東/南/西/下面) / 窓(北/東/南/西) / ドア(北/西)
        それぞれ、対応する記号は次の通りです。
        主たる居室,その他の居室,非居室,床下空間 = MR,OR,NO,UF
        外皮,窓,ドア = "env", "env,win", "env,door"
        上面,北,東,南,西,下面 = top,north,east,south,west,bottom
        例) 参照住戸の主たる居室の外皮の上面の面積の合計の記号は A_env_top_MR_ref
    """
    if tatekata == "共同住宅":
        # 表7 参照住戸の面積(共同住宅の場合)
        return (
            # 主たる居室 / その他の居室 / 非居室 / 床下空間
            (24.23, 29.75, 16.02, 0.00),    # 外皮-上面
            (0.00, 11.80, 4.16, 0.00),      # 外皮-北
            (0.00, 21.59, 8.05, 0.00),      # 外皮-東
            (9.52, 6.45, 0.00, 0.00),       # 外皮-南 
            (17.21, 10.06, 2.37, 0.00),     # 外皮-西
            (24.23, 29.75, 16.02, 0.00),    # 外皮-下面
            (0.00, 2.53, 0.00),             # 窓-北
            (0.00, 0.00, 0.00),             # 窓-東
            (4.52, 3.24, 0.00),             # 窓-南
            (0.00, 0.00, 0.00),             # 窓-西
            (0.00, 0.00, 1.76),             # ドア-北
            (0.00, 0.00, 0.00),             # ドア-西
        )
    
    elif tatekata == "戸建住宅":
        # 表8 参照住戸の面積(戸建住宅の場合)
        return (
            (0.00, 34.79, 17.40, 0.00),     # 外皮-上面
            (5.12, 6.77, 39.08, 2.81),      # 外皮-北
            (17.20, 8.74, 4.36, 3.28),      # 外皮-東
            (14.21, 29.26, 0.00, 2.91),     # 外皮-南 
            (0.00, 17.48, 13.20, 3.28),     # 外皮-西
            (29.81, 16.56, 21.53, 55.48),   # 外皮-下面
            (0.00, 4.59, 3.15),             # 窓-北
            (3.13, 0.66, 0.00),             # 窓-東
            (6.94, 8.17, 0.00),             # 窓-南
            (0.00, 0.99, 1.08),             # 窓-西
            (1.62, 0.00, 0.00),             # ドア-北
            (0.00, 0.00, 1.89),             # ドア-西
        )
    
    else:
        raise ValueError(tatekata)


def get_partition_table_ref(tatekata: str) -> Tuple[float]:
    """参照住戸の間仕切りの面積

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        以下の順番で参照住戸の間仕切り面積をリストで返します。
        1.主たる居室とその他の居室 A_part,MR->OR,ref
        2.主たる居室と非居室 A_part,MR->NO,ref
        3.その他の居室と非居室 A_part,OR->NO,ref
    """
    if tatekata == "共同住宅":
        # 表9 参照住戸の間仕切りの面積(共同住宅)
        return (12.53, 16.19, 40.51)
    elif tatekata == "戸建住宅":
        # 表10 参照住戸の間仕切りの面積(戸建住宅)
        return (8.64, 17.20, 29.51)
    else:
        raise ValueError(tatekata)


def get_partition_bottom_table_ref(tatekata: str) -> Tuple[float]:
    """参照住戸の内壁床の面積

    Args:
        tatekata (str): 住宅区の建て方を "戸建住宅" または "共同住宅" で指定する。

    Returns:
        以下の順番で参照住戸の内壁床の面積をリストで返します。
        1.主たる居室の床面積のうち、別の主たる居室に接する部分      A_part,bottom,MR->MR,ref
        2.主たる居室の床面積のうち、その他の居室に接する部分        A_part,bottom,MR->OR,ref
        3.主たる居室の床面積のうち、非居室に接する部分              A_part,bottom,MR->NO,ref
        4.主たる居室の床面積のうち、床下空間に接する部分            A_part,bottom,MR->UF,ref
        5.その他の居室の床面積のうち、主たる居室に接する部分        A_part,bottom,OR->MR,ref
        6.その他の居室の床面積のうち、別のその他の居室に接する部分  A_part,bottom,OR->OR,ref
        7.その他の居室の床面積のうち、非居室に接する部分            A_part,bottom,OR->NO,ref
        8.その他の居室の床面積のうち、床下空間に接する部分          A_part,bottom,OR->UF,ref
        9.非居室の床面積のうち、主たる居室に接する部分              A_part,bottom,NO->MR,ref
        10.非居室の床面積のうち、その他の居室に接する部分           A_part,bottom,NO->OR,ref
        11.非居室の床面積のうち、別の非居室に接する部分             A_part,bottom,NO->NO,ref
        12.非居室の床面積のうち、床下空間に接する部分               A_part,bottom,NO->UF,ref

    """
    if tatekata == "共同住宅":
        return (00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00, 00.00)
    elif tatekata == "戸建住宅":
        return (00.00, 00.00, 00.00, 29.81, 21.53, 13.25, 00.00, 16.56, 04.14, 00.00, 12.42, 21.53)
    else:
        raise ValueError(tatekata)


@functools.lru_cache
def _get_template_xlsx(tatekata, structure) -> dict[str, pd.DataFrame]:

    if tatekata == "戸建住宅":
        if structure == "床断熱" or structure == '床下断熱':
            template_xlsx_path = 'simple_input_excel_template_kodate_yukadan.xlsx'
        elif structure == "基礎断熱":
            template_xlsx_path = 'simple_input_excel_template_kodate_kisodan.xlsx'
        else:
            raise ValueError(structure)
    elif tatekata == "共同住宅":
            template_xlsx_path = 'simple_input_excel_template_kyodo.xlsx'
    else:
        raise ValueError(tatekata)
    
    if not os.path.isabs(template_xlsx_path):
        template_xlsx_path = os.path.join(os.path.dirname(__file__), 'templates', template_xlsx_path)

    return pd.read_excel(template_xlsx_path, sheet_name=None)


# =============================================================================
# Heat Load Calc JSON direct output API
# =============================================================================
# The original implementation generated an intermediate Excel workbook and then
# converted that workbook to the Heat Load Calc input JSON.  The functions below
# create the Dictionary/JSON structure directly.  The public function name
# `estimate` is kept for compatibility, but it now returns a dictionary that
# follows https://hc-energy.readthedocs.io/ja/latest/contents/02_02_spec_input.html
# instead of writing an Excel file.



_DIRECTION_TO_HC = {
    "south": "s", "east": "e", "north": "n", "west": "w",
    "s": "s", "e": "e", "n": "n", "w": "w",
    "top": "top", "bottom": "bottom",
}

_ROOM_DEFS = {
    "MR": {"id": 0, "name": "main_occupant_room", "label": "主たる居室", "schedule": "main_occupant_room"},
    "OR": {"id": 1, "name": "other_occupant_room", "label": "その他居室", "schedule": "other_occupant_room"},
    "NR": {"id": 2, "name": "non_occupant_room", "label": "非居室", "schedule": "non_occupant_room"},
    "UF": {"id": 3, "name": "under_floor_space", "label": "床下空間", "schedule": "non_occupant_room"},
}


def _r2(x: float, ndigits: int = 6) -> float:
    """Round tiny numerical noise to a JSON-friendly float."""
    if abs(x) < 1e-10:
        return 0.0
    return round(float(x), ndigits)


def _safe_div(num: float, den: float, default: float = 0.0) -> float:
    return num / den if abs(den) > 1e-12 else default


def _hc_h_c(direction: str):
    direction = _DIRECTION_TO_HC[direction]
    if direction in ["s", "sw", "w", "nw", "n", "ne", "e", "se"]:
        return 2.5
    if direction == "bottom":
        return 0.7
    if direction == "top":
        return 5.0
    raise ValueError(direction)


def _hc_is_floor(direction: str) -> bool:
    return _DIRECTION_TO_HC[direction] == "bottom"


def _hc_outside_heat_transfer_resistance(direction: str, temp_dif_coef: float = 1.0) -> float:
    direction = _DIRECTION_TO_HC[direction]
    is_parting = temp_dif_coef != 1.0
    if direction in ["s", "sw", "w", "nw", "n", "ne", "e", "se"]:
        return 0.04 if not is_parting else 0.11
    if direction == "top":
        return 0.04 if not is_parting else 0.09
    if direction == "bottom":
        return 0.15
    raise ValueError(direction)


def _layer(name: str, thermal_resistance: float, thermal_capacity: float = 0.0) -> Dict[str, float | str]:
    return {
        "name": name,
        "thermal_resistance": _r2(max(float(thermal_resistance), 1.0e-6)),
        "thermal_capacity": _r2(max(float(thermal_capacity), 0.0)),
    }


def _layers_from_u(name: str, u_value: float, direction: str, temp_dif_coef: float = 1.0) -> list[dict]:
    """Create one equivalent layer for an external_general_part.

    Heat Load Calc receives surface heat-transfer values separately, so the layer
    resistance here is the practical opaque-part resistance excluding the standard
    inside and outside surface resistances.  A small positive floor is used because
    zero-resistance layers can cause numerical issues in heat_load_calc.
    """
    if u_value <= 0.0 or temp_dif_coef == 0.0:
        return [_layer(f"{name}_adiabatic", 10.0, 165000.0)]
    r_si = 1.0 / _hc_h_c(direction)
    r_se = _hc_outside_heat_transfer_resistance(direction, temp_dif_coef)
    r_layer = max(1.0 / u_value - r_si - r_se, 0.001)
    return [_layer(name, r_layer, 165000.0)]


def _partition_layers(name: str = "internal_partition") -> list[dict]:
    return [_layer(name, 0.12, 90000.0)]


def _rac_equipments(equipment_id: int, room_id: int, floor_area: float) -> tuple[dict, dict]:
    q_rtd_c = 190.5 * floor_area + 45.6
    q_rtd_h = 1.2090 * q_rtd_c - 85.1
    q_max_c = max(0.8462 * q_rtd_c + 1205.9, q_rtd_c)
    q_max_h = max(1.7597 * q_max_c - 413.7, q_rtd_h)
    v_max_c = 11.076 * (q_rtd_c / 1000.0) ** 0.3432
    v_max_h = 11.076 * (q_rtd_h / 1000.0) ** 0.3432
    cooling = {
        "id": equipment_id,
        "name": f"cooling_equipment_{equipment_id}",
        "equipment_type": "rac",
        "property": {
            "space_id": room_id,
            "q_min": 500.0,
            "q_max": _r2(q_max_c),
            "v_min": _r2(v_max_c * 0.55),
            "v_max": _r2(v_max_c),
            "bf": 0.2,
        },
    }
    heating = {
        "id": equipment_id,
        "name": f"heating_equipment_{equipment_id}",
        "equipment_type": "rac",
        "property": {
            "space_id": room_id,
            "q_min": 500.0,
            "q_max": _r2(q_max_h),
            "v_min": _r2(v_max_h * 0.55),
            "v_max": _r2(v_max_h),
            "bf": 0.2,
        },
    }
    return cooling, heating


# -----------------------------------------------------------------------------
# Verification helpers
# -----------------------------------------------------------------------------
_NEU_INDEX_BY_DIRECTION = {
    "top": 0,
    "n": 1,
    "ne": 2,
    "e": 3,
    "se": 4,
    "s": 5,
    "sw": 6,
    "w": 7,
    "nw": 8,
}


def _boundary_layer_resistance(boundary: Dict[str, Any]) -> float:
    """Return the sum of layer thermal resistances for a boundary."""
    return sum(float(layer.get("thermal_resistance", 0.0)) for layer in boundary.get("layers", []))


def _boundary_u_value_for_check(boundary: Dict[str, Any]) -> float:
    """Return the boundary U-value before applying temp_dif_coef.

    external_general_part stores the layer resistance instead of a direct
    u_value.  The equivalent U-value is therefore reconstructed from the inside
    convective resistance, the layer resistance and the outside heat-transfer
    resistance.  external_opaque_part and external_transparent_part store
    u_value directly.
    """
    btype = boundary.get("boundary_type")
    if btype == "external_general_part":
        r_si = 1.0 / float(boundary["h_c"])
        r_layer = _boundary_layer_resistance(boundary)
        r_se = float(boundary.get("outside_heat_transfer_resistance", 0.0))
        r_total = r_si + r_layer + r_se
        return 1.0 / r_total if r_total > 0.0 else 0.0
    if btype in ["external_opaque_part", "external_transparent_part"]:
        return float(boundary.get("u_value", 0.0))
    return 0.0


def _make_check_item(input_value: float, model_value: float, *, unit: str = "",
                     abs_tol: float = 1.0e-3, rel_tol: float = 1.0e-3) -> Dict[str, Any]:
    """Build one comparison record for the verification report."""
    input_value = float(input_value)
    model_value = float(model_value)
    diff = model_value - input_value
    rel_diff = diff / input_value if abs(input_value) > 1.0e-12 else 0.0
    ok = abs(diff) <= max(abs_tol, abs(input_value) * rel_tol)
    return {
        "input": _r2(input_value),
        "model": _r2(model_value),
        "difference": _r2(diff),
        "relative_difference": _r2(rel_diff),
        "unit": unit,
        "ok": bool(ok),
    }


def calculate_model_characteristics(model: Dict[str, Any], region: int) -> Dict[str, float]:
    """Calculate checkable characteristics from the generated HLC input dict.

    The returned values are calculated only from ``rooms`` and ``boundaries`` in
    the generated model.  They can therefore be used to confirm whether the
    Dictionary/JSON actually reproduces the simple input values.
    """
    room_by_id = {int(r["id"]): r for r in model.get("rooms", [])}
    floor_by_room = {room.get("name", ""): float(room.get("floor_area", 0.0)) for room in model.get("rooms", [])}

    A_MR = floor_by_room.get("main_occupant_room", 0.0)
    A_OR = floor_by_room.get("other_occupant_room", 0.0)
    A_NR = floor_by_room.get("non_occupant_room", 0.0)
    A_total = A_MR + A_OR + A_NR

    envelope_types = {"external_general_part", "external_opaque_part", "external_transparent_part", "ground"}

    def is_artificial_same_use_bottom(boundary: Dict[str, Any]) -> bool:
        # Same-use inner floors are represented as adiabatic external floors only
        # for heat-load-calculation connectivity.  They are not part of the
        # user's input A_env, so exclude them from input-vs-model checks.
        return str(boundary.get("name", "")).endswith("_same_use_bottom_adiabatic")

    A_env_model = sum(
        float(b.get("area", 0.0))
        for b in model.get("boundaries", [])
        if b.get("boundary_type") in envelope_types and not is_artificial_same_use_bottom(b)
    )

    heat_loss = 0.0
    m_c = 0.0
    m_h = 0.0
    neu_c, neu_h = get_neu_avg(region)

    for b in model.get("boundaries", []):
        btype = b.get("boundary_type")
        if btype not in envelope_types or is_artificial_same_use_bottom(b):
            continue
        area = float(b.get("area", 0.0))
        temp_dif_coef = float(b.get("temp_dif_coef", 1.0))
        direction = b.get("direction")
        u_value = _boundary_u_value_for_check(b)
        heat_loss += area * u_value * temp_dif_coef

        # Average solar heat gain coefficient check.  Only sun-struck external
        # roofs/walls/doors/windows are counted.  Floors and ground surfaces are
        # intentionally ignored because they do not receive solar radiation in
        # this simplified model.
        if not bool(b.get("is_sun_striked_outside", False)):
            continue
        if direction not in _NEU_INDEX_BY_DIRECTION:
            continue
        idx = _NEU_INDEX_BY_DIRECTION[direction]
        if btype == "external_transparent_part":
            eta = float(b.get("eta_value", 0.0))
            m_c += area * eta * neu_c[idx] * 0.93
            m_h += area * eta * neu_h[idx] * 0.51
        elif btype in ["external_general_part", "external_opaque_part"]:
            # 0.034 is the solar absorption conversion coefficient used in the
            # revised simple-input calculation.
            m_c += area * u_value * temp_dif_coef * neu_c[idx] * 0.034
            m_h += area * u_value * temp_dif_coef * neu_h[idx] * 0.034

    DD_H, DD_C = get_master_days(region)
    eta_ac_model = 100.0 * m_c / A_env_model if A_env_model > 0.0 else 0.0
    eta_ah_model = 100.0 * m_h / A_env_model if A_env_model > 0.0 else 0.0
    eta_avg_model = (eta_ac_model * DD_C + eta_ah_model * DD_H) / max(DD_C + DD_H, 1)

    return {
        "total_floor_area": A_total,
        "main_floor_area": A_MR,
        "other_floor_area": A_OR,
        "non_occupant_floor_area": A_NR,
        "A_env": A_env_model,
        "UA": heat_loss / A_env_model if A_env_model > 0.0 else 0.0,
        "eta_ac": eta_ac_model,
        "eta_ah": eta_ah_model,
        "eta_avg": eta_avg_model,
        "heat_loss_coefficient": heat_loss,
    }


def verify_estimated_characteristics(model: Dict[str, Any], *, region: int,
                                     total_floor_area: float, main_floor_area: float,
                                     other_floor_area: float, A_env: float, ua: float,
                                     eta_ah: float, eta_ac: float,
                                     abs_tol: float = 1.0e-2,
                                     rel_tol: float = 1.0e-2) -> Dict[str, Any]:
    """Compare simple inputs with characteristics recalculated from the model.

    The check report is JSON-serializable and is intended to be stored in the
    generated dictionary under ``_simple_input_verification``.
    """
    characteristics = calculate_model_characteristics(model, region)
    DD_H, DD_C = get_master_days(region)
    eta_avg_input = (float(eta_ac) * DD_C + float(eta_ah) * DD_H) / max(DD_C + DD_H, 1)
    non_occupant_floor_area = float(total_floor_area) - float(main_floor_area) - float(other_floor_area)

    items = {
        "total_floor_area": _make_check_item(total_floor_area, characteristics["total_floor_area"], unit="m2", abs_tol=abs_tol, rel_tol=rel_tol),
        "main_floor_area": _make_check_item(main_floor_area, characteristics["main_floor_area"], unit="m2", abs_tol=abs_tol, rel_tol=rel_tol),
        "other_floor_area": _make_check_item(other_floor_area, characteristics["other_floor_area"], unit="m2", abs_tol=abs_tol, rel_tol=rel_tol),
        "non_occupant_floor_area": _make_check_item(non_occupant_floor_area, characteristics["non_occupant_floor_area"], unit="m2", abs_tol=abs_tol, rel_tol=rel_tol),
        "A_env": _make_check_item(A_env, characteristics["A_env"], unit="m2", abs_tol=abs_tol, rel_tol=rel_tol),
        "UA": _make_check_item(ua, characteristics["UA"], unit="W/m2K", abs_tol=abs_tol, rel_tol=rel_tol),
        "eta_ac": _make_check_item(eta_ac, characteristics["eta_ac"], unit="-", abs_tol=abs_tol, rel_tol=rel_tol),
        "eta_ah": _make_check_item(eta_ah, characteristics["eta_ah"], unit="-", abs_tol=abs_tol, rel_tol=rel_tol),
        "eta_avg": _make_check_item(eta_avg_input, characteristics["eta_avg"], unit="-", abs_tol=abs_tol, rel_tol=rel_tol),
    }
    return {
        "ok": all(item["ok"] for item in items.values()),
        "abs_tol": abs_tol,
        "rel_tol": rel_tol,
        "items": items,
        "model_characteristics": {k: _r2(v) for k, v in characteristics.items()},
        "note": "UA and eta values are recalculated from the generated Dictionary/JSON. Internal partitions are not included in A_env or UA.",
    }


def print_verification_report(verification: Dict[str, Any]) -> None:
    """Pretty-print the verification report created by verify_estimated_characteristics."""
    print("入力値と暖冷房負荷モデルの特性値の照合")
    print("-------------------------------------------------")
    for name, item in verification.get("items", {}).items():
        status = "OK" if item.get("ok") else "NG"
        unit = item.get("unit", "")
        print(
            f"{name:24s}: input={item['input']:12.6g}  "
            f"model={item['model']:12.6g}  diff={item['difference']:12.6g}  {unit}  [{status}]"
        )
    print("overall:", "OK" if verification.get("ok") else "NG")


def _add_external_general(boundaries: list[dict], *, room: str, name: str, area: float,
                          direction: str, u_value: float, temp_dif_coef: float,
                          sun: Optional[bool] = None, layer_name: Optional[str] = None) -> None:
    if area <= 1.0e-8:
        return
    d = _DIRECTION_TO_HC[direction]
    if sun is None:
        sun = (temp_dif_coef > 0.0 and d in ["s", "e", "n", "w", "top"])
    boundaries.append({
        "id": len(boundaries),
        "name": name,
        "sub_name": "",
        "connected_room_id": _ROOM_DEFS[room]["id"],
        "boundary_type": "external_general_part",
        "area": _r2(area),
        "is_sun_striked_outside": bool(sun),
        "temp_dif_coef": _r2(temp_dif_coef),
        "is_solar_absorbed_inside": _hc_is_floor(d),
        "is_floor": _hc_is_floor(d),
        "direction": d,
        "h_c": _hc_h_c(d),
        "outside_emissivity": 0.9,
        "outside_heat_transfer_resistance": _hc_outside_heat_transfer_resistance(d, temp_dif_coef),
        "outside_solar_absorption": 0.8,
        "layers": _layers_from_u(layer_name or name, u_value, d, temp_dif_coef),
        "solar_shading_part": {"existence": False},
    })


def _add_window(boundaries: list[dict], *, room: str, name: str, area: float, direction: str,
                u_value: float, eta_value: float) -> None:
    if area <= 1.0e-8:
        return
    d = _DIRECTION_TO_HC[direction]
    boundaries.append({
        "id": len(boundaries),
        "name": name,
        "sub_name": "",
        "connected_room_id": _ROOM_DEFS[room]["id"],
        "boundary_type": "external_transparent_part",
        "area": _r2(area),
        "is_sun_striked_outside": True,
        "temp_dif_coef": 1.0,
        "is_solar_absorbed_inside": False,
        "is_floor": False,
        "direction": d,
        "h_c": _hc_h_c(d),
        "outside_emissivity": 0.9,
        "outside_heat_transfer_resistance": _hc_outside_heat_transfer_resistance(d, 1.0),
        "u_value": _r2(max(u_value, 0.1)),
        "inside_heat_transfer_resistance": 0.11,
        "eta_value": _r2(max(eta_value, 1.0e-8)),
        "glass_area_ratio": 0.8,
        "incident_angle_characteristics": "multiple",
        "solar_shading_part": {"existence": False},
    })


def _add_door(boundaries: list[dict], *, room: str, name: str, area: float, direction: str,
              u_value: float) -> None:
    if area <= 1.0e-8:
        return
    d = _DIRECTION_TO_HC[direction]
    boundaries.append({
        "id": len(boundaries),
        "name": name,
        "sub_name": "",
        "connected_room_id": _ROOM_DEFS[room]["id"],
        "boundary_type": "external_opaque_part",
        "area": _r2(area),
        "is_sun_striked_outside": True,
        "temp_dif_coef": 1.0,
        "is_solar_absorbed_inside": False,
        "is_floor": False,
        "direction": d,
        "h_c": _hc_h_c(d),
        "outside_emissivity": 0.9,
        "outside_heat_transfer_resistance": _hc_outside_heat_transfer_resistance(d, 1.0),
        "u_value": _r2(max(u_value, 0.1)),
        "inside_heat_transfer_resistance": 0.11,
        "outside_solar_absorption": 0.8,
        "solar_shading_part": {"existence": False},
    })


def _add_internal_pair(boundaries: list[dict], *, room_a: str, room_b: str, name: str,
                       area: float, direction: str = "horizontal") -> None:
    if area <= 1.0e-8 or room_a == room_b:
        return
    if direction == "upward":
        h_a, h_b = 5.0, 0.7
        floor_a, floor_b = False, True
    elif direction == "downward":
        h_a, h_b = 0.7, 5.0
        floor_a, floor_b = True, False
    else:
        h_a = h_b = 2.5
        floor_a = floor_b = False
    id_a = len(boundaries)
    id_b = id_a + 1
    layers = _partition_layers(name)
    boundaries.append({
        "id": id_a,
        "name": f"{name}_{room_a}_side",
        "sub_name": "",
        "connected_room_id": _ROOM_DEFS[room_a]["id"],
        "boundary_type": "internal",
        "area": _r2(area),
        "rear_surface_boundary_id": id_b,
        "is_solar_absorbed_inside": floor_a,
        "is_floor": floor_a,
        "h_c": h_a,
        "layers": layers,
        "solar_shading_part": {"existence": False},
    })
    boundaries.append({
        "id": id_b,
        "name": f"{name}_{room_b}_side",
        "sub_name": "",
        "connected_room_id": _ROOM_DEFS[room_b]["id"],
        "boundary_type": "internal",
        "area": _r2(area),
        "rear_surface_boundary_id": id_a,
        "is_solar_absorbed_inside": floor_b,
        "is_floor": floor_b,
        "h_c": h_b,
        "layers": list(reversed(layers)),
        "solar_shading_part": {"existence": False},
    })


def _add_ground(boundaries: list[dict], *, room: str, name: str, area: float) -> None:
    if area <= 1.0e-8:
        return
    boundaries.append({
        "id": len(boundaries),
        "name": name,
        "sub_name": "",
        "connected_room_id": _ROOM_DEFS[room]["id"],
        "boundary_type": "ground",
        "area": _r2(area),
        "is_solar_absorbed_inside": True,
        "is_floor": True,
        "h_c": 0.7,
        "layers": [_layer("ground_slab", 0.15, 165000.0)],
        "solar_shading_part": {"existence": False},
    })


def _estimate_area_and_properties(region: int, total_floor_area: float, main_floor_area: float,
                                  other_floor_area: float, A_env: float, ua: float,
                                  eta_ah: float, eta_ac: float, tatekata: str,
                                  structure: str, has_vertical_internal: str = "有") -> Dict[str, Any]:
    """Estimate intermediate areas/properties from the revised simple-input spec."""
    if tatekata not in ["戸建住宅", "共同住宅"]:
        raise ValueError("tatekata must be '戸建住宅' or '共同住宅'.")
    if region < 1 or region > 8:
        raise ValueError("region must be 1..8.")
    A_MR = float(main_floor_area)
    A_OR = float(other_floor_area)
    A_NR = float(total_floor_area) - A_MR - A_OR
    if min(A_MR, A_OR, A_NR) <= 0.0:
        raise ValueError("床面積の合計は、主たる居室＋その他居室より大きくしてください。")

    A_MR_ref, A_OR_ref, A_NR_ref = get_floor_area_ref(tatekata)
    area_ref = get_area_table_ref(tatekata)
    # rows: top, north, east, south, west, bottom, win_n/e/s/w, door_n/w
    ref = {
        "top": dict(zip(["MR", "OR", "NR", "UF"], area_ref[0])),
        "north": dict(zip(["MR", "OR", "NR", "UF"], area_ref[1])),
        "east": dict(zip(["MR", "OR", "NR", "UF"], area_ref[2])),
        "south": dict(zip(["MR", "OR", "NR", "UF"], area_ref[3])),
        "west": dict(zip(["MR", "OR", "NR", "UF"], area_ref[4])),
        "bottom": dict(zip(["MR", "OR", "NR", "UF"], area_ref[5])),
        "win_north": dict(zip(["MR", "OR", "NR"], area_ref[6])),
        "win_east": dict(zip(["MR", "OR", "NR"], area_ref[7])),
        "win_south": dict(zip(["MR", "OR", "NR"], area_ref[8])),
        "win_west": dict(zip(["MR", "OR", "NR"], area_ref[9])),
        "door_north": dict(zip(["MR", "OR", "NR"], area_ref[10])),
        "door_west": dict(zip(["MR", "OR", "NR"], area_ref[11])),
    }
    if tatekata == "戸建住宅":
        if structure == "基礎断熱":
            for r in ["MR", "OR", "NR"]:
                ref["bottom"][r] = 0.0
        elif structure in ["床断熱", "床下断熱"]:
            for d in ["north", "east", "south", "west", "bottom"]:
                ref[d]["UF"] = 0.0
        else:
            raise ValueError("structure must be '床断熱', '床下断熱' or '基礎断熱'.")
    else:
        structure = "共同住宅"

    floors = {"MR": A_MR, "OR": A_OR, "NR": A_NR}
    floors_ref = {"MR": A_MR_ref, "OR": A_OR_ref, "NR": A_NR_ref}
    # horizontal envelope areas
    top = {r: ref["top"][r] * floors[r] / floors_ref[r] for r in ["MR", "OR", "NR"]}
    bottom = {r: ref["bottom"][r] * floors[r] / floors_ref[r] for r in ["MR", "OR", "NR"]}
    bottom["UF"] = ref["bottom"]["UF"] * sum(floors.values()) / sum(floors_ref.values()) if ref["bottom"]["UF"] > 0 else 0.0
    top["UF"] = 0.0

    ref_vert_uf = sum(ref[d]["UF"] for d in ["north", "east", "south", "west"])
    vert_uf = ref_vert_uf * _safe_div(bottom["UF"], ref["bottom"]["UF"]) if ref["bottom"]["UF"] > 0.0 else 0.0
    living_horizontal = sum(top[r] + bottom[r] for r in ["MR", "OR", "NR"])
    total_living_vertical = max(float(A_env) - living_horizontal - vert_uf - bottom["UF"], 0.0)
    vert = {r: total_living_vertical * floors[r] / sum(floors.values()) for r in ["MR", "OR", "NR"]}
    vert["UF"] = vert_uf

    dir_area = {d: {} for d in ["south", "east", "north", "west"]}
    for r in ["MR", "OR", "NR", "UF"]:
        ref_sum = sum(ref[d][r] for d in ["south", "east", "north", "west"])
        for d in ["south", "east", "north", "west"]:
            dir_area[d][r] = vert[r] * _safe_div(ref[d][r], ref_sum)

    # External-contact ratios
    if tatekata == "戸建住宅":
        ratio_ex = {"top": 1.0, "bottom": 1.0, "south": 1.0, "east": 1.0, "north": 1.0, "west": 1.0}
    else:
        min_ex = sum(dir_area[d][r] for d in ["south", "north"] for r in ["MR", "OR", "NR"]) + sum(bottom[r] for r in ["MR", "OR", "NR"])
        r_env_ex_min = _safe_div(min_ex, A_env)
        r_env_ex = min(max(r_env_ex_min, 1.0 / (1.0 + math.exp(-9.10907512 * (ua - 1.05204145)))), 1.0)
        A_env_ex = A_env * r_env_ex
        denom = sum(top.values()) + sum(dir_area[d][r] for d in ["east", "west"] for r in ["MR", "OR", "NR"])
        r_side = max(_safe_div(A_env_ex - min_ex, denom), 0.0)
        ratio_ex = {"top": r_side, "bottom": 1.0, "south": 1.0, "north": 1.0, "east": r_side, "west": r_side}

    top_ex = {r: top[r] * ratio_ex["top"] for r in top}
    bottom_ex = {r: bottom[r] * ratio_ex["bottom"] for r in bottom}
    dir_ex = {d: {r: dir_area[d][r] * ratio_ex[d] for r in dir_area[d]} for d in dir_area}
    top_in = {r: top[r] - top_ex[r] for r in top}
    bottom_in = {r: bottom[r] - bottom_ex[r] for r in bottom}
    dir_in = {d: {r: dir_area[d][r] - dir_ex[d][r] for r in dir_area[d]} for d in dir_area}

    total_ex_area = sum(top_ex.values()) + sum(bottom_ex.values()) + sum(dir_ex[d][r] for d in dir_ex for r in dir_ex[d])
    total_open = total_ex_area * get_open_rate(float(eta_ac))

    ref_open = {}
    for r in ["MR", "OR", "NR"]:
        ref_open[r] = sum(ref[f"win_{d}"][r] for d in ["south", "east", "north", "west"]) + ref["door_north"][r] + ref["door_west"][r]
    denom_open = sum(floors[r] * _safe_div(ref_open[r], floors_ref[r]) for r in ["MR", "OR", "NR"])
    open_room = {r: total_open * floors[r] * _safe_div(ref_open[r], floors_ref[r]) / denom_open if denom_open > 0 else 0.0 for r in ["MR", "OR", "NR"]}

    win = {d: {} for d in ["south", "east", "north", "west"]}
    door = {"north": {}, "west": {}}
    for r in ["MR", "OR", "NR"]:
        for d in ["south", "east", "north", "west"]:
            win[d][r] = min(open_room[r] * _safe_div(ref[f"win_{d}"][r], ref_open[r]), dir_ex[d][r])
        for d in ["north", "west"]:
            door[d][r] = min(open_room[r] * _safe_div(ref[f"door_{d}"][r], ref_open[r]), max(dir_ex[d][r] - win[d][r], 0.0))

    wall_ex = {d: {} for d in ["south", "east", "north", "west"]}
    for d in wall_ex:
        for r in ["MR", "OR", "NR"]:
            wall_ex[d][r] = max(dir_ex[d][r] - win[d][r] - (door[d][r] if d in door else 0.0), 0.0)
        wall_ex[d]["UF"] = dir_ex[d].get("UF", 0.0)

    A_roof_ex = sum(top_ex.values())
    A_floor_ex = sum(bottom_ex[r] for r in ["MR", "OR", "NR"])
    A_base_ex = sum(wall_ex[d].get("UF", 0.0) for d in wall_ex)
    A_wall_ex = sum(wall_ex[d][r] for d in wall_ex for r in ["MR", "OR", "NR"])
    A_win = sum(win[d][r] for d in win for r in ["MR", "OR", "NR"])
    A_door = sum(door[d][r] for d in door for r in ["MR", "OR", "NR"])

    # Allocate UA to parts with a common multiplier, avoiding impossible zero/negative U-values.
    u_ref = {"roof": 0.35, "wall": 0.53, "floor": 0.48, "base": 0.75, "win": 4.65, "door": 4.65}
    q_ref = (u_ref["roof"] * A_roof_ex + u_ref["wall"] * A_wall_ex + u_ref["floor"] * A_floor_ex +
             u_ref["base"] * A_base_ex + u_ref["win"] * A_win + u_ref["door"] * A_door)
    scale = _safe_div(float(ua) * float(A_env), q_ref, 1.0)
    u = {
        "roof": min(max(u_ref["roof"] * scale, 0.05), 5.0),
        "wall": min(max(u_ref["wall"] * scale, 0.05), 5.0),
        "floor": min(max(u_ref["floor"] * scale, 0.05), 5.0),
        "base": min(max(u_ref["base"] * scale, 0.05), 5.0),
        "win": min(max(u_ref["win"] * scale, 0.5), 10.0),
        "door": min(max(u_ref["door"] * scale, 0.5), 10.0),
    }

    DD_H, DD_C = get_master_days(region)
    neu_c, neu_h = get_neu_avg(region)
    eta_avg = (float(eta_ac) * DD_C + float(eta_ah) * DD_H) / max(DD_C + DD_H, 1)
    m_total = eta_avg / 100.0 * float(A_env)
    m_opaq_c = (A_roof_ex * u["roof"] * neu_c[0] +
                sum(wall_ex["south"][r] for r in wall_ex["south"]) * u["wall"] * neu_c[5] +
                sum(wall_ex["east"][r] for r in wall_ex["east"]) * u["wall"] * neu_c[3] +
                sum(wall_ex["north"][r] for r in wall_ex["north"]) * u["wall"] * neu_c[1] +
                sum(wall_ex["west"][r] for r in wall_ex["west"]) * u["wall"] * neu_c[7] +
                sum(door["north"].values()) * u["door"] * neu_c[1] +
                sum(door["west"].values()) * u["door"] * neu_c[7]) * 0.034
    m_opaq_h = (A_roof_ex * u["roof"] * neu_h[0] +
                sum(wall_ex["south"][r] for r in wall_ex["south"]) * u["wall"] * neu_h[5] +
                sum(wall_ex["east"][r] for r in wall_ex["east"]) * u["wall"] * neu_h[3] +
                sum(wall_ex["north"][r] for r in wall_ex["north"]) * u["wall"] * neu_h[1] +
                sum(wall_ex["west"][r] for r in wall_ex["west"]) * u["wall"] * neu_h[7] +
                sum(door["north"].values()) * u["door"] * neu_h[1] +
                sum(door["west"].values()) * u["door"] * neu_h[7]) * 0.034
    m_opaq = (m_opaq_c * DD_C + m_opaq_h * DD_H) / max(DD_C + DD_H, 1)
    eta_win = calc_eta_win(max(m_total - m_opaq, 1.0e-8),
                           (sum(win["south"].values()), sum(win["east"].values()), sum(win["north"].values()), sum(win["west"].values())),
                           neu_c, neu_h, DD_C, DD_H) if A_win > 0 else 1.0e-8
    eta_win = min(max(eta_win, 1.0e-8), 0.88)

    part = dict(zip(["MR_OR", "MR_NR", "OR_NR"], get_partition_table_ref(tatekata)))
    ref_vert = {r: sum(ref[d][r] for d in ["south", "east", "north", "west"]) for r in ["MR", "OR", "NR"]}
    if has_vertical_internal == "無":
        part_area = {"MR_OR": 0.0, "MR_NR": 0.0, "OR_NR": 0.0}
    else:
        part_area = {
            "MR_OR": part["MR_OR"] * _safe_div(vert["MR"] + vert["OR"], ref_vert["MR"] + ref_vert["OR"]),
            "MR_NR": part["MR_NR"] * _safe_div(vert["MR"] + vert["NR"], ref_vert["MR"] + ref_vert["NR"]),
            "OR_NR": part["OR_NR"] * _safe_div(vert["OR"] + vert["NR"], ref_vert["OR"] + ref_vert["NR"]),
        }

    pb_names = ["MR_MR", "MR_OR", "MR_NR", "MR_UF", "OR_MR", "OR_OR", "OR_NR", "OR_UF", "NR_MR", "NR_OR", "NR_NR", "NR_UF"]
    pb_ref = dict(zip(pb_names, get_partition_bottom_table_ref(tatekata)))
    if not (tatekata == "戸建住宅" and structure == "基礎断熱"):
        for k in ["MR_UF", "OR_UF", "NR_UF"]:
            pb_ref[k] = 0.0
    pb_total = {"MR": max(A_MR - bottom["MR"], 0.0), "OR": max(A_OR - bottom["OR"], 0.0), "NR": max(A_NR - bottom["NR"], 0.0)}
    pb = {}
    for src in ["MR", "OR", "NR"]:
        keys = [k for k in pb_names if k.startswith(src + "_")]
        denom = sum(pb_ref[k] for k in keys)
        for k in keys:
            pb[k] = pb_total[src] * _safe_div(pb_ref[k], denom)

    return {
        "floor": {"MR": A_MR, "OR": A_OR, "NR": A_NR, "UF": bottom["UF"]},
        "volume": {"MR": 2.4 * A_MR, "OR": 2.4 * A_OR, "NR": 2.4 * A_NR, "UF": 0.4 * bottom["UF"]},
        "top_ex": top_ex, "top_in": top_in, "bottom_ex": bottom_ex, "bottom_in": bottom_in,
        "wall_ex": wall_ex, "wall_in": dir_in, "win": win, "door": door,
        "u": u, "eta_win": eta_win, "part": part_area, "part_bottom": pb,
        "summary": {
            "A_env_ex": total_ex_area,
            "A_open": total_open,
            "A_wall_ex": A_wall_ex,
            "A_roof_ex": A_roof_ex,
            "A_floor_ex": A_floor_ex,
            "A_win": A_win,
            "A_door": A_door,
            "eta_win": eta_win,
        },
    }


def create_heat_load_calc_input(region: int, total_floor_area: float, main_floor_area: float,
                                other_floor_area: float, A_env: float, ua: float,
                                eta_ah: float, eta_ac: float, tatekata: str,
                                structure: str = "床断熱", has_vertical_internal: str = "有",
                                ac_method: str = "ot", interval: str = "1h",
                                c_value: float = 2.0, inside_pressure: str = "negative",
                                include_verification: bool = True,
                                verification_abs_tol: float = 1.0e-2,
                                verification_rel_tol: float = 1.0e-2) -> Dict[str, Any]:
    """Create a Heat Load Calc input dictionary directly from simple inputs.

    Parameters match the revised simple-input specification: building type,
    region, floor areas by room use, total envelope area, UA, eta_AC, eta_AH,
    and the floor/foundation insulation distinction for detached houses.

    If ``include_verification`` is True, the returned dictionary also contains
    ``_simple_input_verification``, which compares the simple inputs with the
    characteristics recalculated from the generated model.
    """
    est = _estimate_area_and_properties(region, total_floor_area, main_floor_area, other_floor_area,
                                        A_env, ua, eta_ah, eta_ac, tatekata, structure,
                                        has_vertical_internal)
    rooms = []
    for r in ["MR", "OR", "NR"] + (["UF"] if est["floor"]["UF"] > 1.0e-8 else []):
        rooms.append({
            "id": _ROOM_DEFS[r]["id"],
            "name": _ROOM_DEFS[r]["name"],
            "sub_name": _ROOM_DEFS[r]["label"],
            "floor_area": _r2(est["floor"][r]),
            "volume": _r2(est["volume"][r]),
            "ventilation": {"natural": 0.0},
            "furniture": {"input_method": "default"},
            "schedule": {"name": _ROOM_DEFS[r]["schedule"]},
        })

    boundaries: list[dict] = []
    # External/in-contact general parts
    for r in ["MR", "OR", "NR"]:
        _add_external_general(boundaries, room=r, name=f"{r}_roof_external", area=est["top_ex"][r], direction="top", u_value=est["u"]["roof"], temp_dif_coef=1.0, layer_name="roof_external")
        _add_external_general(boundaries, room=r, name=f"{r}_roof_parting", area=est["top_in"][r], direction="top", u_value=0.0, temp_dif_coef=0.0, sun=False, layer_name="roof_parting")
        if est["floor"]["UF"] <= 1.0e-8:
            _add_external_general(boundaries, room=r, name=f"{r}_floor_external", area=est["bottom_ex"][r], direction="bottom", u_value=est["u"]["floor"], temp_dif_coef=0.7, sun=False, layer_name="floor_external")
        _add_external_general(boundaries, room=r, name=f"{r}_floor_parting", area=est["bottom_in"][r], direction="bottom", u_value=0.0, temp_dif_coef=0.0, sun=False, layer_name="floor_parting")
        for d in ["south", "east", "north", "west"]:
            _add_external_general(boundaries, room=r, name=f"{r}_wall_{d}_external", area=est["wall_ex"][d][r], direction=d, u_value=est["u"]["wall"], temp_dif_coef=1.0, layer_name="wall_external")
            _add_external_general(boundaries, room=r, name=f"{r}_wall_{d}_parting", area=est["wall_in"][d][r], direction=d, u_value=0.0, temp_dif_coef=0.0, sun=False, layer_name="wall_parting")
            _add_window(boundaries, room=r, name=f"{r}_window_{d}", area=est["win"][d][r], direction=d, u_value=est["u"]["win"], eta_value=est["eta_win"])
        for d in ["north", "west"]:
            _add_door(boundaries, room=r, name=f"{r}_door_{d}", area=est["door"][d][r], direction=d, u_value=est["u"]["door"])

    if est["floor"]["UF"] > 1.0e-8:
        for d in ["south", "east", "north", "west"]:
            _add_external_general(boundaries, room="UF", name=f"UF_base_{d}_external", area=est["wall_ex"][d].get("UF", 0.0), direction=d, u_value=est["u"]["base"], temp_dif_coef=1.0, layer_name="base_external")
        _add_ground(boundaries, room="UF", name="UF_ground", area=est["floor"]["UF"])
        for r in ["MR", "OR", "NR"]:
            _add_internal_pair(boundaries, room_a=r, room_b="UF", name=f"bottom_{r}_UF", area=est["part_bottom"].get(f"{r}_UF", 0.0), direction="downward")

    # Same-use inner floors are represented as adiabatic external floors per the revised spec.
    for r in ["MR", "OR", "NR"]:
        _add_external_general(boundaries, room=r, name=f"{r}_same_use_bottom_adiabatic", area=est["part_bottom"].get(f"{r}_{r}", 0.0), direction="bottom", u_value=0.0, temp_dif_coef=0.0, sun=False, layer_name="same_use_bottom")

    # Vertical internal partitions.
    _add_internal_pair(boundaries, room_a="MR", room_b="OR", name="partition_MR_OR", area=est["part"].get("MR_OR", 0.0))
    _add_internal_pair(boundaries, room_a="MR", room_b="NR", name="partition_MR_NR", area=est["part"].get("MR_NR", 0.0))
    _add_internal_pair(boundaries, room_a="OR", room_b="NR", name="partition_OR_NR", area=est["part"].get("OR_NR", 0.0))

    # Different-use inner floors.  Direction is kept as a floor/ceiling pair.
    for a, b in [("MR", "OR"), ("MR", "NR"), ("OR", "NR")]:
        area_ab = est["part_bottom"].get(f"{a}_{b}", 0.0)
        area_ba = est["part_bottom"].get(f"{b}_{a}", 0.0)
        _add_internal_pair(boundaries, room_a=a, room_b=b, name=f"bottom_{a}_{b}", area=area_ab, direction="downward")
        _add_internal_pair(boundaries, room_a=b, room_b=a, name=f"bottom_{b}_{a}", area=area_ba, direction="downward")

    V_MR = est["volume"]["MR"]
    V_OR = est["volume"]["OR"]
    V_NR = est["volume"]["NR"]
    ventilation_rate = 0.5
    v_mr = ventilation_rate * (V_MR + V_NR * V_MR / (V_MR + V_OR))
    v_or = ventilation_rate * (V_OR + V_NR * V_OR / (V_MR + V_OR))
    mechanical_ventilations = [
        {"id": 0, "root_type": "type3", "volume": _r2(v_mr), "route": [0, 2]},
        {"id": 1, "root_type": "type3", "volume": _r2(v_or), "route": [1, 2]},
    ]

    c0, h0 = _rac_equipments(0, 0, est["floor"]["MR"])
    c1, h1 = _rac_equipments(1, 1, est["floor"]["OR"])
    result = {
        "common": {
            "calculation_day": {"main": 365, "run_up": 365, "run_up_building": 183},
            "interval": interval,
            "ac_method": ac_method,
            "ac_config": [
                {"mode": 1, "lower": 20.0, "upper": 27.0},
            ],
            "weather": {"method": "ees", "region": int(region)},
            "mutual_radiation_method": "Nagata",
        },
        "building": {
            "infiltration": {
                "method": "balance_residential",
                "c_value_estimate": "specify",
                "story": 2 if tatekata == "戸建住宅" else 1,
                "c_value": float(c_value),
                "inside_pressure": inside_pressure,
            }
        },
        "rooms": rooms,
        "boundaries": boundaries,
        "mechanical_ventilations": mechanical_ventilations,
        "equipments": {"heating_equipments": [h0, h1], "cooling_equipments": [c0, c1]},
        "_simple_input_summary": {k: _r2(v) for k, v in est["summary"].items()},
    }
    if include_verification:
        result["_simple_input_verification"] = verify_estimated_characteristics(
            result,
            region=region,
            total_floor_area=total_floor_area,
            main_floor_area=main_floor_area,
            other_floor_area=other_floor_area,
            A_env=A_env,
            ua=ua,
            eta_ah=eta_ah,
            eta_ac=eta_ac,
            abs_tol=verification_abs_tol,
            rel_tol=verification_rel_tol,
        )
    return result


def estimate(region: int, total_floor_area: float, main_floor_area: float, other_floor_area: float,
             A_env: float, ua: float, eta_ah: float, eta_ac: float, tatekata: str,
             structure: str = "床断熱", xlsx_path: Optional[str] = None,
             has_vertical_internal: str = "有", json_path: Optional[str] = None,
             **kwargs) -> Dict[str, Any]:
    """Backward-compatible wrapper that returns Heat Load Calc JSON dict.

    `xlsx_path` is retained only for compatibility.  If `xlsx_path` or
    `json_path` ends with `.json`, the dictionary is also written as UTF-8 JSON.
    No Excel workbook is created by this function.
    """
    data = create_heat_load_calc_input(
        region=region,
        total_floor_area=total_floor_area,
        main_floor_area=main_floor_area,
        other_floor_area=other_floor_area,
        A_env=A_env,
        ua=ua,
        eta_ah=eta_ah,
        eta_ac=eta_ac,
        tatekata=tatekata,
        structure=structure,
        has_vertical_internal=has_vertical_internal,
        **kwargs,
    )
    out = json_path or (xlsx_path if xlsx_path and str(xlsx_path).lower().endswith(".json") else None)
    if out:
        Path(out).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def write_heat_load_calc_input_json(path: str | Path, **estimate_kwargs) -> Dict[str, Any]:
    """Create and write Heat Load Calc input JSON, returning the same dict."""
    data = create_heat_load_calc_input(**estimate_kwargs)
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


if __name__ == "__main__":
    sample = estimate(
        region=3,
        total_floor_area=83.38,
        main_floor_area=29.225,
        other_floor_area=34.47,
        A_env=264.12,
        ua=1.991226,
        eta_ah=2,
        eta_ac=1,
        tatekata="戸建住宅",
        structure="基礎断熱",
        json_path="test.json",
    )
    if "_simple_input_verification" in sample:
        print_verification_report(sample["_simple_input_verification"])
    print(json.dumps(sample, ensure_ascii=False, indent=2))
