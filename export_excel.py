# -*- coding: utf-8 -*-
"""
export_excel.py
==============================================================================
1つの暖冷房負荷モデルにつき1つの Excel を出力する。
シート構成（一覧性重視）:
  1. 検証          入力値 vs モデル再現値と誤差%（=式で算出）
  2. 面積一覧      空間×方位の外皮/外気接/窓/ドア/外壁等/非外気
  3. 熱貫流率      部位別 U値・断熱材厚・基礎ψ
  4. 室・換気      室容積・自然風換気量
  5. 境界一覧      HLC boundary を平坦化（行=境界）
==============================================================================
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import reference_data as ref
import verify as vfy

VERT = ref.VERT_DIRS
DIRS = ["top", "north", "east", "south", "west", "bottom"]
DIR_JP = {"top": "上面", "north": "北", "east": "東", "south": "南",
          "west": "西", "bottom": "下面"}
SPACE_JP = {"MR": "主たる居室", "OR": "その他居室", "NR": "非居室", "UF": "床下空間"}

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(name=FONT, bold=True, size=13, color="1F4E78")
NG_FILL = PatternFill("solid", fgColor="FFC7CE")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncol, start=1):
    for c in range(start, start + ncol):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _base_font(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.font is None or cell.font.name != FONT:
                if not cell.font.bold and cell.font.color is None:
                    cell.font = Font(name=FONT, size=10)


def export(m, results, path, model_label):
    wb = Workbook()

    # ===== Sheet1 検証 =====
    ws = wb.active
    ws.title = "検証"
    ws["A1"] = f"暖冷房負荷モデル 推定結果 ─ {model_label}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "【入力パラメータ（設計住戸）】"
    ws["A2"].font = Font(name=FONT, bold=True, size=11)
    inp = [
        ("建て方", "戸建住宅" if m.bt == "detached" else "集合住宅"),
        ("地域の区分", m.region),
        ("断熱方式", "床断熱" if m.is_floor_ins else "基礎断熱"),
        ("主たる居室床面積 A_MR [m2]", m.A["MR"]),
        ("その他居室床面積 A_OR [m2]", m.A["OR"]),
        ("床面積の合計 A_A [m2]", m.A_A),
        ("総外皮面積 A_env [m2]", m.A_env_input),
        ("外皮平均熱貫流率 UA [W/m2K]", m.UA),
        ("冷房期平均日射熱取得率 ηAC", m.eta_AC),
        ("暖房期平均日射熱取得率 ηAH", m.eta_AH),
    ]
    r = 3
    for k, v in inp:
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, size=10)
        c = ws.cell(row=r, column=2, value=v)
        c.font = Font(name=FONT, size=10, color="0000FF")
        c.alignment = Alignment(horizontal="right")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="【再現性検証】入力値が暖冷房負荷モデルで再現できているか")
    ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=11)
    r += 1
    hdr = ["検証項目", "目標(入力)", "モデル再現値", "誤差 [%]", "判定"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, len(hdr))
    r += 1
    keys = [("A_A", "床面積合計 A_A [m2]"), ("A_MR", "主たる居室 A_MR [m2]"),
            ("A_OR", "その他居室 A_OR [m2]"), ("A_env", "総外皮面積 A_env [m2]"),
            ("UA", "外皮平均熱貫流率 UA [W/m2K]"), ("etaA", "平均日射熱取得率 ηA [-]")]
    first = r
    for key, label in keys:
        res = results[key]
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=round(res["target"], 4))
        ws.cell(row=r, column=3, value=round(res["model"], 4))
        # 誤差は式で算出（動的）
        ws.cell(row=r, column=4,
                value=f'=IF(B{r}=0,0,(C{r}-B{r})/B{r}*100)')
        ws.cell(row=r, column=4).number_format = "0.000"
        jcell = ws.cell(row=r, column=5, value=f'=IF(ABS(D{r})<=0.5,"OK","要確認")')
        for cc in range(1, 6):
            ws.cell(row=r, column=cc).border = BORDER
        if not res["ok"]:
            for cc in range(1, 6):
                ws.cell(row=r, column=cc).fill = NG_FILL
        else:
            jcell.fill = OK_FILL
        r += 1

    if results["_window_clipped"]:
        r += 1
        note = (f"※ 窓の垂直入射時日射熱取得率 暫定値={results['_eta_win_temp']:.4f} が "
                f"上下限[0.10, 0.73]を外れ、η={results['_eta_win']:.2f} にクリップ。"
                f"仕様3.4.9により、この場合は総外皮面積 A_env が再現されない。")
        ws.cell(row=r, column=1, value=note).font = Font(name=FONT, size=9, italic=True)
        ws.cell(row=r, column=1).fill = WARN_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    ws.column_dimensions["A"].width = 34
    for col in "BCDE":
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A1"

    # ===== Sheet2 面積一覧 =====
    ws2 = wb.create_sheet("面積一覧")
    ws2["A1"] = "面積一覧 [m2]（空間×方位）"
    ws2["A1"].font = TITLE_FONT
    head = ["空間", "方位", "外皮面積", "外気接外皮", "窓(補正後)", "ドア", "外壁等(外気接)", "非外気外皮"]
    hr = 3
    for i, h in enumerate(head, 1):
        ws2.cell(row=hr, column=i, value=h)
    _style_header(ws2, hr, len(head))
    rr = hr + 1
    data_start = rr
    for s in m.spaces:
        for d in DIRS:
            A_env = m._A_env_rd(s, d)
            A_ex = m.A_env_rd_ex[s][d]
            win = m.A_win_mod[s][d] if d in VERT else 0.0
            door = m.A_door[s][d] if d in VERT else 0.0
            wall = m.A_wall_ex[s][d]
            nonext = m.A_in[s][d]
            if max(A_env, A_ex, win, door, wall, nonext) <= 1e-9:
                continue
            vals = [SPACE_JP[s], DIR_JP[d], A_env, A_ex, win, door, wall, nonext]
            for i, v in enumerate(vals, 1):
                cell = ws2.cell(row=rr, column=i, value=(round(v, 3) if isinstance(v, float) else v))
                cell.border = BORDER
                if i >= 3:
                    cell.number_format = "0.000"
            rr += 1
    # 合計行（式）
    ws2.cell(row=rr, column=1, value="合計").font = Font(name=FONT, bold=True)
    for col in range(3, 9):
        L = get_column_letter(col)
        ws2.cell(row=rr, column=col, value=f"=SUM({L}{data_start}:{L}{rr-1})")
        ws2.cell(row=rr, column=col).number_format = "0.000"
        ws2.cell(row=rr, column=col).font = Font(name=FONT, bold=True)
        ws2.cell(row=rr, column=col).fill = SUB_FILL
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 8
    for col in "CDEFGH":
        ws2.column_dimensions[col].width = 14
    ws2.freeze_panes = "A4"

    # ===== Sheet3 熱貫流率 =====
    ws3 = wb.create_sheet("熱貫流率")
    ws3["A1"] = "部位別 熱貫流率・断熱材厚さ"
    ws3["A1"].font = TITLE_FONT
    head3 = ["部位", "熱貫流率 U [W/m2K]", "断熱材厚さ d [m]", "備考"]
    for i, h in enumerate(head3, 1):
        ws3.cell(row=3, column=i, value=h)
    _style_header(ws3, 3, len(head3))
    rows3 = [
        ("屋根・天井", m.U.get("roof", 0), m.d_ins.get("roof"), ""),
        ("外壁", m.U.get("wall", 0), m.d_ins.get("wall"), ""),
        ("窓", m.U.get("win", 0), None, f"η={m.eta_win:.3f}, ガラス面積率={m.glass_area_ratio}"),
        ("ドア", m.U.get("door", 0), None, ""),
    ]
    if m.has_uf:
        rows3.append(("床(対外気)", 0.0, None, "基礎断熱のため対外気床なし"))
        rows3.append(("基礎壁", m.U.get("uf_wall", 0), m.d_ins.get("uf_wall"), ""))
        rows3.append(("土間床外周部 ψ [W/mK]", m.psi_uf, None, f"外周長 L={m.L_uf_ex:.2f} m（線熱貫流, エンジン実装待ち）"))
    else:
        rows3.append(("床", m.U.get("floor", 0), m.d_ins.get("floor"), ""))
    rr = 4
    for name, u, d, note in rows3:
        ws3.cell(row=rr, column=1, value=name).border = BORDER
        cu = ws3.cell(row=rr, column=2, value=round(u, 4)); cu.number_format = "0.0000"; cu.border = BORDER
        cd = ws3.cell(row=rr, column=3, value=(round(d, 4) if d is not None else "-"))
        cd.number_format = "0.0000"; cd.border = BORDER
        ws3.cell(row=rr, column=4, value=note).border = BORDER
        rr += 1
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 44

    # ===== Sheet4 室・換気 =====
    ws4 = wb.create_sheet("室・換気")
    ws4["A1"] = "室容積・自然風換気量"
    ws4["A1"].font = TITLE_FONT
    head4 = ["室ID", "室名", "床面積 [m2]", "天井高/床下高 [m]", "室容積 [m3]", "換気回数 [回/h]", "自然風換気量 [m3/h]"]
    for i, h in enumerate(head4, 1):
        ws4.cell(row=3, column=i, value=h)
    _style_header(ws4, 3, len(head4))
    from hlc_builder import ROOM_ID
    rr = 4
    for s in m.spaces:
        h = 0.4 if s == "UF" else 2.4
        af = m.A_UF if s == "UF" else m.A[s]
        vals = [ROOM_ID[s], SPACE_JP[s], round(af, 3), h, round(m.V[s], 3),
                m.n_r.get(s, 0.0), round(m.Q_ntrl[s], 3)]
        for i, v in enumerate(vals, 1):
            ws4.cell(row=rr, column=i, value=v).border = BORDER
        rr += 1
    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 14
    for col in "CDEFG":
        ws4.column_dimensions[col].width = 18

    # ===== Sheet5 境界一覧 =====
    from hlc_builder import build as build_hlc
    hlc = build_hlc(m)
    ws5 = wb.create_sheet("境界一覧")
    ws5["A1"] = "境界一覧（heat_load_calc boundaries）"
    ws5["A1"].font = TITLE_FONT
    head5 = ["id", "name", "接室", "境界種別", "面積[m2]", "方位", "温度差係数",
             "U値[W/m2K]", "η値", "床", "裏面id"]
    for i, h in enumerate(head5, 1):
        ws5.cell(row=3, column=i, value=h)
    _style_header(ws5, 3, len(head5))
    rr = 4
    rid2name = {0: "MR", 1: "OR", 2: "NR", 3: "UF"}
    for b in hlc["boundaries"]:
        # 一般部位のU値は層から逆算（表示用）
        u = b.get("u_value")
        if u is None and "layers" in b:
            Rsum = sum(l["thermal_resistance"] for l in b["layers"]) + 0.04 + 0.11
            u = round(1.0 / Rsum, 4) if Rsum > 0 else None
        vals = [
            b["id"], b["name"], rid2name.get(b["connected_room_id"], b["connected_room_id"]),
            b["boundary_type"], round(b["area"], 3), b.get("direction", "-"),
            b.get("temp_dif_coef", "-"), (round(u, 4) if u else "-"),
            b.get("eta_value", "-"), "○" if b.get("is_floor") else "",
            b.get("rear_surface_boundary_id", "-"),
        ]
        for i, v in enumerate(vals, 1):
            cell = ws5.cell(row=rr, column=i, value=v)
            cell.border = BORDER
            cell.font = Font(name=FONT, size=9)
        rr += 1
    ws5.cell(row=rr, column=4, value="面積合計")
    ws5.cell(row=rr, column=4).font = Font(name=FONT, bold=True)
    ws5.cell(row=rr, column=5, value=f"=SUM(E4:E{rr-1})")
    ws5.cell(row=rr, column=5).number_format = "0.000"
    ws5.cell(row=rr, column=5).font = Font(name=FONT, bold=True)
    widths5 = [6, 34, 8, 24, 10, 8, 11, 12, 8, 6, 8]
    for i, w in enumerate(widths5, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.freeze_panes = "A4"

    # 既定フォント
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for cell in row:
                if cell.value is not None and (cell.font is None or cell.font.name != FONT):
                    f = cell.font
                    cell.font = Font(name=FONT, size=f.size or 10, bold=f.bold,
                                     color=f.color, italic=f.italic)

    wb.save(path)
    return path
